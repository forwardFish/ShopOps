from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from scripts.run_jushuitan_orders_to_feishu import FeishuOrderWriter, F_RAW
from shopops.config import load_settings


TEXT_FIELD = 1
NUMBER_FIELD = 2

F_UNIQUE_KEY = "unique_key"
F_PLATFORM = "\u5e73\u53f0"
F_DATA_SOURCE = "\u6570\u636e\u6765\u6e90"
F_SHOP_ID = "\u5e97\u94faID"
F_SHOP_NAME = "\u5e97\u94fa\u540d\u79f0"
F_FETCHED_AT = "\u91c7\u96c6\u65f6\u95f4"
F_ORDER_NO = "\u8ba2\u5355\u53f7"
F_CREATED_AT = "\u521b\u5efa\u65f6\u95f4"
F_BUYER_NICK = "\u4e70\u5bb6\u6635\u79f0"
F_PRODUCT = "\u5546\u54c1\u540d\u79f0"
F_UNIT_PRICE = "\u5355\u4ef7"
F_QUANTITY = "\u6570\u91cf"
F_FULFILL_STATUS = "\u5c65\u7ea6/\u552e\u540e\u72b6\u6001"
F_TRADE_STATUS = "\u4ea4\u6613\u72b6\u6001"
F_PAID_AMOUNT = "\u5b9e\u6536\u6b3e"
F_REFUND_AMOUNT = "\u9000\u6b3e\u91d1\u989d"
F_OPERATION = "\u64cd\u4f5c\u4fe1\u606f"

PDD_EXTRA_FIELDS = [
    ("\u62fc\u591a\u591a\u8ba2\u5355\u72b6\u6001", TEXT_FIELD),
    ("\u62fc\u591a\u591a\u552e\u540e\u72b6\u6001", TEXT_FIELD),
    ("\u62fc\u591a\u591a\u627f\u8bfa\u53d1\u8d27\u65f6\u95f4", TEXT_FIELD),
    ("\u62fc\u591a\u591a\u5feb\u9012\u5355\u53f7", TEXT_FIELD),
    ("\u62fc\u591a\u591a\u7269\u6d41\u516c\u53f8", TEXT_FIELD),
    ("\u62fc\u591a\u591a\u5546\u54c1ID", TEXT_FIELD),
    ("\u62fc\u591a\u591a\u5546\u54c1\u89c4\u683c", TEXT_FIELD),
    ("\u62fc\u591a\u591a\u5546\u5bb6\u7f16\u7801", TEXT_FIELD),
    ("\u62fc\u591a\u591a\u652f\u4ed8\u65b9\u5f0f", TEXT_FIELD),
    ("\u62fc\u591a\u591a\u7528\u6237\u5b9e\u4ed8\u91d1\u989d", NUMBER_FIELD),
    (F_REFUND_AMOUNT, NUMBER_FIELD),
]

REQUIRED_HEADER_ALIASES = {
    "order_no": ["\u8ba2\u5355\u53f7", "\u8ba2\u5355\u7f16\u53f7", "\u8ba2\u5355id", "\u8ba2\u5355ID"],
    "product": ["\u5546\u54c1", "\u5546\u54c1\u540d\u79f0", "\u5546\u54c1\u6807\u9898"],
    "paid_amount": ["\u5546\u5bb6\u5b9e\u6536\u91d1\u989d(\u5143)", "\u5546\u5bb6\u5b9e\u6536\u91d1\u989d", "\u8ba2\u5355\u91d1\u989d", "\u652f\u4ed8\u91d1\u989d", "\u5b9e\u6536\u6b3e"],
}

PDD_SIGNATURE_HEADERS = [
    "\u6536\u8d27\u4eba",
    "\u6536\u4ef6\u4eba",
    "\u6536\u8d27\u5730\u5740",
    "\u8be6\u7ec6\u5730\u5740",
    "\u627f\u8bfa\u53d1\u8d27\u65f6\u95f4",
    "\u5546\u5bb6\u5b9e\u6536\u91d1\u989d",
    "\u5feb\u9012\u5355\u53f7",
    "\u552e\u540e\u72b6\u6001",
    "\u7528\u6237\u5b9e\u4ed8\u91d1\u989d(\u5143)",
    "\u652f\u4ed8\u65f6\u95f4",
]

SENSITIVE_HEADER_PARTS = [
    "\u6536\u8d27",
    "\u6536\u4ef6",
    "\u624b\u673a",
    "\u7535\u8bdd",
    "\u5730\u5740",
    "\u6d88\u8d39\u8005\u8d44\u6599",
    "\u7528\u6237\u8d2d\u4e70",
    "\u914d\u9001\u5458",
    "\u5145\u503c\u53f7\u7801",
]


class PinduoduoOrderWriter(FeishuOrderWriter):
    def ensure_pdd_fields(self) -> list[str]:
        existing = self._field_names()
        created: list[str] = []
        for field_name, field_type in PDD_EXTRA_FIELDS:
            if field_name in existing:
                continue
            self._request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/fields",
                {"field_name": field_name, "type": field_type},
                allow_duplicate=True,
            )
            created.append(field_name)
            existing.add(field_name)
        return created


def load_source_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    if path.suffix.lower() == ".csv":
        return load_csv_rows(path)
    return load_excel_rows(path)


def load_csv_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        headers = [str(header or "").strip() for header in (reader.fieldnames or [])]
        rows: list[dict[str, Any]] = []
        for row in reader:
            normalized = {str(key or "").strip(): value for key, value in row.items()}
            if any(value not in (None, "") for value in normalized.values()):
                rows.append(normalized)
    return headers, rows


def load_excel_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return headers, rows


def diagnose_headers(headers: list[str]) -> dict[str, Any]:
    resolved = {key: resolve_header(headers, aliases) for key, aliases in REQUIRED_HEADER_ALIASES.items()}
    missing_required = [key for key, value in resolved.items() if not value]
    signature_hits = [header for header in headers if any(part in header for part in PDD_SIGNATURE_HEADERS)]
    incompatible_hits = [header for header in headers if header in {"\u4f5c\u8005\u8d26\u53f7", "\u6296\u97f3/\u706b\u5c71\u53f7", "\u4f63\u91d1\u7387"}]
    is_pdd_like = not missing_required and len(signature_hits) >= 2 and not incompatible_hits
    return {
        "is_pdd_like": is_pdd_like,
        "resolved": resolved,
        "missing_required": missing_required,
        "signature_hits": signature_hits,
        "incompatible_hits": incompatible_hits,
    }


def resolve_header(headers: list[str], aliases: list[str]) -> str | None:
    for alias in aliases:
        if alias in headers:
            return alias
    return None


def pdd_feishu_rows(headers: list[str], excel_rows: list[dict[str, Any]], source_file: Path) -> list[dict[str, Any]]:
    resolved = diagnose_headers(headers)["resolved"]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows: list[dict[str, Any]] = []
    for row in excel_rows:
        order_no = clean_text(row.get(resolved["order_no"]))
        if not order_no:
            continue
        product = clean_text(row.get(resolved["product"]))
        paid_amount = number_value(row.get(resolved["paid_amount"]))
        quantity = number_value(first_present(row, ["\u5546\u54c1\u6570\u91cf(\u4ef6)", "\u5546\u54c1\u6570\u91cf", "\u6570\u91cf", "\u8d2d\u4e70\u6570\u91cf"]))
        unit_price = round(paid_amount / quantity, 2) if paid_amount is not None and quantity not in (None, 0) else None
        order_status = clean_text(first_present(row, ["\u8ba2\u5355\u72b6\u6001", "\u4ea4\u6613\u72b6\u6001"]))
        aftersale_status = clean_text(first_present(row, ["\u552e\u540e\u72b6\u6001", "\u9000\u6b3e\u72b6\u6001"]))
        refund_amount = pdd_refund_amount(row, paid_amount, order_status, aftersale_status)
        actual_quantity = actual_sold_quantity(
            quantity=quantity,
            product=product,
            unit_price=unit_price,
            refund_amount=refund_amount,
            trade_status=order_status,
            fulfill_status=aftersale_status,
        )
        rows.append(
            {
                F_UNIQUE_KEY: f"pdd_{order_no}",
                F_PLATFORM: "\u62fc\u591a\u591a",
                F_DATA_SOURCE: "\u62fc\u591a\u591a\u5bfc\u51faExcel",
                F_SHOP_ID: clean_text(first_present(row, ["\u5e97\u94faID", "\u5e97\u94faid", "\u5e97\u94fa\u7f16\u53f7"])),
                F_SHOP_NAME: clean_text(first_present(row, ["\u5e97\u94fa\u540d\u79f0", "\u5e97\u94fa"])),
                F_FETCHED_AT: now,
                F_ORDER_NO: order_no,
                F_CREATED_AT: pdd_created_at(row, order_no),
                F_BUYER_NICK: clean_text(first_present(row, ["\u4e70\u5bb6\u6635\u79f0", "\u7528\u6237\u6635\u79f0"])),
                F_PRODUCT: product,
                F_UNIT_PRICE: unit_price,
                F_QUANTITY: actual_quantity,
                F_FULFILL_STATUS: "/".join(item for item in [order_status, aftersale_status] if item),
                F_TRADE_STATUS: order_status,
                F_PAID_AMOUNT: paid_amount,
                F_REFUND_AMOUNT: refund_amount,
                F_OPERATION: "\u62fc\u591a\u591aExcel\u5bfc\u5165",
                "\u62fc\u591a\u591a\u8ba2\u5355\u72b6\u6001": order_status,
                "\u62fc\u591a\u591a\u552e\u540e\u72b6\u6001": aftersale_status,
                "\u62fc\u591a\u591a\u627f\u8bfa\u53d1\u8d27\u65f6\u95f4": clean_text(first_present(row, ["\u627f\u8bfa\u53d1\u8d27\u65f6\u95f4"])),
                "\u62fc\u591a\u591a\u5feb\u9012\u5355\u53f7": clean_text(first_present(row, ["\u5feb\u9012\u5355\u53f7", "\u7269\u6d41\u5355\u53f7"])),
                "\u62fc\u591a\u591a\u7269\u6d41\u516c\u53f8": clean_text(first_present(row, ["\u7269\u6d41\u516c\u53f8", "\u5feb\u9012\u516c\u53f8"])),
                "\u62fc\u591a\u591a\u5546\u54c1ID": clean_text(first_present(row, ["\u5546\u54c1id", "\u5546\u54c1ID"])),
                "\u62fc\u591a\u591a\u5546\u54c1\u89c4\u683c": clean_text(first_present(row, ["\u5546\u54c1\u89c4\u683c"])),
                "\u62fc\u591a\u591a\u5546\u5bb6\u7f16\u7801": clean_text(first_present(row, ["\u5546\u5bb6\u7f16\u7801-\u89c4\u683c\u7ef4\u5ea6", "\u5546\u5bb6\u7f16\u7801-\u5546\u54c1\u7ef4\u5ea6", "\u5546\u5bb6\u7f16\u7801"])),
                "\u62fc\u591a\u591a\u652f\u4ed8\u65b9\u5f0f": clean_text(first_present(row, ["\u652f\u4ed8\u65b9\u5f0f"])),
                "\u62fc\u591a\u591a\u7528\u6237\u5b9e\u4ed8\u91d1\u989d": number_value(first_present(row, ["\u7528\u6237\u5b9e\u4ed8\u91d1\u989d(\u5143)", "\u7528\u6237\u5b9e\u4ed8\u91d1\u989d"])),
                F_RAW: json.dumps(redact_row(row), ensure_ascii=False, sort_keys=True, default=str),
            }
        )
    return rows


def pdd_refund_amount(
    row: dict[str, Any],
    paid_amount: float | None,
    order_status: str,
    aftersale_status: str,
) -> float:
    explicit_refund = number_value(
        first_present(
            row,
            [
                "\u9000\u6b3e\u91d1\u989d",
                "\u5df2\u9000\u6b3e\u91d1\u989d",
                "\u552e\u540e\u9000\u6b3e\u91d1\u989d",
                "\u8ba2\u5355\u9000\u6b3e\u91d1\u989d",
                "\u5546\u54c1\u5df2\u9000\u6b3e\u91d1\u989d",
            ],
        )
    )
    if explicit_refund is not None:
        return explicit_refund
    if "\u9000\u6b3e\u6210\u529f" in f"{order_status}/{aftersale_status}":
        return paid_amount or 0
    return 0


NON_SOLD_STATUS_KEYWORDS = ("\u9000\u6b3e", "\u4ea4\u6613\u5173\u95ed", "\u5df2\u5173\u95ed", "\u5df2\u53d6\u6d88", "\u8ba2\u5355\u5173\u95ed", "\u5f85\u4ed8\u6b3e", "\u7b49\u5f85\u4e70\u5bb6\u4ed8\u6b3e", "\u672a\u4ed8\u6b3e")
NON_SOLD_PRODUCT_KEYWORDS = ("\u8865\u6536\u5dee\u4ef7", "\u5dee\u4ef7\u4e13\u7528", "\u8d2d\u4e70\u524d\u987b\u8054\u7cfb\u5ba2\u670d", "\u8054\u7cfb\u5ba2\u670d\u786e\u8ba4")


def actual_sold_quantity(
    *,
    quantity: float | None,
    product: str,
    unit_price: float | None,
    refund_amount: float | None,
    trade_status: str,
    fulfill_status: str,
) -> float | None:
    if refund_amount and refund_amount > 0:
        return 0
    text = f"{trade_status}/{fulfill_status}".replace("\u65e0\u552e\u540e\u6216\u552e\u540e\u53d6\u6d88", "")
    if any(keyword in text for keyword in NON_SOLD_STATUS_KEYWORDS):
        return 0
    if unit_price == 0 or any(keyword in product for keyword in NON_SOLD_PRODUCT_KEYWORDS):
        return 0
    return quantity


def first_present(row: dict[str, Any], keys: list[str]) -> Any:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


def pdd_created_at(row: dict[str, Any], order_no: str) -> str:
    value = normalize_datetime_text(
        first_present(
            row,
            [
                "\u8ba2\u5355\u6210\u4ea4\u65f6\u95f4",
                "\u652f\u4ed8\u65f6\u95f4",
                "\u4e0b\u5355\u65f6\u95f4",
                "\u6210\u56e2\u65f6\u95f4",
                "\u8ba2\u5355\u521b\u5efa\u65f6\u95f4",
            ],
        )
    )
    if value:
        return value
    return pdd_date_from_order_no(order_no)


def pdd_date_from_order_no(order_no: str) -> str:
    prefix = clean_text(order_no).split("-", 1)[0]
    if len(prefix) != 6 or not prefix.isdigit():
        return ""
    month = int(prefix[2:4])
    day = int(prefix[4:6])
    if not (1 <= month <= 12 and 1 <= day <= 31):
        return ""
    return f"20{prefix[:2]}-{month:02d}-{day:02d}"


def clean_text(value: Any) -> str:
    if value in (None, "-"):
        return ""
    return str(value).strip().rstrip("\t")


def number_value(value: Any) -> float | None:
    text = clean_text(value)
    if not text:
        return None
    return round(float(text.replace(",", "")), 2)


def normalize_datetime_text(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    normalized = text.replace("/", "-")
    for candidate in (normalized, normalized[:19], normalized[:16], normalized[:10]):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(candidate, fmt)
                if fmt == "%Y-%m-%d":
                    return parsed.strftime("%Y-%m-%d")
                return parsed.strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
    return text


def redact_row(row: dict[str, Any]) -> dict[str, Any]:
    redacted: dict[str, Any] = {}
    for key, value in row.items():
        if any(part in str(key) for part in SENSITIVE_HEADER_PARTS):
            redacted[key] = "[REDACTED]"
        else:
            redacted[key] = value
    return redacted


def readback_pdd_rows(writer: FeishuOrderWriter, unique_keys: set[str]) -> dict[str, dict[str, Any]]:
    found: dict[str, dict[str, Any]] = {}
    page_token = None
    while True:
        params: dict[str, Any] = {"page_size": 500}
        if page_token:
            params["page_token"] = page_token
        data = writer._request(
            "GET",
            f"/bitable/v1/apps/{writer.app_token}/tables/{writer.table_id}/records",
            params=params,
        )
        for item in data.get("items", []) or []:
            fields = item.get("fields") or {}
            unique_key = fields.get(F_UNIQUE_KEY)
            if unique_key in unique_keys:
                found[str(unique_key)] = fields
        if not data.get("has_more"):
            return found
        page_token = data.get("page_token")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True, help="Path to a Pinduoduo .csv or .xlsx order export.")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    headers, excel_rows = load_source_rows(excel_path)
    diagnosis = diagnose_headers(headers)
    if not diagnosis["is_pdd_like"]:
        print(
            json.dumps(
                {
                    "status": "not_pinduoduo_order_export",
                    "excel_file": str(excel_path),
                    "row_count": len(excel_rows),
                    "headers": headers,
                    "diagnosis": diagnosis,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 2

    feishu_rows = pdd_feishu_rows(headers, excel_rows, excel_path)
    if args.dry_run:
        print(
            json.dumps(
                {
                    "status": "dry_run",
                    "excel_file": str(excel_path),
                    "row_count": len(excel_rows),
                    "mapped_count": len(feishu_rows),
                    "sample": feishu_rows[:3],
                    "diagnosis": diagnosis,
                },
                ensure_ascii=False,
                indent=2,
                default=str,
            )
        )
        return 0

    writer = PinduoduoOrderWriter(load_settings())
    writer.ensure_fields()
    created_fields = writer.ensure_pdd_fields()
    saved_count = writer.write_orders(feishu_rows)
    readback = readback_pdd_rows(writer, {row[F_UNIQUE_KEY] for row in feishu_rows})
    missing = sorted(set(row[F_UNIQUE_KEY] for row in feishu_rows) - set(readback))
    mismatched = [
        row[F_UNIQUE_KEY]
        for row in feishu_rows
        if row[F_UNIQUE_KEY] in readback and str(readback[row[F_UNIQUE_KEY]].get(F_ORDER_NO) or "") != str(row[F_ORDER_NO])
    ]
    print(
        json.dumps(
            {
                "status": "success" if not missing and not mismatched else "readback_mismatch",
                "excel_file": str(excel_path),
                "row_count": len(excel_rows),
                "saved_count": saved_count,
                "readback_count": len(readback),
                "created_fields": created_fields,
                "missing_unique_keys": missing[:30],
                "mismatched_unique_keys": mismatched[:30],
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0 if not missing and not mismatched else 4


if __name__ == "__main__":
    raise SystemExit(main())
