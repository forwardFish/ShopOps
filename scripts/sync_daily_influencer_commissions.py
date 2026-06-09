from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from shopops.config import load_settings
from shopops.storage.feishu_bootstrap import FEISHU_BASE_URL, FeishuOpenApiClient


TEXT_FIELD = 1
NUMBER_FIELD = 2

F_UNIQUE_KEY = "unique_key"
F_PLATFORM = "平台"
F_SOURCE = "数据来源"
F_ORDER_NO = "订单号"
F_CREATED_AT = "下单时间"
F_PAY_AT = "支付时间"
F_ORDER_STATUS = "订单状态"
F_PRODUCT_ID = "商品ID"
F_PRODUCT_NAME = "商品名称"
F_QUANTITY = "商品数量"
F_PAID_AMOUNT = "支付金额"
F_INFLUENCER_ID = "带货达人ID"
F_INFLUENCER_NICK = "带货达人昵称"
F_COMMISSION_RATE = "带货佣金率"
F_COMMISSION = "带货费用"
F_COMMISSION_BASIS = "带货费用口径"
F_SHOP_ID = "店铺ID"
F_SHOP_NAME = "店铺名称"
F_FETCHED_AT = "采集时间"
F_SOURCE_FILE = "来源文件"
F_RAW = "原始数据"

TARGET_FIELDS = [
    (F_UNIQUE_KEY, TEXT_FIELD),
    (F_PLATFORM, TEXT_FIELD),
    (F_SOURCE, TEXT_FIELD),
    (F_ORDER_NO, TEXT_FIELD),
    (F_CREATED_AT, TEXT_FIELD),
    (F_PAY_AT, TEXT_FIELD),
    (F_ORDER_STATUS, TEXT_FIELD),
    (F_PRODUCT_ID, TEXT_FIELD),
    (F_PRODUCT_NAME, TEXT_FIELD),
    (F_QUANTITY, NUMBER_FIELD),
    (F_PAID_AMOUNT, NUMBER_FIELD),
    (F_INFLUENCER_ID, TEXT_FIELD),
    (F_INFLUENCER_NICK, TEXT_FIELD),
    (F_COMMISSION_RATE, TEXT_FIELD),
    (F_COMMISSION, NUMBER_FIELD),
    (F_COMMISSION_BASIS, TEXT_FIELD),
    (F_SHOP_ID, TEXT_FIELD),
    (F_SHOP_NAME, TEXT_FIELD),
    (F_FETCHED_AT, TEXT_FIELD),
    (F_SOURCE_FILE, TEXT_FIELD),
    (F_RAW, TEXT_FIELD),
]

SENSITIVE_KEY_PARTS = ("收件", "收货", "手机", "电话", "地址", "详细地址", "消费者资料")


class FeishuInfluencerTableClient:
    def __init__(self, table_id: str) -> None:
        settings = load_settings()
        self.app_token = settings.shopops_data_center_app_token or settings.feishu_app_token
        if not self.app_token:
            raise RuntimeError("Missing SHOPOPS_DATA_CENTER_APP_TOKEN or FEISHU_APP_TOKEN")
        self.table_id = table_id
        self.session = requests.Session()
        self.session.trust_env = False
        os.environ["NO_PROXY"] = "open.feishu.cn"
        os.environ["no_proxy"] = "open.feishu.cn"
        self.auth = FeishuOpenApiClient(settings.feishu_app_id, settings.feishu_app_secret)

    def request(
        self,
        method: str,
        path: str,
        payload: dict[str, Any] | None = None,
        params: dict[str, Any] | None = None,
        allow_duplicate_field: bool = False,
    ) -> dict[str, Any]:
        response = self.session.request(
            method,
            f"{FEISHU_BASE_URL}{path}",
            headers=self.auth.headers(),
            json=payload,
            params=params,
            timeout=30,
        )
        body = response.json()
        if allow_duplicate_field and body.get("code") == 1254014:
            return {}
        if response.status_code >= 400 or body.get("code") != 0:
            raise RuntimeError(f"Feishu API {method} {path} failed HTTP {response.status_code}: {body}")
        return body.get("data") or {}

    def field_names(self) -> set[str]:
        names: set[str] = set()
        page_token = None
        while True:
            params: dict[str, Any] = {"page_size": 100}
            if page_token:
                params["page_token"] = page_token
            data = self.request(
                "GET",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/fields",
                params=params,
            )
            for item in data.get("items") or []:
                if item.get("field_name"):
                    names.add(str(item["field_name"]))
            if not data.get("has_more"):
                return names
            page_token = data.get("page_token")

    def ensure_fields(self) -> list[str]:
        existing = self.field_names()
        created: list[str] = []
        for field_name, field_type in TARGET_FIELDS:
            if field_name in existing:
                continue
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/fields",
                {"field_name": field_name, "type": field_type},
                allow_duplicate_field=True,
            )
            existing.add(field_name)
            created.append(field_name)
        return created

    def record_index(self) -> dict[str, str]:
        records: dict[str, str] = {}
        for item in self.iter_records():
            fields = item.get("fields") or {}
            unique_key = scalar_text(fields.get(F_UNIQUE_KEY))
            if unique_key:
                records[unique_key] = str(item.get("record_id"))
        return records

    def iter_records(self) -> Iterable[dict[str, Any]]:
        page_token = None
        while True:
            params: dict[str, Any] = {"page_size": 500}
            if page_token:
                params["page_token"] = page_token
            data = self.request(
                "GET",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records",
                params=params,
            )
            yield from data.get("items") or []
            if not data.get("has_more"):
                return
            page_token = data.get("page_token")

    def upsert_rows(self, rows: list[dict[str, Any]]) -> dict[str, int]:
        existing = self.record_index()
        create_rows: list[dict[str, Any]] = []
        update_rows: list[dict[str, Any]] = []
        for row in rows:
            clean = {key: value for key, value in row.items() if value not in (None, "")}
            unique_key = scalar_text(row.get(F_UNIQUE_KEY))
            record_id = existing.get(unique_key)
            if record_id:
                update_rows.append({"record_id": record_id, "fields": clean})
            else:
                create_rows.append({"fields": clean})
        for chunk in chunks(create_rows, 500):
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records/batch_create", {"records": chunk})
        for chunk in chunks(update_rows, 500):
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records/batch_update", {"records": chunk})
        return {"created": len(create_rows), "updated": len(update_rows), "saved": len(create_rows) + len(update_rows)}

    def readback(self, unique_keys: set[str]) -> dict[str, dict[str, Any]]:
        found: dict[str, dict[str, Any]] = {}
        if not unique_keys:
            return found
        for item in self.iter_records():
            fields = item.get("fields") or {}
            unique_key = scalar_text(fields.get(F_UNIQUE_KEY))
            if unique_key in unique_keys:
                found[unique_key] = fields
        return found


def latest_file(directory: Path) -> Path | None:
    candidates = [path for path in directory.iterdir() if path.suffix.lower() in {".csv", ".xlsx"}]
    if not candidates:
        return None
    return sorted(candidates, key=lambda item: item.stat().st_mtime, reverse=True)[0]


def load_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            headers = [str(header or "").strip() for header in reader.fieldnames or []]
            rows = [
                {str(key or "").strip(): value for key, value in row.items()}
                for row in reader
                if any(value not in (None, "") for value in row.values())
            ]
        return headers, rows
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    headers = [str(value or "").strip() for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return headers, rows


def build_rows_for_file(platform: str, path: Path, fetched_at: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    headers, source_rows = load_rows(path)
    if platform == "抖音":
        rows = douyin_rows(source_rows, path, fetched_at)
    elif platform == "视频号":
        rows = wechat_channels_rows(source_rows, path, fetched_at)
    else:
        raise ValueError(f"Unsupported influencer platform: {platform}")
    duplicate_keys = [key for key, count in Counter(row[F_UNIQUE_KEY] for row in rows).items() if count > 1]
    return rows, {
        "file": str(path),
        "headers": headers,
        "source_row_count": len(source_rows),
        "mapped_row_count": len(rows),
        "duplicate_unique_keys": duplicate_keys[:50],
    }


def collapse_rows_by_unique_key(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    collapsed: dict[str, dict[str, Any]] = {}
    raw_rows: dict[str, list[Any]] = {}
    for row in rows:
        unique_key = row[F_UNIQUE_KEY]
        if unique_key not in collapsed:
            collapsed[unique_key] = dict(row)
            raw_rows[unique_key] = [json.loads(row[F_RAW])] if row.get(F_RAW) else []
            continue
        current = collapsed[unique_key]
        raw_rows[unique_key].append(json.loads(row[F_RAW]) if row.get(F_RAW) else {})
        for field in (F_PRODUCT_ID, F_PRODUCT_NAME, F_INFLUENCER_ID, F_INFLUENCER_NICK, F_COMMISSION_BASIS):
            current[field] = join_unique(current.get(field), row.get(field))
        current[F_QUANTITY] = add_numbers(current.get(F_QUANTITY), row.get(F_QUANTITY))
        current[F_COMMISSION] = add_numbers(current.get(F_COMMISSION), row.get(F_COMMISSION))
        current[F_PAID_AMOUNT] = first_number(current.get(F_PAID_AMOUNT), row.get(F_PAID_AMOUNT))
        current[F_RAW] = json.dumps(raw_rows[unique_key], ensure_ascii=False, sort_keys=True, default=str)
    return list(collapsed.values())


def douyin_rows(source_rows: list[dict[str, Any]], source_file: Path, fetched_at: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in source_rows:
        order_no = clean_text(first_present(row, "主订单编号", "订单id", "订单ID", "订单号"))
        influencer_id = clean_text(row.get("达人ID"))
        influencer_nick = clean_text(row.get("达人昵称"))
        commission, basis = first_number_with_basis(row, ("达人实际承担优惠金额", "达人优惠"))
        if not order_no or not any((influencer_id, influencer_nick, commission)):
            continue
        rows.append(
            base_row(
                platform="抖音",
                source="抖店订单导出",
                order_no=order_no,
                created_at=clean_text(row.get("订单提交时间")),
                pay_at=clean_text(row.get("支付完成时间")),
                order_status=clean_text(row.get("订单状态")),
                product_id=clean_text(first_present(row, "商品ID", "商品id")),
                product_name=clean_text(first_present(row, "选购商品", "商品名称")),
                quantity=number_value(row.get("商品数量")),
                paid_amount=number_value(first_present(row, "订单应付金额", "支付金额")),
                influencer_id=influencer_id,
                influencer_nick=influencer_nick,
                commission_rate="",
                commission=commission,
                commission_basis=basis,
                shop_id=clean_text(first_present(row, "所属门店ID", "店铺id", "店铺ID")),
                shop_name="抖音",
                fetched_at=fetched_at,
                source_file=source_file,
                raw=row,
            )
        )
    return rows


def wechat_channels_rows(source_rows: list[dict[str, Any]], source_file: Path, fetched_at: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in source_rows:
        order_no = clean_text(row.get("订单号"))
        influencer_nick = clean_text(row.get("带货账号昵称"))
        commission = number_value(row.get("带货费用"))
        if not order_no or not any((influencer_nick, commission, clean_text(row.get("带货方式")), clean_text(row.get("带货费用渠道")))):
            continue
        rows.append(
            base_row(
                platform="视频号",
                source="微信小店订单导出",
                order_no=order_no,
                created_at=clean_text(row.get("订单下单时间")),
                pay_at=clean_text(row.get("支付时间")),
                order_status=clean_text(row.get("订单状态")),
                product_id=clean_text(row.get("商品编码(平台)")),
                product_name=clean_text(row.get("商品名称")),
                quantity=number_value(row.get("商品数量")),
                paid_amount=number_value(row.get("订单实际支付金额")),
                influencer_id="",
                influencer_nick=influencer_nick,
                commission_rate=clean_text(row.get("带货佣金率")),
                commission=commission,
                commission_basis=clean_text(row.get("带货费用类型")) or "带货费用",
                shop_id="",
                shop_name="视频号",
                fetched_at=fetched_at,
                source_file=source_file,
                raw=redact_row(row),
            )
        )
    return rows


def base_row(
    *,
    platform: str,
    source: str,
    order_no: str,
    created_at: str,
    pay_at: str,
    order_status: str,
    product_id: str,
    product_name: str,
    quantity: float | None,
    paid_amount: float | None,
    influencer_id: str,
    influencer_nick: str,
    commission_rate: str,
    commission: float | None,
    commission_basis: str,
    shop_id: str,
    shop_name: str,
    fetched_at: str,
    source_file: Path,
    raw: dict[str, Any],
) -> dict[str, Any]:
    return {
        F_UNIQUE_KEY: canonical_unique_key(platform, order_no),
        F_PLATFORM: platform,
        F_SOURCE: source,
        F_ORDER_NO: order_no,
        F_CREATED_AT: created_at,
        F_PAY_AT: pay_at,
        F_ORDER_STATUS: order_status,
        F_PRODUCT_ID: product_id,
        F_PRODUCT_NAME: product_name,
        F_QUANTITY: quantity,
        F_PAID_AMOUNT: paid_amount,
        F_INFLUENCER_ID: influencer_id,
        F_INFLUENCER_NICK: influencer_nick,
        F_COMMISSION_RATE: commission_rate,
        F_COMMISSION: commission,
        F_COMMISSION_BASIS: commission_basis,
        F_SHOP_ID: shop_id,
        F_SHOP_NAME: shop_name,
        F_FETCHED_AT: fetched_at,
        F_SOURCE_FILE: str(source_file),
        F_RAW: json.dumps(raw, ensure_ascii=False, sort_keys=True, default=str),
    }


def canonical_unique_key(platform: str, order_no: Any) -> str:
    cleaned = clean_text(order_no).replace("\t", "").strip("'")
    return f"{platform}{cleaned}" if cleaned else ""


def first_present(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


def first_number_with_basis(row: dict[str, Any], keys: tuple[str, ...]) -> tuple[float | None, str]:
    for key in keys:
        value = number_value(row.get(key))
        if value not in (None, 0):
            return value, key
    return None, ""


def join_unique(left: Any, right: Any) -> str:
    values: list[str] = []
    for value in (left, right):
        for part in clean_text(value).split(";"):
            cleaned = part.strip()
            if cleaned and cleaned not in values:
                values.append(cleaned)
    return "; ".join(values)


def add_numbers(left: Any, right: Any) -> float | None:
    values = [value for value in (number_value(left), number_value(right)) if value is not None]
    if not values:
        return None
    return round(sum(values), 4)


def first_number(left: Any, right: Any) -> float | None:
    left_number = number_value(left)
    return left_number if left_number is not None else number_value(right)


def clean_text(value: Any) -> str:
    if value in (None, "-", "--"):
        return ""
    return str(value).strip().strip("\t").rstrip("\t")


def scalar_text(value: Any) -> str:
    if isinstance(value, list):
        return "".join(str(item.get("text") if isinstance(item, dict) else item) for item in value).strip()
    return clean_text(value)


def number_value(value: Any) -> float | None:
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1]
    try:
        return round(float(text), 4)
    except ValueError:
        return None


def redact_row(row: dict[str, Any]) -> dict[str, Any]:
    return {key: ("[REDACTED]" if any(part in str(key) for part in SENSITIVE_KEY_PARTS) else value) for key, value in row.items()}


def chunks(values: list[dict[str, Any]], size: int) -> list[list[dict[str, Any]]]:
    return [values[index : index + size] for index in range(0, len(values), size)]


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync daily Douyin/WeChat Channels influencer commission exports into Feishu.")
    parser.add_argument("--data-root", default=r"D:\lyh\ShopOps")
    parser.add_argument("--date-dir", required=True, help="Daily directory name such as 0607.")
    parser.add_argument("--target-table", default="")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--evidence", default="")
    args = parser.parse_args()

    settings = load_settings()
    table_id = args.target_table or settings.table_douyin_influencer_commission
    if not table_id or not table_id.startswith("tbl"):
        raise RuntimeError("Missing FEISHU_TABLE_DOUYIN_INFLUENCER_COMMISSION or --target-table")

    fetched_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    data_root = Path(args.data_root)
    all_rows: list[dict[str, Any]] = []
    files: dict[str, Any] = {}
    missing_platform_dirs: list[str] = []
    for platform in ("抖音", "视频号"):
        platform_dir = data_root / platform / args.date_dir
        if not platform_dir.exists():
            missing_platform_dirs.append(str(platform_dir))
            continue
        source_file = latest_file(platform_dir)
        if not source_file:
            files[platform] = {"status": "missing_export_file", "directory": str(platform_dir)}
            continue
        rows, info = build_rows_for_file(platform, source_file, fetched_at)
        all_rows.extend(rows)
        files[platform] = info

    mapped_line_count = len(all_rows)
    duplicate_unique_keys = [key for key, count in Counter(row[F_UNIQUE_KEY] for row in all_rows).items() if count > 1]
    all_rows = collapse_rows_by_unique_key(all_rows)
    summary: dict[str, Any] = {
        "status": "dry_run" if args.dry_run else "started",
        "target_table": table_id,
        "unique_key_rule": "平台+订单编号",
        "data_root": str(data_root),
        "date_dir": args.date_dir,
        "files": files,
        "missing_platform_dirs": missing_platform_dirs,
        "mapped_line_count": mapped_line_count,
        "source_row_count": len(all_rows),
        "platform_counts": dict(Counter(row[F_PLATFORM] for row in all_rows)),
        "collapsed_duplicate_unique_key_count": len(duplicate_unique_keys),
        "collapsed_duplicate_unique_keys_sample": duplicate_unique_keys[:50],
        "sample_unique_keys": [row[F_UNIQUE_KEY] for row in all_rows[:20]],
    }
    if args.dry_run:
        print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
        return 0

    client = FeishuInfluencerTableClient(table_id)
    created_fields = client.ensure_fields()
    upsert = client.upsert_rows(all_rows)
    readback = client.readback({row[F_UNIQUE_KEY] for row in all_rows})
    missing = sorted(set(row[F_UNIQUE_KEY] for row in all_rows) - set(readback))
    mismatched = [
        row[F_UNIQUE_KEY]
        for row in all_rows
        if row[F_UNIQUE_KEY] in readback and scalar_text(readback[row[F_UNIQUE_KEY]].get(F_ORDER_NO)) != row[F_ORDER_NO]
    ]
    summary.update(
        {
            "status": "success" if not missing and not mismatched else "readback_mismatch",
            "created_fields": created_fields,
            "upsert": upsert,
            "readback_count": len(readback),
            "missing_unique_keys": missing[:50],
            "mismatched_unique_keys": mismatched[:50],
        }
    )
    evidence = Path(args.evidence) if args.evidence else Path("docs/live-evidence") / f"influencer-commission-{args.date_dir}.json"
    evidence.parent.mkdir(parents=True, exist_ok=True)
    evidence.write_text(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True, default=str), encoding="utf-8")
    summary["evidence_path"] = str(evidence.resolve())
    print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True, default=str))
    return 0 if summary["status"] == "success" else 4


if __name__ == "__main__":
    raise SystemExit(main())
