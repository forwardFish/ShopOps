from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import subprocess
import threading
import time
import traceback
import warnings
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from contextlib import contextmanager
from dataclasses import replace
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
warnings.filterwarnings("ignore", message="Workbook contains no default style.*", category=UserWarning)

from shopops.config import _load_dotenv, load_settings
from shopops.collectors.jushuitan_order_api import (
    JushuitanOrderApiCollector,
    extract_jushuitan_orders,
    first_value as jushuitan_first_value,
    is_unpaid as is_unpaid_jushuitan_order,
    jushuitan_public_params,
    normalize_amount as normalize_jushuitan_amount,
)
from shopops.services.product_breakdown import (
    DEFAULT_PRODUCT_CATALOG_TABLE_ID,
    effective_sales_amount,
    extract_product_code_from_raw,
    product_breakdown_values,
    product_field_names,
    product_rules_from_records,
)
from shopops.storage.feishu_bootstrap import FEISHU_BASE_URL, FeishuOpenApiClient
from scripts.write_douyin_influencer_excel_to_feishu import (
    doudian_influencer_rows as doudian_commission_excel_rows,
    parse_doudian_xlsx as parse_doudian_commission_xlsx,
)


TEXT_FIELD = 1
NUMBER_FIELD = 2
FORMULA_FIELD = 20

PLATFORMS = ("天猫", "抖音", "拼多多", "视频号")
PLATFORM_CODES = {"天猫": "tmall", "抖音": "douyin", "拼多多": "pdd", "视频号": "wechat_channels"}
DOUYIN_ORDER_DETAIL_REQUIRED_HEADER = "支付方式"
JUSHUITAN_DOUYIN_LOOKBACK_DAYS = 90
JUSHUITAN_DOUYIN_CHUNK_DAYS = 7
ORDER_ROLLING_LOOKBACK_DAYS = 90
INFLUENCER_ROLLING_LOOKBACK_DAYS = 90
ORDER_TABLE_ENV = {
    "天猫": "SHOPOPS_ORDER_TABLE_TMALL_ID",
    "抖音": "SHOPOPS_ORDER_TABLE_DOUYIN_ID",
    "拼多多": "SHOPOPS_ORDER_TABLE_PINDUODUO_ID",
    "视频号": "SHOPOPS_ORDER_TABLE_WECHAT_CHANNELS_ID",
}

F_UNIQUE_KEY = "unique_key"
F_PLATFORM = "平台"
F_DATA_SOURCE = "数据来源"
F_SHOP_ID = "店铺ID"
F_SHOP_NAME = "店铺名称"
F_FETCHED_AT = "采集时间"
F_ORDER_NO = "订单号"
F_CREATED_AT = "创建时间"
F_BUYER_NICK = "买家昵称"
F_PRODUCT_NAME = "商品名称"
F_PRODUCT_CODE = "商品编码"
F_ACCESSORY_FLAG = "是否是配件"
F_UNIT_PRICE = "单价"
F_QUANTITY = "数量"
F_FULFILL_STATUS = "履约/售后状态"
F_TRADE_STATUS = "交易状态"
F_PAID_AMOUNT = "实收款"
F_REFUND_AMOUNT = "退款金额"
F_PRODUCT_COST = "商品成本"
F_FREIGHT_COST = "运费成本"
F_PLATFORM_FEE = "平台扣点"
F_OTHER_FEE = "其他费用"
F_OPERATION = "操作信息"
F_RAW = "原始数据"

F_DATE = "投放日期"
F_SPEND = "花费"
F_PROMOTION_SPEND = "推广花费(元)"
F_ACTUAL_SPEND = "实际消耗"
F_DEAL_AMOUNT = "成交金额"
F_IMPRESSIONS = "展现量"
F_EXPOSURES = "曝光量"
F_CLICKS = "点击量"
F_CLICK_RATE = "点击率"
F_CPC = "点击单价"
F_ROI = "ROI"
F_PLATFORM_ROI = "平台显示ROI"
F_TRUE_ROI = "平台真实ROI"
F_PRODUCT_ID = "商品ID"
F_DEAL_SPEND = "成交花费(元)"
F_TOTAL_SPEND = "总花费(元)"
F_TRADE_AMOUNT = "交易额(元)"
F_NET_TRADE_AMOUNT = "净交易额(元)"
F_NET_ACTUAL_ROI = "净实际投产比"
F_NET_DEAL_COUNT = "净成交笔数"
F_COST_PER_NET_DEAL = "每笔净成交花费(元)"
F_NET_TRADE_AMOUNT_RATE = "净交易额占比"
F_NET_DEAL_COUNT_RATE = "净成交笔数占比"
F_AMOUNT_PER_NET_DEAL = "每笔净成交金额(元)"
F_SETTLED_TRADE_AMOUNT = "结算交易额(元)"
F_SETTLED_ROI = "结算投产比"
F_SETTLED_DEAL_COUNT = "结算成交笔数"
F_REFUND_EXEMPTION_RATE = "退款豁免率"
F_REFUND_ORDER_EXEMPTION_RATE = "退单豁免率"
F_COST_PER_SETTLED_DEAL = "每笔结算成交花费(元)"
F_TRADE_AMOUNT_SETTLEMENT_RATE = "交易额结算率"
F_ORDER_SETTLEMENT_RATE = "订单结算率"
F_AMOUNT_PER_SETTLED_DEAL = "每笔结算成交金额(元)"
F_DEAL_COUNT = "成交笔数"
F_COST_PER_DEAL = "每笔成交花费(元)"
F_AMOUNT_PER_DEAL = "每笔成交金额(元)"
INTERNAL_PRODUCT_BREAKDOWN_QUANTITY = "__product_breakdown_quantity"

I_SOURCE = "数据来源"
I_CREATED_AT = "下单时间"
I_PAY_AT = "支付时间"
I_STATUS = "订单状态"
I_INFLUENCER_ID = "带货达人ID"
I_INFLUENCER_NICK = "带货达人昵称"
I_COMMISSION_RATE = "带货佣金率"
I_COMMISSION = "带货费用"
I_COMMISSION_BASIS = "带货费用口径"
I_COMMISSION_RATE_NUM = "佣金率"
I_ESTIMATED_COMMISSION = "预估佣金支出"
I_ACTUAL_COMMISSION = "实际佣金支出"
I_SOURCE_FILE = "来源文件"

ORDER_FIELDS = [
    F_UNIQUE_KEY,
    F_PLATFORM,
    F_DATA_SOURCE,
    F_SHOP_ID,
    F_SHOP_NAME,
    F_FETCHED_AT,
    F_ORDER_NO,
    F_CREATED_AT,
    F_BUYER_NICK,
    F_PRODUCT_NAME,
    F_PRODUCT_CODE,
    F_ACCESSORY_FLAG,
    F_UNIT_PRICE,
    F_QUANTITY,
    F_FULFILL_STATUS,
    F_TRADE_STATUS,
    F_PAID_AMOUNT,
    F_REFUND_AMOUNT,
    F_PRODUCT_COST,
    F_FREIGHT_COST,
    F_PLATFORM_FEE,
    F_OTHER_FEE,
    F_OPERATION,
    F_RAW,
]
ORDER_UPDATE_FIELDS = set(ORDER_FIELDS)

AD_FIELDS = [
    F_UNIQUE_KEY,
    F_PLATFORM,
    F_DATA_SOURCE,
    F_SHOP_ID,
    F_SHOP_NAME,
    F_FETCHED_AT,
    F_DATE,
    F_PRODUCT_ID,
    F_PRODUCT_NAME,
    F_SPEND,
    F_PROMOTION_SPEND,
    F_ACTUAL_SPEND,
    F_DEAL_AMOUNT,
    F_DEAL_SPEND,
    F_TOTAL_SPEND,
    F_TRADE_AMOUNT,
    F_NET_TRADE_AMOUNT,
    F_NET_ACTUAL_ROI,
    F_NET_DEAL_COUNT,
    F_COST_PER_NET_DEAL,
    F_NET_TRADE_AMOUNT_RATE,
    F_NET_DEAL_COUNT_RATE,
    F_AMOUNT_PER_NET_DEAL,
    F_SETTLED_TRADE_AMOUNT,
    F_SETTLED_ROI,
    F_SETTLED_DEAL_COUNT,
    F_REFUND_EXEMPTION_RATE,
    F_REFUND_ORDER_EXEMPTION_RATE,
    F_COST_PER_SETTLED_DEAL,
    F_TRADE_AMOUNT_SETTLEMENT_RATE,
    F_ORDER_SETTLEMENT_RATE,
    F_AMOUNT_PER_SETTLED_DEAL,
    F_DEAL_COUNT,
    F_COST_PER_DEAL,
    F_AMOUNT_PER_DEAL,
    F_IMPRESSIONS,
    F_EXPOSURES,
    F_CLICKS,
    F_CLICK_RATE,
    F_CPC,
    F_ROI,
    F_PLATFORM_ROI,
    F_TRUE_ROI,
    F_RAW,
]

AD_FIELD_TYPES = {
    field: TEXT_FIELD
    for field in (F_UNIQUE_KEY, F_PLATFORM, F_DATA_SOURCE, F_SHOP_ID, F_SHOP_NAME, F_FETCHED_AT, F_DATE, F_PRODUCT_ID, F_PRODUCT_NAME, F_RAW)
} | {
    field: NUMBER_FIELD
    for field in AD_FIELDS
    if field not in {F_UNIQUE_KEY, F_PLATFORM, F_DATA_SOURCE, F_SHOP_ID, F_SHOP_NAME, F_FETCHED_AT, F_DATE, F_PRODUCT_ID, F_PRODUCT_NAME, F_RAW}
}

INFLUENCER_FIELDS = [
    F_UNIQUE_KEY,
    F_PLATFORM,
    I_SOURCE,
    F_ORDER_NO,
    I_CREATED_AT,
    I_PAY_AT,
    I_STATUS,
    F_PRODUCT_ID,
    F_PRODUCT_NAME,
    F_QUANTITY,
    F_PAID_AMOUNT,
    I_INFLUENCER_ID,
    I_INFLUENCER_NICK,
    I_COMMISSION_RATE,
    I_COMMISSION,
    I_COMMISSION_BASIS,
    I_COMMISSION_RATE_NUM,
    I_ESTIMATED_COMMISSION,
    I_ACTUAL_COMMISSION,
    F_SHOP_ID,
    F_SHOP_NAME,
    F_FETCHED_AT,
    I_SOURCE_FILE,
    F_RAW,
]

SENSITIVE_KEY_PARTS = (
    "收件",
    "收货",
    "手机",
    "电话",
    "地址",
    "消费者资料",
    "用户购买手机号",
    "详细地址",
    "receiver",
    "address",
    "phone",
    "mobile",
)


class FeishuDailyClient:
    def __init__(self) -> None:
        settings = load_settings()
        self.app_token = settings.shopops_data_center_app_token or settings.feishu_app_token
        if not self.app_token:
            raise RuntimeError("Missing SHOPOPS_DATA_CENTER_APP_TOKEN or FEISHU_APP_TOKEN")
        self.auth = FeishuOpenApiClient(settings.feishu_app_id, settings.feishu_app_secret)
        self.session = requests.Session()
        self.session.trust_env = False
        os.environ["NO_PROXY"] = "open.feishu.cn"
        os.environ["no_proxy"] = "open.feishu.cn"

    def request(
        self,
        method: str,
        path: str,
        payload: dict[str, Any] | None = None,
        params: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        last_error: Exception | None = None
        max_attempts = 6
        for attempt in range(1, max_attempts + 1):
            try:
                response = self.session.request(
                    method,
                    f"{FEISHU_BASE_URL}{path}",
                    headers=self.auth.headers(),
                    json=payload,
                    params=params,
                    timeout=(10, 90),
                )
                body = response.json()
                if not is_retryable_feishu_response(response.status_code, body):
                    break
                if attempt == max_attempts:
                    break
                time.sleep(min(30, attempt * 5))
            except requests.RequestException as exc:
                last_error = exc
                if attempt == max_attempts:
                    raise
                time.sleep(min(30, attempt * 5))
        else:
            raise RuntimeError(f"Feishu API request failed: {last_error}")
        if response.status_code >= 400 or body.get("code") != 0:
            raise RuntimeError(f"Feishu API {method} {path} failed HTTP {response.status_code}: {body}")
        return body.get("data") or {}

    def field_names(self, table_id: str) -> set[str]:
        names: set[str] = set()
        page_token = None
        while True:
            params: dict[str, Any] = {"page_size": 100}
            if page_token:
                params["page_token"] = page_token
            data = self.request("GET", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/fields", params=params)
            for item in data.get("items") or []:
                if item.get("field_name"):
                    names.add(str(item["field_name"]))
            if not data.get("has_more"):
                return names
            page_token = data.get("page_token")

    def field_index(self, table_id: str) -> dict[str, dict[str, Any]]:
        fields: dict[str, dict[str, Any]] = {}
        page_token = None
        while True:
            params: dict[str, Any] = {"page_size": 100}
            if page_token:
                params["page_token"] = page_token
            data = self.request("GET", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/fields", params=params)
            for item in data.get("items") or []:
                if item.get("field_name"):
                    fields[str(item["field_name"])] = item
            if not data.get("has_more"):
                return fields
            page_token = data.get("page_token")

    def ensure_formula_field(self, table_id: str, name: str, expression: str, formatter: str) -> None:
        existing = self.field_index(table_id)
        self.ensure_formula_field_with_index(table_id, existing, name, expression, formatter)

    def ensure_formula_field_with_index(
        self,
        table_id: str,
        existing: dict[str, dict[str, Any]],
        name: str,
        expression: str,
        formatter: str,
    ) -> None:
        payload = {
            "field_name": name,
            "type": FORMULA_FIELD,
            "property": {"formatter": formatter, "formula_expression": expression},
        }
        current = existing.get(name)
        if not current:
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/fields", payload)
            existing[name] = {"field_name": name, "type": FORMULA_FIELD}
            return
        if int(current.get("type") or 0) != FORMULA_FIELD:
            raise RuntimeError(f"Field {name} exists in table {table_id}, but it is not a formula field")
        current_property = current.get("property") or {}
        if (
            current_property.get("formula_expression") == expression
            and str(current_property.get("formatter") or "") == str(formatter or "")
        ):
            return
        self.request("PUT", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/fields/{current['field_id']}", payload)

    def ensure_number_field_with_index(self, table_id: str, existing: dict[str, dict[str, Any]], name: str) -> str:
        current = existing.get(name)
        if not current:
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/fields", {"field_name": name, "type": NUMBER_FIELD})
            existing[name] = {"field_name": name, "type": NUMBER_FIELD}
            return "created"
        if int(current.get("type") or 0) == NUMBER_FIELD:
            return "reused"
        field_id = current.get("field_id")
        self.request("DELETE", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/fields/{field_id}")
        self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/fields", {"field_name": name, "type": NUMBER_FIELD})
        existing[name] = {"field_name": name, "type": NUMBER_FIELD}
        return "replaced"

    def product_rules(self, product_table_id: str) -> list[Any]:
        return product_rules_from_records(list(self.iter_records(product_table_id)))

    def ensure_product_breakdown_fields(self, table_id: str, rules: list[Any]) -> dict[str, str]:
        existing = self.field_index(table_id)
        return {
            name: self.ensure_number_field_with_index(table_id, existing, name)
            for name in product_field_names(rules)
        }

    def iter_records(self, table_id: str, field_names: list[str] | None = None) -> Iterable[dict[str, Any]]:
        page_token = None
        while True:
            params: dict[str, Any] = {"page_size": 500}
            if page_token:
                params["page_token"] = page_token
            if field_names:
                params["field_names"] = json.dumps(field_names, ensure_ascii=False)
            data = self.request("GET", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records", params=params)
            yield from data.get("items") or []
            if not data.get("has_more"):
                return
            page_token = data.get("page_token")

    def upsert_rows(
        self,
        *,
        table_id: str,
        rows: list[dict[str, Any]],
        required_fields: list[str],
        fallback_match_fields: tuple[str, ...],
        allow_partial_fields: bool = True,
        update_existing_fields: set[str] | None = None,
        clear_empty_fields: set[str] | None = None,
    ) -> dict[str, Any]:
        fields = self.field_names(table_id)
        missing_required = [field for field in required_fields if field not in fields]
        if missing_required:
            raise RuntimeError(f"Target table {table_id} is missing required existing fields: {missing_required}")

        incoming_by_identity: dict[tuple[str, ...], dict[str, Any]] = {}
        deduplicated_incoming_rows = 0
        for row in rows:
            unique_key = scalar_text(row.get(F_UNIQUE_KEY))
            fallback_key = tuple(scalar_text(row.get(field)) for field in fallback_match_fields)
            if all(fallback_key):
                identity = ("fallback", *fallback_key)
            elif unique_key:
                identity = ("unique", unique_key)
            else:
                raise RuntimeError(
                    f"Incoming row for {table_id} has neither unique_key nor complete fallback identity {fallback_match_fields}"
                )
            previous = incoming_by_identity.get(identity)
            if previous is None:
                incoming_by_identity[identity] = row
                continue
            if previous != row:
                raise RuntimeError(f"Conflicting incoming rows share identity {'|'.join(identity)}")
            deduplicated_incoming_rows += 1
        rows = list(incoming_by_identity.values())

        unique_index: dict[str, list[str]] = defaultdict(list)
        fallback_index: dict[tuple[str, ...], list[str]] = defaultdict(list)
        index_fields = sorted(fields)
        existing_by_record_id: dict[str, dict[str, Any]] = {}
        for record in self.iter_records(table_id, index_fields):
            record_id = str(record.get("record_id") or "")
            record_fields = record.get("fields") or {}
            if record_id:
                existing_by_record_id[record_id] = record_fields
            unique_key = scalar_text(record_fields.get(F_UNIQUE_KEY))
            if unique_key and record_id:
                unique_index[unique_key].append(record_id)
            fallback_key = tuple(scalar_text(record_fields.get(field)) for field in fallback_match_fields)
            if all(fallback_key) and record_id:
                fallback_index[fallback_key].append(record_id)

        dropped_fields: Counter[str] = Counter()
        changed_fields: Counter[str] = Counter()
        to_create: list[dict[str, Any]] = []
        to_update: list[dict[str, Any]] = []
        duplicate_record_ids: set[str] = set()
        duplicate_keys: set[str] = set()
        duplicate_conflicts: list[dict[str, Any]] = []
        preserved_duplicate_fields: Counter[str] = Counter()
        field_metadata: dict[str, dict[str, Any]] | None = None
        for row in rows:
            clean: dict[str, Any] = {}
            for key, value in row.items():
                if key in fields:
                    if value in (None, ""):
                        if key in (clear_empty_fields or set()):
                            clean[key] = None
                    else:
                        clean[key] = value
                else:
                    if value not in (None, ""):
                        dropped_fields[key] += 1
            if not allow_partial_fields:
                missing = [key for key in row if row.get(key) not in (None, "") and key not in fields]
                if missing:
                    raise RuntimeError(f"Target table {table_id} does not contain fields used by import: {sorted(set(missing))}")

            unique_key = scalar_text(row.get(F_UNIQUE_KEY))
            fallback_key = tuple(scalar_text(row.get(field)) for field in fallback_match_fields)
            exact_matches = list(unique_index.get(unique_key) or [])
            fallback_matches = list(fallback_index.get(fallback_key) or []) if all(fallback_key) else []
            candidate_record_ids = list(dict.fromkeys([*exact_matches, *fallback_matches]))
            record_id = ""
            preserved_field_names_for_row: set[str] = set()
            if candidate_record_ids:
                def survivor_score(candidate_id: str) -> tuple[int, int, int, str]:
                    candidate_fields = existing_by_record_id.get(candidate_id, {})
                    source_matches = sum(
                        1
                        for field, value in clean.items()
                        if field_value_equal(candidate_fields.get(field), value)
                    )
                    populated_fields = sum(value not in (None, "", []) for value in candidate_fields.values())
                    exact_unique_key = int(scalar_text(candidate_fields.get(F_UNIQUE_KEY)) == unique_key)
                    return source_matches, populated_fields, exact_unique_key, candidate_id

                record_id = max(candidate_record_ids, key=survivor_score)
            if record_id:
                extras = set(candidate_record_ids) - {record_id}
                if extras:
                    if field_metadata is None:
                        field_metadata = self.field_index(table_id)
                    managed_fields = set(row)
                    survivor_fields = dict(existing_by_record_id.get(record_id, {}))
                    fields_to_preserve: dict[str, Any] = {}
                    conflicting_fields: dict[str, list[Any]] = {}
                    for field in sorted(fields - managed_fields):
                        field_type = int((field_metadata.get(field) or {}).get("type") or 0)
                        if field_type == FORMULA_FIELD or field_type >= 1000:
                            continue
                        distinct_values: list[Any] = []
                        for candidate_id in candidate_record_ids:
                            value = existing_by_record_id.get(candidate_id, {}).get(field)
                            if value in (None, "", []):
                                continue
                            if not any(field_value_equal(value, current) for current in distinct_values):
                                distinct_values.append(value)
                        if len(distinct_values) > 1:
                            conflicting_fields[field] = distinct_values[:5]
                            continue
                        if not distinct_values or survivor_fields.get(field) not in (None, "", []):
                            continue
                        if field_type in {TEXT_FIELD, NUMBER_FIELD}:
                            fields_to_preserve[field] = distinct_values[0]
                        else:
                            conflicting_fields[field] = distinct_values
                    if conflicting_fields:
                        if len(duplicate_conflicts) < 20:
                            duplicate_conflicts.append(
                                {
                                    F_UNIQUE_KEY: unique_key,
                                    "fallback_key": list(fallback_key),
                                    "record_ids": candidate_record_ids,
                                    "conflicting_fields": conflicting_fields,
                                }
                            )
                    else:
                        clean.update(fields_to_preserve)
                        preserved_field_names_for_row.update(fields_to_preserve)
                        preserved_duplicate_fields.update(fields_to_preserve.keys())
                        duplicate_record_ids.update(extras)
                        duplicate_keys.add(unique_key or "|".join(fallback_key))
                if update_existing_fields is not None:
                    clean = {
                        key: value
                        for key, value in clean.items()
                        if key in update_existing_fields or key in preserved_field_names_for_row
                    }
                changed = {
                    key: value
                    for key, value in clean.items()
                    if not field_value_equal(existing_by_record_id.get(record_id, {}).get(key), value)
                }
                if changed:
                    changed_fields.update(changed.keys())
                    to_update.append({"record_id": record_id, "fields": changed})
            else:
                to_create.append({"fields": {key: value for key, value in clean.items() if value is not None}})

        for chunk in chunks(to_create, 500):
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_create", {"records": chunk})
        for chunk in chunks(to_update, 500):
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_update", {"records": chunk})
        for chunk in chunks([{"record_id": record_id} for record_id in sorted(duplicate_record_ids)], 500):
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_delete",
                {"records": [item["record_id"] for item in chunk]},
            )
        return {
            "created": len(to_create),
            "updated": len(to_update),
            "saved": len(to_create) + len(to_update),
            "deduplicated_incoming_rows": deduplicated_incoming_rows,
            "deleted_duplicate_records": len(duplicate_record_ids),
            "repaired_duplicate_keys": len(duplicate_keys),
            "sample_repaired_duplicate_keys": sorted(duplicate_keys)[:20],
            "duplicate_conflicts": duplicate_conflicts,
            "preserved_duplicate_fields": dict(preserved_duplicate_fields),
            "changed_fields": dict(changed_fields),
            "dropped_nonexistent_fields": dict(dropped_fields),
        }

    def ensure_missing_fields_for_rows(self, table_id: str, rows: list[dict[str, Any]], field_types: dict[str, int]) -> list[str]:
        existing = self.field_names(table_id)
        needed = sorted(
            key
            for row in rows
            for key, value in row.items()
            if value not in (None, "") and key in field_types and key not in existing
        )
        created: list[str] = []
        for field_name in dict.fromkeys(needed):
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/fields",
                {"field_name": field_name, "type": field_types[field_name]},
            )
            existing.add(field_name)
            created.append(field_name)
        return created

    def deduplicate_records(self, table_id: str, key_fields: tuple[str, ...]) -> dict[str, Any]:
        fields = self.field_names(table_id)
        missing_required = [field for field in key_fields if field not in fields]
        if missing_required:
            raise RuntimeError(f"Target table {table_id} is missing dedupe fields: {missing_required}")

        seen: set[tuple[str, ...]] = set()
        duplicate_record_ids: list[str] = []
        duplicate_keys: Counter[str] = Counter()
        for record in self.iter_records(table_id, list(key_fields)):
            record_id = str(record.get("record_id") or "")
            record_fields = record.get("fields") or {}
            key = tuple(scalar_text(record_fields.get(field)) for field in key_fields)
            if not record_id or not all(key):
                continue
            if key in seen:
                duplicate_record_ids.append(record_id)
                duplicate_keys["|".join(key)] += 1
            else:
                seen.add(key)

        for chunk in chunks([{"record_id": record_id} for record_id in duplicate_record_ids], 500):
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_delete",
                {"records": [item["record_id"] for item in chunk]},
            )
        return {
            "key_fields": list(key_fields),
            "deleted_duplicate_records": len(duplicate_record_ids),
            "duplicate_keys": len(duplicate_keys),
            "sample_duplicate_keys": list(duplicate_keys)[:20],
        }

    def canonicalize_ad_unique_keys(self, table_id: str) -> dict[str, Any]:
        fields = self.field_names(table_id)
        required = {F_UNIQUE_KEY, F_PLATFORM, F_DATE}
        missing_required = sorted(required - fields)
        if missing_required:
            raise RuntimeError(f"Target ad table {table_id} is missing fields for key canonicalization: {missing_required}")

        rows: list[tuple[str, str, str, str, str]] = []
        canonical_counts: Counter[str] = Counter()
        for record in self.iter_records(table_id, [F_UNIQUE_KEY, F_PLATFORM, F_DATE]):
            record_id = str(record.get("record_id") or "")
            record_fields = record.get("fields") or {}
            original_platform = scalar_text(record_fields.get(F_PLATFORM))
            platform = normalize_platform(original_platform)
            date_text = normalize_date(record_fields.get(F_DATE))
            unique_key = scalar_text(record_fields.get(F_UNIQUE_KEY))
            if not record_id or platform not in PLATFORM_CODES or not date_text:
                continue
            canonical = ad_unique_key(platform, date_text)
            rows.append((record_id, unique_key, canonical, original_platform, platform))
            canonical_counts[canonical] += 1

        to_update: list[dict[str, Any]] = []
        duplicate_record_ids: list[str] = []
        seen: set[str] = set()
        for record_id, unique_key, canonical, original_platform, platform in rows:
            if canonical in seen:
                duplicate_record_ids.append(record_id)
                continue
            seen.add(canonical)
            if unique_key != canonical or original_platform != platform:
                to_update.append({"record_id": record_id, "fields": {F_UNIQUE_KEY: canonical, F_PLATFORM: platform}})

        for chunk in chunks([{"record_id": record_id} for record_id in duplicate_record_ids], 500):
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_delete",
                {"records": [item["record_id"] for item in chunk]},
            )
        for chunk in chunks(to_update, 500):
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_update", {"records": chunk})
        duplicates = {key: count for key, count in canonical_counts.items() if count > 1}
        return {
            "updated": len(to_update),
            "deleted_duplicate_records": len(duplicate_record_ids),
            "duplicate_canonical_keys": len(duplicates),
            "sample_duplicate_keys": list(duplicates)[:20],
        }

    def delete_platform_records(self, table_id: str, platform: str) -> dict[str, Any]:
        record_ids: list[str] = []
        for record in self.iter_records(table_id, [F_PLATFORM]):
            record_id = str(record.get("record_id") or "")
            record_fields = record.get("fields") or {}
            if record_id and scalar_text(record_fields.get(F_PLATFORM)) == platform:
                record_ids.append(record_id)
        for chunk in chunks([{"record_id": record_id} for record_id in record_ids], 500):
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_delete",
                {"records": [item["record_id"] for item in chunk]},
            )
        return {"platform": platform, "deleted_records": len(record_ids)}

    def readback_by_unique_key(self, table_id: str, unique_keys: set[str]) -> dict[str, dict[str, Any]]:
        found: dict[str, dict[str, Any]] = {}
        if not unique_keys:
            return found
        for record in self.iter_records(table_id, [F_UNIQUE_KEY]):
            fields = record.get("fields") or {}
            unique_key = scalar_text(fields.get(F_UNIQUE_KEY))
            if unique_key in unique_keys:
                found[unique_key] = fields
        return found

    def verify_rows_by_unique_key(
        self,
        table_id: str,
        rows: list[dict[str, Any]],
        compare_fields: set[str] | None = None,
        fallback_match_fields: tuple[str, ...] = (),
        compare_empty_fields: bool = False,
    ) -> dict[str, Any]:
        expected_by_key: dict[str, dict[str, Any]] = {}
        expected_key_by_fallback: dict[tuple[str, ...], str] = {}
        for row in rows:
            unique_key = scalar_text(row.get(F_UNIQUE_KEY))
            if not unique_key:
                raise RuntimeError(f"Cannot verify row without {F_UNIQUE_KEY} in {table_id}")
            if unique_key in expected_by_key:
                raise RuntimeError(f"Cannot verify duplicate incoming unique_key {unique_key} in {table_id}")
            expected_by_key[unique_key] = row
            fallback_key = tuple(scalar_text(row.get(field)) for field in fallback_match_fields)
            if fallback_match_fields and all(fallback_key):
                previous = expected_key_by_fallback.get(fallback_key)
                if previous and previous != unique_key:
                    raise RuntimeError(f"Cannot verify duplicate incoming fallback identity {fallback_key} in {table_id}")
                expected_key_by_fallback[fallback_key] = unique_key
        if not expected_by_key:
            return {
                "status": "success",
                "checked_rows": 0,
                "matched_rows": 0,
                "readback_count": 0,
                "missing_unique_keys": [],
                "duplicate_unique_keys": {},
                "mismatched_row_count": 0,
                "mismatched_rows": [],
            }

        table_fields = self.field_names(table_id)
        fields_to_compare = {
            field
            for row in rows
            for field, value in row.items()
            if field in table_fields and (compare_empty_fields or value not in (None, ""))
        }
        if compare_fields is not None:
            fields_to_compare &= compare_fields
        requested_fields = sorted({F_UNIQUE_KEY, *fallback_match_fields, *fields_to_compare})
        found_by_key: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for record in self.iter_records(table_id, requested_fields):
            record_fields = record.get("fields") or {}
            unique_key = scalar_text(record_fields.get(F_UNIQUE_KEY))
            fallback_key = tuple(scalar_text(record_fields.get(field)) for field in fallback_match_fields)
            expected_key = expected_key_by_fallback.get(fallback_key) if all(fallback_key) else None
            if not expected_key and unique_key in expected_by_key:
                expected_key = unique_key
            if expected_key:
                found_by_key[expected_key].append(record_fields)

        missing_unique_keys = sorted(set(expected_by_key) - set(found_by_key))
        duplicate_unique_keys = {
            unique_key: len(records)
            for unique_key, records in sorted(found_by_key.items())
            if len(records) != 1
        }
        mismatched_rows: list[dict[str, Any]] = []
        mismatched_row_count = 0
        for unique_key, expected in expected_by_key.items():
            records = found_by_key.get(unique_key) or []
            if len(records) != 1:
                continue
            actual = records[0]
            mismatched_fields = {
                field: {"expected": expected.get(field), "actual": actual.get(field)}
                for field in sorted(fields_to_compare)
                if not field_value_equal(actual.get(field), expected.get(field))
            }
            if mismatched_fields:
                mismatched_row_count += 1
                if len(mismatched_rows) < 20:
                    mismatched_rows.append({F_UNIQUE_KEY: unique_key, "fields": mismatched_fields})

        checked_rows = len(expected_by_key)
        matched_rows = checked_rows - len(missing_unique_keys) - len(duplicate_unique_keys) - mismatched_row_count
        return {
            "status": "success" if matched_rows == checked_rows else "mismatch",
            "checked_rows": checked_rows,
            "matched_rows": matched_rows,
            "readback_count": sum(len(records) for records in found_by_key.values()),
            "missing_unique_keys": missing_unique_keys[:50],
            "duplicate_unique_keys": duplicate_unique_keys,
            "mismatched_row_count": mismatched_row_count,
            "mismatched_rows": mismatched_rows,
        }

    def verify_unique_identities(
        self,
        table_id: str,
        *,
        key_fields: tuple[str, ...],
    ) -> dict[str, Any]:
        identities: dict[tuple[str, ...], list[dict[str, str]]] = defaultdict(list)
        empty_identity_records: list[dict[str, Any]] = []
        total_records = 0
        requested_fields = list(dict.fromkeys([*key_fields, F_UNIQUE_KEY, F_DATA_SOURCE]))
        for record in self.iter_records(table_id, requested_fields):
            total_records += 1
            record_id = str(record.get("record_id") or "")
            record_fields = record.get("fields") or {}
            identity = tuple(scalar_text(record_fields.get(field)) for field in key_fields)
            evidence = {
                "record_id": record_id,
                F_UNIQUE_KEY: scalar_text(record_fields.get(F_UNIQUE_KEY)),
                F_DATA_SOURCE: scalar_text(record_fields.get(F_DATA_SOURCE)),
            }
            if not all(identity):
                if len(empty_identity_records) < 20:
                    empty_identity_records.append(
                        {
                            **evidence,
                            "identity": list(identity),
                        }
                    )
                continue
            identities[identity].append(evidence)

        duplicates = {
            identity: records
            for identity, records in identities.items()
            if len(records) > 1
        }
        duplicate_samples = [
            {
                "identity": list(identity),
                "count": len(records),
                "records": records[:10],
            }
            for identity, records in list(sorted(duplicates.items()))[:20]
        ]
        duplicate_extra_rows = sum(len(records) - 1 for records in duplicates.values())
        keyed_records = sum(len(records) for records in identities.values())
        empty_identity_record_count = total_records - keyed_records
        return {
            "status": "success" if not duplicates and empty_identity_record_count == 0 else "mismatch",
            "key_fields": list(key_fields),
            "total_records": total_records,
            "keyed_records": keyed_records,
            "unique_identities": len(identities),
            "empty_identity_record_count": empty_identity_record_count,
            "empty_identity_records": empty_identity_records,
            "duplicate_identities": len(duplicates),
            "duplicate_extra_rows": duplicate_extra_rows,
            "duplicate_samples": duplicate_samples,
        }

    def delete_blank_identity_records(
        self,
        table_id: str,
        *,
        key_fields: tuple[str, ...],
        content_fields: Iterable[str],
    ) -> dict[str, Any]:
        available_fields = self.field_names(table_id)
        requested_fields = [
            field
            for field in dict.fromkeys([*key_fields, F_UNIQUE_KEY, F_DATA_SOURCE, *content_fields])
            if field in available_fields
        ]
        blank_record_ids: list[str] = []
        unresolved_records: list[dict[str, Any]] = []

        def has_value(value: Any) -> bool:
            if value is None:
                return False
            if isinstance(value, str):
                return bool(value.strip())
            if isinstance(value, (list, tuple, set, dict)):
                return bool(value)
            return True

        for record in self.iter_records(table_id, requested_fields):
            record_id = str(record.get("record_id") or "")
            fields = record.get("fields") or {}
            identity = tuple(scalar_text(fields.get(field)) for field in key_fields)
            if not record_id or all(identity):
                continue
            meaningful_fields = {
                field: fields.get(field)
                for field in requested_fields
                if has_value(fields.get(field))
            }
            if meaningful_fields:
                if len(unresolved_records) < 20:
                    unresolved_records.append(
                        {
                            "record_id": record_id,
                            "identity": list(identity),
                            "meaningful_fields": meaningful_fields,
                        }
                    )
                continue
            blank_record_ids.append(record_id)

        for chunk in chunks(blank_record_ids, 500):
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_delete",
                {"records": chunk},
            )
        return {
            "deleted_blank_records": len(blank_record_ids),
            "sample_deleted_record_ids": blank_record_ids[:20],
            "unresolved_empty_identity_records": unresolved_records,
        }

    def prune_order_records_for_dates(
        self,
        table_id: str,
        *,
        source_rows: list[dict[str, Any]],
        dates: set[str],
    ) -> dict[str, Any]:
        source_keys = {scalar_text(row.get(F_UNIQUE_KEY)) for row in source_rows}
        source_keys.discard("")
        normalized_dates = {normalize_date(date) for date in dates}
        normalized_dates.discard("")
        if not normalized_dates:
            return {
                "status": "skipped",
                "reason": "no_dates",
                "deleted_records": 0,
                "source_keys": len(source_keys),
            }

        stale_record_ids: list[str] = []
        stale_samples: list[dict[str, Any]] = []
        scanned_date_records = 0
        fields = [F_UNIQUE_KEY, F_ORDER_NO, F_CREATED_AT, F_PAID_AMOUNT, F_REFUND_AMOUNT]
        for record in self.iter_records(table_id, fields):
            record_id = str(record.get("record_id") or "")
            record_fields = record.get("fields") or {}
            if normalize_date(record_fields.get(F_CREATED_AT)) not in normalized_dates:
                continue
            scanned_date_records += 1
            unique_key = scalar_text(record_fields.get(F_UNIQUE_KEY))
            if not record_id or not unique_key or unique_key in source_keys:
                continue
            stale_record_ids.append(record_id)
            if len(stale_samples) < 20:
                stale_samples.append(
                    {
                        "record_id": record_id,
                        F_UNIQUE_KEY: unique_key,
                        F_ORDER_NO: scalar_text(record_fields.get(F_ORDER_NO)),
                        F_CREATED_AT: scalar_text(record_fields.get(F_CREATED_AT)),
                        F_PAID_AMOUNT: record_fields.get(F_PAID_AMOUNT),
                        F_REFUND_AMOUNT: record_fields.get(F_REFUND_AMOUNT),
                    }
                )

        for chunk in chunks([{"record_id": record_id} for record_id in stale_record_ids], 500):
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_delete",
                {"records": [item["record_id"] for item in chunk]},
            )
        return {
            "status": "complete",
            "dates": sorted(normalized_dates),
            "source_keys": len(source_keys),
            "scanned_date_records": scanned_date_records,
            "deleted_records": len(stale_record_ids),
            "sample_deleted_records": stale_samples,
        }

    def prune_records_to_snapshot(
        self,
        table_id: str,
        *,
        source_rows: list[dict[str, Any]],
        fallback_match_fields: tuple[str, ...],
        date_field: str,
        start_date: str,
        end_date: str,
    ) -> dict[str, Any]:
        start_date = normalize_date(start_date)
        end_date = normalize_date(end_date)
        if not source_rows:
            raise RuntimeError(f"Refusing to prune {table_id} against an empty source snapshot")
        if not start_date or not end_date or start_date > end_date:
            raise RuntimeError(f"Invalid snapshot date range for {table_id}: {start_date!r} to {end_date!r}")
        source_unique_keys = {scalar_text(row.get(F_UNIQUE_KEY)) for row in source_rows}
        source_unique_keys.discard("")
        source_fallback_keys = {
            tuple(scalar_text(row.get(field)) for field in fallback_match_fields)
            for row in source_rows
        }
        source_fallback_keys = {key for key in source_fallback_keys if all(key)}
        stale_record_ids: list[str] = []
        stale_samples: list[dict[str, Any]] = []
        scanned_records = 0
        preserved_outside_scope = 0
        fields = [F_UNIQUE_KEY, *fallback_match_fields, date_field]
        for record in self.iter_records(table_id, fields):
            record_id = str(record.get("record_id") or "")
            record_fields = record.get("fields") or {}
            if not record_id:
                continue
            record_date = normalize_date(record_fields.get(date_field))
            if not record_date or record_date < start_date or record_date > end_date:
                preserved_outside_scope += 1
                continue
            scanned_records += 1
            unique_key = scalar_text(record_fields.get(F_UNIQUE_KEY))
            fallback_key = tuple(scalar_text(record_fields.get(field)) for field in fallback_match_fields)
            if unique_key in source_unique_keys or (all(fallback_key) and fallback_key in source_fallback_keys):
                continue
            stale_record_ids.append(record_id)
            if len(stale_samples) < 20:
                stale_samples.append(
                    {
                        "record_id": record_id,
                        F_UNIQUE_KEY: unique_key,
                        "fallback_key": list(fallback_key),
                    }
                )

        for chunk in chunks([{"record_id": record_id} for record_id in stale_record_ids], 500):
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_delete",
                {"records": [item["record_id"] for item in chunk]},
            )
        return {
            "status": "complete",
            "source_rows": len(source_rows),
            "start_date": start_date,
            "end_date": end_date,
            "scanned_records": scanned_records,
            "preserved_outside_scope": preserved_outside_scope,
            "deleted_records": len(stale_record_ids),
            "sample_deleted_records": stale_samples,
        }

    def count_records_in_date_range(
        self,
        table_id: str,
        *,
        date_field: str,
        start_date: str,
        end_date: str,
    ) -> dict[str, int]:
        in_scope = 0
        outside_scope = 0
        for record in self.iter_records(table_id, [date_field]):
            record_date = normalize_date((record.get("fields") or {}).get(date_field))
            if record_date and start_date <= record_date <= end_date:
                in_scope += 1
            else:
                outside_scope += 1
        return {"in_scope": in_scope, "outside_scope": outside_scope, "total": in_scope + outside_scope}


def is_retryable_feishu_response(status_code: int, body: dict[str, Any]) -> bool:
    return status_code in {429, 500, 502, 503, 504} or body.get("code") in {1254607, 1255002}


def discover_daily_files(batch_dir: Path) -> dict[str, dict[str, list[Path]]]:
    result: dict[str, dict[str, list[Path]]] = {platform: {"orders": [], "ads": [], "influencer": []} for platform in PLATFORMS}
    if not batch_dir.exists():
        raise FileNotFoundError(batch_dir)
    for platform_dir in batch_dir.iterdir():
        if not platform_dir.is_dir():
            continue
        platform = normalize_platform(platform_dir.name)
        if platform not in result:
            continue
        for path in platform_dir.rglob("*"):
            if not path.is_file() or path.suffix.lower() not in {".csv", ".xlsx", ".xls"}:
                continue
            if is_temporary_export_file(path):
                continue
            kind = classify_file(platform, path)
            if kind:
                result[platform][kind].append(path)
    for platform in result:
        for kind in result[platform]:
            result[platform][kind].sort(
                key=lambda item: (item.stat().st_mtime_ns, str(item.resolve()).casefold()),
                reverse=True,
            )
    return result


def is_temporary_export_file(path: Path) -> bool:
    return path.name.startswith(("~$", ".~"))


def classify_file(platform: str, path: Path) -> str | None:
    headers = peek_headers(path)
    if platform == "抖音":
        header_kind = classify_douyin_headers(headers)
        if header_kind:
            return header_kind
        name = path.name.lower()
        if any(token in name for token in ("达人佣金", "佣金", "daren", "commission")):
            return "influencer"
        if any(token in name for token in ("投流", "推广", "全域推广", "商品推广", "分天数据")):
            return "ads"
        return None
    header_kind = classify_headers(headers)
    if header_kind:
        return header_kind
    name = path.name.lower()
    if any(token in name for token in ("投流", "推广", "全域推广", "商品推广", "分天数据")):
        return "ads"
    if any(token in name for token in ("order", "订单", "exportorderlist", "orders_export")):
        return "orders"
    return "orders" if platform in {"拼多多"} and path.suffix.lower() == ".csv" else None


def peek_headers(path: Path) -> set[str]:
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            for encoding in ("utf-8-sig", "gb18030", "utf-16"):
                try:
                    with path.open("r", encoding=encoding, newline="") as fh:
                        return {clean_header(value) for value in next(csv.reader(fh), [])}
                except UnicodeError:
                    continue
            return set()
        if suffix == ".xls":
            frame = pd.read_excel(path, engine="xlrd", header=None, nrows=20)
            rows = [tuple(row) for row in frame.itertuples(index=False, name=None)]
        else:
            workbook = load_workbook(path, data_only=True, read_only=True)
            sheet = preferred_worksheet(workbook)
            reset_worksheet_dimensions(sheet)
            rows = [tuple(row) for row in sheet.iter_rows(values_only=True, max_row=20)]
        if not rows:
            return set()
        header_index = find_header_index(rows)
        return {clean_header(value) for value in rows[header_index]}
    except Exception:
        return set()


def classify_headers(headers: set[str]) -> str | None:
    if not headers:
        return None
    douyin_commission_signals = {
        "订单id",
        "作者账号",
        "抖音/火山号",
        "支付金额",
        "佣金率",
        "预估佣金支出",
        "结算金额",
        "实际佣金支出",
        "流量来源",
        "APP渠道",
    }
    if "订单id" in headers and len(headers & douyin_commission_signals) >= 4:
        return "influencer"
    order_signals = {
        "订单号",
        "订单编号",
        "主订单编号",
        "子订单编号",
        "订单状态",
        "订单下单时间",
        "订单创建时间",
    }
    ad_signals = {
        "日期",
        "投放日期",
        "花费",
        "推广花费(元)",
        "实际消耗",
        "点击量",
        "展现量",
        "曝光量",
        "投入产出比",
        "ROI",
    }
    if len(headers & order_signals) >= 2 or (headers & {"订单号", "订单编号", "主订单编号"} and headers & {"订单状态"}):
        return "orders"
    if len(headers & ad_signals) >= 2 or ("日期" in headers and headers & {"花费", "点击量", "展现量", "投入产出比"}):
        return "ads"
    return None


def classify_douyin_headers(headers: set[str]) -> str | None:
    if not headers:
        return None
    if DOUYIN_ORDER_DETAIL_REQUIRED_HEADER in headers:
        return "orders"
    return classify_headers(headers) if classify_headers(headers) != "orders" else None


def load_tabular(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_csv(path)
    if suffix == ".xls":
        return load_xls(path)
    return load_xlsx(path)


def load_csv(path: Path) -> list[dict[str, Any]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            with path.open("r", encoding=encoding, newline="") as fh:
                reader = csv.DictReader(fh)
                return [
                    {clean_header(key): value for key, value in row.items()}
                    for row in reader
                    if any(value not in (None, "") for value in row.values())
                ]
        except UnicodeError as exc:
            last_error = exc
    raise UnicodeError(f"Cannot decode CSV file {path}: {last_error}")


def load_xlsx(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = preferred_worksheet(workbook)
    reset_worksheet_dimensions(sheet)
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header_index = find_header_index(rows)
    headers = [clean_header(value) for value in rows[header_index]]
    result: list[dict[str, Any]] = []
    for values in rows[header_index + 1 :]:
        row = dict(zip(headers, values))
        if any(value not in (None, "") for value in row.values()):
            result.append(row)
    return result


def preferred_worksheet(workbook: Any) -> Any:
    for sheet_name in workbook.sheetnames:
        if clean_text(sheet_name).casefold() == "export":
            return workbook[sheet_name]
    return workbook.active


def reset_worksheet_dimensions(sheet: Any) -> None:
    reset = getattr(sheet, "reset_dimensions", None)
    if callable(reset):
        reset()


def load_xls(path: Path) -> list[dict[str, Any]]:
    frame = pd.read_excel(path, engine="xlrd", header=None)
    rows = [tuple(row) for row in frame.itertuples(index=False, name=None)]
    header_index = find_header_index(rows)
    headers = [clean_header(value) for value in rows[header_index]]
    result: list[dict[str, Any]] = []
    for values in rows[header_index + 1 :]:
        row = dict(zip(headers, values))
        if any(value not in (None, "") and not is_nan(value) for value in row.values()):
            result.append(row)
    return result


def find_header_index(rows: list[tuple[Any, ...]] | list[Any]) -> int:
    for index, row in enumerate(rows[:20]):
        headers = {clean_header(value) for value in row}
        if any(key in headers for key in ("订单号", "订单编号", "主订单编号", "支付方式", "日期", "商品", "商品ID")):
            return index
    return 0


def parse_order_rows(platform: str, path: Path) -> list[dict[str, Any]]:
    source_rows = load_tabular(path)
    if platform == "天猫":
        rows = [tmall_order_row(row, path) for row in source_rows]
    elif platform == "抖音":
        rows = [douyin_order_row(row, path) for row in source_rows]
    elif platform == "拼多多":
        rows = [pdd_order_row(row, path) for row in source_rows]
    elif platform == "视频号":
        rows = [wechat_order_row(row, path) for row in source_rows]
    else:
        rows = []
    return collapse_order_rows([row for row in rows if row])


def add_product_breakdown_to_orders(rows: list[dict[str, Any]], rules: list[Any]) -> list[dict[str, Any]]:
    if not rules:
        return rows
    for row in rows:
        actual_quantity = number_value(row.get(F_QUANTITY)) or 0
        valid_sales = effective_sales_amount(row.get(F_PAID_AMOUNT), row.get(F_REFUND_AMOUNT)) if actual_quantity > 0 else 0
        row.update(
            product_breakdown_values(
                rules,
                product_name=row.get(F_PRODUCT_NAME),
                product_code=row.get(F_PRODUCT_CODE) or extract_product_code_from_raw(row.get(F_RAW)),
                actual_quantity=actual_quantity,
                valid_sales=valid_sales,
            )
        )
        row.pop(INTERNAL_PRODUCT_BREAKDOWN_QUANTITY, None)
    return rows


def tmall_order_row(row: dict[str, Any], path: Path) -> dict[str, Any] | None:
    order_no = clean_text(first_present(row, "订单编号", "订单号"))
    if not order_no:
        return None
    quantity = number_value(first_present(row, "宝贝总数量", "数量", "商品数量"))
    refund = number_value(first_present(row, "退款金额")) or 0
    paid = number_value(first_present(row, "买家实付金额", "实收款", "总金额"))
    if refund > 0 and paid is not None and paid < refund:
        paid = round(paid + refund, 2)
    unit_price = ratio(paid if paid not in (None, 0) else number_value(first_present(row, "买家应付货款", "总金额")), quantity)
    return order_base(
        platform="天猫",
        source="天猫订单Excel导入",
        order_no=order_no,
        created_at=order_created_at("天猫", row, order_no),
        product=clean_text(first_present(row, "商品标题", "商品名称")),
        product_code=clean_text(first_present(row, "商品编码", "商家编码", "商品编号", "商品ID")),
        quantity=quantity,
        unit_price=unit_price,
        paid_amount=paid,
        refund_amount=refund,
        freight=number_value(first_present(row, "买家应付邮费")) or 0,
        platform_fee=number_value(first_present(row, "卖家服务费")) or 0,
        fulfill_status=clean_text(first_present(row, "订单关闭原因")),
        trade_status=clean_text(first_present(row, "订单状态")),
        operation="天猫订单Excel导入",
        source_file=path,
        raw=row,
        shop_id=clean_text(first_present(row, "店铺ID")),
        shop_name=clean_text(first_present(row, "店铺名称")) or "天猫",
    )


def douyin_order_row(row: dict[str, Any], path: Path) -> dict[str, Any] | None:
    order_no = clean_text(first_present(row, "主订单编号", "子订单编号", "订单id", "订单ID", "订单号"))
    if not order_no:
        return None
    quantity = number_value(first_present(row, "商品数量", "数量"))
    paid = number_value(first_present(row, "订单应付金额", "支付金额", "实收款"))
    status = clean_text(first_present(row, "订单状态"))
    aftersale = clean_text(first_present(row, "售后状态"))
    refund = refund_from_status(number_value(first_present(row, "退款金额", "已退款金额")), paid, f"{status}/{aftersale}")
    return order_base(
        platform="抖音",
        source="抖店订单CSV导入",
        order_no=order_no,
        created_at=order_created_at("抖音", row, order_no),
        product=clean_text(first_present(row, "选购商品", "商品名称")),
        product_code=clean_text(first_present(row, "商品编码", "商品编码(平台)", "商家编码", "商品ID")),
        quantity=quantity,
        unit_price=number_value(first_present(row, "商品单价", "单价")),
        paid_amount=paid,
        refund_amount=refund,
        freight=number_value(first_present(row, "运费")) or 0,
        platform_fee=number_value(first_present(row, "手续费")) or 0,
        fulfill_status=aftersale,
        trade_status=status,
        operation="抖店订单CSV导入",
        source_file=path,
        raw=row,
        shop_id=clean_text(first_present(row, "所属门店ID", "店铺id", "店铺ID")),
        shop_name="抖音",
    )


def fetch_jushuitan_douyin_order_rows(settings: Any, selected_dates: set[str] | None = None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not settings.jushuitan_douyin_shop_id:
        raise RuntimeError("Missing JUSHUITAN_SHOP_ID_DOUYIN for Douyin Jushuitan fallback")
    missing = [
        name
        for name, value in (
            ("JUSHUITAN_PARTNER_ID", settings.jushuitan_partner_id),
            ("JUSHUITAN_PARTNER_KEY", settings.jushuitan_partner_key),
            ("JUSHUITAN_TOKEN", settings.jushuitan_token),
        )
        if not value
    ]
    if missing:
        raise RuntimeError(f"Missing Jushuitan credentials for Douyin fallback: {', '.join(missing)}")

    session = requests.Session()
    session.trust_env = jushuitan_trust_env_proxy()
    fetched_at = datetime.now()
    start_at, end_at, query_info = jushuitan_query_window(selected_dates, fetched_at)
    api_settings = replace(
        settings,
        shop_id=settings.jushuitan_douyin_shop_id,
        shop_name="抖音",
        shop_platform="doudian",
        order_source="jushuitan",
        use_mock_collectors=False,
        jushuitan_shop_ids=settings.jushuitan_douyin_shop_id,
    )

    def transport(method: str, url: str, params: dict[str, Any] | None, body: dict[str, Any] | None) -> dict[str, Any]:
        if method != "POST_JSON":
            raise ValueError(f"Unsupported Jushuitan request method: {method}")
        response = session.post(url, params=params, json=body, timeout=30)
        response.raise_for_status()
        payload = response.json()
        if not isinstance(payload, dict):
            raise ValueError("Jushuitan API response is not a JSON object")
        return payload

    collector = JushuitanOrderApiCollector(api_settings, transport=transport)
    raw_orders: list[dict[str, Any]] = []
    page_count = 0
    chunk_count = 0
    chunk_start = start_at
    while chunk_start < end_at:
        chunk_end = min(chunk_start + timedelta(days=JUSHUITAN_DOUYIN_CHUNK_DAYS), end_at)
        chunk_count += 1
        page_index = 1
        while True:
            body = collector._request_body(chunk_start, chunk_end, page_index)
            params = jushuitan_public_params(
                partner_id=api_settings.jushuitan_partner_id,
                partner_key=api_settings.jushuitan_partner_key,
                token=api_settings.jushuitan_token,
                method=api_settings.jushuitan_order_query_method,
                ts=int(datetime.now().timestamp()),
            )
            payload = transport("POST_JSON", api_settings.jushuitan_api_url, params, body)
            collector._raise_jushuitan_error(payload)
            page_orders = extract_jushuitan_orders(payload)
            raw_orders.extend(page_orders)
            page_count += 1
            if len(page_orders) < collector.page_size:
                break
            page_index += 1
        chunk_start = chunk_end

    unique_raw_orders = dedupe_jushuitan_orders(raw_orders)
    paid_orders = [order for order in unique_raw_orders if not is_unpaid_jushuitan_order(order)]
    rows = [row for row in (jushuitan_douyin_order_row(order, fetched_at) for order in paid_orders) if row]
    date_filter_applied = bool(selected_dates)
    if selected_dates:
        rows = [row for row in rows if normalize_date(row.get(F_CREATED_AT)) in selected_dates]
    rows = collapse_order_rows(rows)
    return rows, {
        "source": "jushuitan",
        "lookback_days": query_info["lookback_days"],
        "query_window_source": query_info["source"],
        "requested_dates": query_info["requested_dates"],
        "start_at": start_at.strftime("%Y-%m-%d %H:%M:%S"),
        "end_at": end_at.strftime("%Y-%m-%d %H:%M:%S"),
        "chunk_days": JUSHUITAN_DOUYIN_CHUNK_DAYS,
        "chunks": chunk_count,
        "shop_id": settings.jushuitan_douyin_shop_id,
        "pages": page_count,
        "raw_orders": len(raw_orders),
        "unique_raw_orders": len(unique_raw_orders),
        "paid_orders": len(paid_orders),
        "rows": len(rows),
        "date_filter_applied": date_filter_applied,
        "trust_env_proxy": session.trust_env,
    }


def jushuitan_trust_env_proxy() -> bool:
    value = os.getenv("SHOPOPS_JUSHUITAN_TRUST_ENV_PROXY")
    if value is None:
        return True
    return value.strip().lower() not in {"0", "false", "no", "off"}


def jushuitan_query_window(selected_dates: set[str] | None, fetched_at: datetime) -> tuple[datetime, datetime, dict[str, Any]]:
    parsed_dates = []
    for value in selected_dates or set():
        try:
            parsed_dates.append(datetime.strptime(value, "%Y-%m-%d").date())
        except ValueError:
            continue
    if parsed_dates:
        start_date = min(parsed_dates)
        end_date = max(parsed_dates)
        start_at = datetime.combine(start_date, datetime.min.time())
        end_at = datetime.combine(end_date + timedelta(days=1), datetime.min.time())
        if end_at > fetched_at:
            end_at = fetched_at
        if end_at <= start_at:
            end_at = min(fetched_at, start_at + timedelta(days=1))
        return start_at, end_at, {
            "source": "selected_dates",
            "lookback_days": 0,
            "requested_dates": sorted(date.isoformat() for date in parsed_dates),
        }
    return fetched_at - timedelta(days=JUSHUITAN_DOUYIN_LOOKBACK_DAYS), fetched_at, {
        "source": "rolling_lookback",
        "lookback_days": JUSHUITAN_DOUYIN_LOOKBACK_DAYS,
        "requested_dates": [],
    }


def dedupe_jushuitan_orders(raw_orders: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    result: list[dict[str, Any]] = []
    for order in raw_orders:
        key = clean_text(jushuitan_first_value(order, "shop_id", "shopid")) + "|" + clean_text(
            jushuitan_first_value(order, "so_id", "raw_so_id", "outer_so_id", "shop_order_id", "order_id", "o_id")
        )
        if not key.strip("|"):
            key = json.dumps(order, ensure_ascii=False, sort_keys=True, default=str)
        if key in seen:
            continue
        seen.add(key)
        result.append(order)
    return result


def jushuitan_douyin_order_row(raw: dict[str, Any], fetched_at: datetime) -> dict[str, Any] | None:
    order_no = clean_text(jushuitan_first_value(raw, "so_id", "raw_so_id", "outer_so_id", "shop_order_id", "order_id", "o_id"))
    if not order_no:
        return None
    items = [item for item in raw.get("items", []) if isinstance(item, dict)]
    product = join_jushuitan_item_names(items) or clean_text(jushuitan_first_value(raw, "product_name", "name", "title"))
    quantity = sum_jushuitan_item_quantity(items) or number_value(jushuitan_first_value(raw, "qty", "quantity"))
    paid = normalize_jushuitan_amount(raw)
    status = clean_text(jushuitan_first_value(raw, "status", "order_status", "so_status"))
    aftersale = clean_text(jushuitan_first_value(raw, "aftersale_status", "refund_status", "question_type"))
    refund = refund_from_status(
        number_value(jushuitan_first_value(raw, "refund_amount", "refund_fee", "refund")),
        paid,
        f"{status}/{aftersale}",
    )
    created_at = normalize_datetime(jushuitan_first_value(raw, "order_date", "created", "created_at", "shop_order_date", "pay_date", "pay_time"))
    row = order_base(
        platform="抖音",
        source="聚水潭抖音订单API",
        order_no=order_no,
        created_at=created_at,
        product=product,
        product_code=extract_product_code_from_raw(raw),
        quantity=quantity,
        unit_price=ratio(paid, quantity),
        paid_amount=paid,
        refund_amount=refund,
        freight=number_value(jushuitan_first_value(raw, "freight", "freight_amount", "post_fee")) or 0,
        platform_fee=number_value(jushuitan_first_value(raw, "platform_fee", "commission_fee", "service_fee")) or 0,
        fulfill_status=aftersale,
        trade_status=status,
        operation="聚水潭抖音订单API导入",
        source_file=Path(f"jushuitan://orders.single.query/{order_no}"),
        raw=raw,
        shop_id=clean_text(jushuitan_first_value(raw, "shop_id", "shopid")),
        shop_name=clean_text(jushuitan_first_value(raw, "shop_name", "shopname")) or "抖音",
    )
    row[F_FETCHED_AT] = fetched_at.strftime("%Y-%m-%d %H:%M:%S")
    return row


def join_jushuitan_item_names(items: list[dict[str, Any]]) -> str:
    names: list[str] = []
    for item in items:
        name = clean_text(jushuitan_first_value(item, "name", "item_name", "sku_name", "title"))
        if name and name not in names:
            names.append(name)
    return "; ".join(names)


def sum_jushuitan_item_quantity(items: list[dict[str, Any]]) -> float | None:
    total = 0.0
    found = False
    for item in items:
        quantity = number_value(jushuitan_first_value(item, "qty", "quantity", "sale_qty", "num"))
        if quantity is None:
            continue
        total += quantity
        found = True
    return round(total, 6) if found else None


def pdd_order_row(row: dict[str, Any], path: Path) -> dict[str, Any] | None:
    order_no = clean_text(first_present(row, "订单号", "订单编号", "订单ID"))
    if not order_no:
        return None
    quantity = number_value(first_present(row, "商品数量(件)", "商品数量", "数量"))
    paid = number_value(first_present(row, "商家实收金额(元)", "用户实付金额(元)", "商品总价(元)", "实收款"))
    status = clean_text(first_present(row, "订单状态"))
    aftersale = clean_text(first_present(row, "售后状态"))
    refund = refund_from_status(number_value(first_present(row, "退款金额", "已退款金额")), paid, f"{status}/{aftersale}")
    return order_base(
        platform="拼多多",
        source="拼多多订单CSV导入",
        order_no=order_no,
        created_at=order_created_at("拼多多", row, order_no),
        product=clean_text(first_present(row, "商品", "商品名称")),
        product_code=clean_text(first_present(row, "商品编码", "商品编码(平台)", "商家编码", "商品ID")),
        quantity=quantity,
        unit_price=ratio(paid, quantity),
        paid_amount=paid,
        refund_amount=refund,
        freight=number_value(first_present(row, "邮费(元)", "运费")) or 0,
        platform_fee=0,
        fulfill_status="/".join(item for item in (status, aftersale) if item),
        trade_status=status,
        operation="拼多多订单CSV导入",
        source_file=path,
        raw=row,
        shop_name="拼多多",
    )


def wechat_order_row(row: dict[str, Any], path: Path) -> dict[str, Any] | None:
    order_no = clean_text(first_present(row, "订单号"))
    if not order_no:
        return None
    paid = number_value(first_present(row, "订单实际收款金额", "订单实际支付金额", "商品实际价格(总共)"))
    return order_base(
        platform="视频号",
        source="微信小店订单Excel导入",
        order_no=order_no,
        created_at=order_created_at("视频号", row, order_no),
        product=clean_text(first_present(row, "商品名称")),
        product_code=clean_text(first_present(row, "商品编码(平台)", "商品编码", "商家编码", "商品ID")),
        quantity=number_value(first_present(row, "商品数量")),
        unit_price=number_value(first_present(row, "商品实际价格(单件)", "商品价格(单件)")),
        paid_amount=paid,
        refund_amount=number_value(first_present(row, "商品已退款金额")) or 0,
        freight=number_value(first_present(row, "订单运费", "商品平均运费")) or 0,
        platform_fee=number_value(first_present(row, "技术服务费")) or 0,
        other_fee=number_value(first_present(row, "运费险预计投保费用")) or 0,
        fulfill_status=clean_text(first_present(row, "商品发货", "商品售后")),
        trade_status=clean_text(first_present(row, "订单状态")),
        operation="微信小店订单Excel导入",
        source_file=path,
        raw=redact_row(row),
        shop_name="视频号",
    )


def order_base(
    *,
    platform: str,
    source: str,
    order_no: str,
    created_at: str,
    product: str,
    quantity: float | None,
    unit_price: float | None,
    paid_amount: float | None,
    refund_amount: float | None,
    freight: float | None,
    platform_fee: float | None,
    fulfill_status: str,
    trade_status: str,
    operation: str,
    source_file: Path,
    raw: dict[str, Any],
    product_code: str = "",
    product_cost: float | None = 0,
    other_fee: float | None = 0,
    shop_id: str = "",
    shop_name: str = "",
) -> dict[str, Any]:
    product_breakdown_quantity = quantity
    quantity = actual_sold_quantity(
        quantity=quantity,
        product=product,
        unit_price=unit_price,
        refund_amount=refund_amount,
        trade_status=trade_status,
        fulfill_status=fulfill_status,
    )
    return {
        F_UNIQUE_KEY: order_unique_key(platform, order_no),
        F_PLATFORM: platform,
        F_DATA_SOURCE: source,
        F_SHOP_ID: shop_id,
        F_SHOP_NAME: shop_name or platform,
        F_FETCHED_AT: datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        F_ORDER_NO: order_no,
        F_CREATED_AT: created_at,
        F_BUYER_NICK: "",
        F_PRODUCT_NAME: product,
        F_PRODUCT_CODE: product_code,
        F_ACCESSORY_FLAG: "是" if is_accessory_product(product) else "否",
        INTERNAL_PRODUCT_BREAKDOWN_QUANTITY: product_breakdown_quantity,
        F_UNIT_PRICE: unit_price,
        F_QUANTITY: quantity,
        F_FULFILL_STATUS: fulfill_status,
        F_TRADE_STATUS: trade_status,
        F_PAID_AMOUNT: paid_amount,
        F_REFUND_AMOUNT: refund_amount,
        F_PRODUCT_COST: product_cost,
        F_FREIGHT_COST: freight,
        F_PLATFORM_FEE: platform_fee,
        F_OTHER_FEE: other_fee,
        F_OPERATION: operation,
        F_RAW: json.dumps({"source_file": str(source_file), "row": redact_row(raw)}, ensure_ascii=False, sort_keys=True, default=str),
    }


def collapse_order_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = row[F_UNIQUE_KEY]
        current = merged.get(key)
        if current is None:
            merged[key] = dict(row)
            continue
        current[F_PRODUCT_NAME] = join_unique(current.get(F_PRODUCT_NAME), row.get(F_PRODUCT_NAME))
        current[F_PRODUCT_CODE] = join_unique(current.get(F_PRODUCT_CODE), row.get(F_PRODUCT_CODE))
        current[F_ACCESSORY_FLAG] = "是" if current.get(F_ACCESSORY_FLAG) == "是" and row.get(F_ACCESSORY_FLAG) == "是" else "否"
        current[INTERNAL_PRODUCT_BREAKDOWN_QUANTITY] = round(
            (number_value(current.get(INTERNAL_PRODUCT_BREAKDOWN_QUANTITY)) or 0)
            + (number_value(row.get(INTERNAL_PRODUCT_BREAKDOWN_QUANTITY)) or 0),
            2,
        )
        for field in (F_QUANTITY, F_PAID_AMOUNT, F_REFUND_AMOUNT, F_PRODUCT_COST, F_FREIGHT_COST, F_PLATFORM_FEE, F_OTHER_FEE):
            current[field] = round((number_value(current.get(field)) or 0) + (number_value(row.get(field)) or 0), 2)
        current[F_FULFILL_STATUS] = join_unique(current.get(F_FULFILL_STATUS), row.get(F_FULFILL_STATUS), "/")
        current[F_TRADE_STATUS] = join_unique(current.get(F_TRADE_STATUS), row.get(F_TRADE_STATUS), "/")
    return list(merged.values())


def parse_ad_rows(platform: str, path: Path) -> list[dict[str, Any]]:
    source_rows = load_tabular(path)
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for source in source_rows:
        date_text = normalize_date(first_present(source, "日期", "投放日期"))
        if not date_text or date_text in {"全部", "总计"}:
            continue
        grouped[date_text].append(source)
    rows: list[dict[str, Any]] = []
    for date_text, items in sorted(grouped.items()):
        rows.append(ad_row(platform, date_text, items, path))
    return rows


def ad_row(platform: str, date_text: str, items: list[dict[str, Any]], path: Path) -> dict[str, Any]:
    deal_spend = sum_numbers(items, F_DEAL_SPEND)
    total_spend = sum_numbers(items, F_TOTAL_SPEND)
    spend = sum_numbers(items, "花费", "整体消耗", F_DEAL_SPEND, F_TOTAL_SPEND)
    deal_amount = sum_numbers(items, "总成交金额", "整体成交金额", F_TRADE_AMOUNT, "成交金额")
    impressions = sum_numbers(items, "展现量", "曝光量", "整体展示次数")
    clicks = sum_numbers(items, "点击量", "整体点击次数")
    roi_value = ratio(deal_amount, spend)
    pdd_fields = pdd_ad_extra_fields(items) if platform == "拼多多" else {}
    return {
        F_UNIQUE_KEY: ad_unique_key(platform, date_text),
        F_PLATFORM: platform,
        F_DATA_SOURCE: f"{platform}投流文件导入",
        F_SHOP_ID: "",
        F_SHOP_NAME: platform,
        F_FETCHED_AT: datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        F_DATE: date_text,
        F_SPEND: spend,
        F_PROMOTION_SPEND: spend,
        F_ACTUAL_SPEND: spend,
        F_DEAL_AMOUNT: deal_amount,
        F_DEAL_SPEND: deal_spend if deal_spend else None,
        F_TOTAL_SPEND: total_spend if total_spend else None,
        F_TRADE_AMOUNT: deal_amount,
        **pdd_fields,
        F_IMPRESSIONS: impressions,
        F_EXPOSURES: impressions,
        F_CLICKS: clicks,
        F_CLICK_RATE: ratio(clicks, impressions),
        F_CPC: ratio(spend, clicks),
        F_ROI: roi_value,
        F_PLATFORM_ROI: first_number(items, "投入产出比", "整体支付ROI", "实际投产比") or roi_value,
        F_TRUE_ROI: roi_value,
        F_RAW: json.dumps({"source_file": str(path), "rows": items}, ensure_ascii=False, sort_keys=True, default=str),
    }


PDD_AD_SUM_FIELDS = (
    F_NET_TRADE_AMOUNT,
    F_NET_DEAL_COUNT,
    F_SETTLED_TRADE_AMOUNT,
    F_SETTLED_DEAL_COUNT,
    F_DEAL_COUNT,
)

PDD_AD_RATIO_OR_AVERAGE_FIELDS = (
    F_NET_ACTUAL_ROI,
    F_COST_PER_NET_DEAL,
    F_NET_TRADE_AMOUNT_RATE,
    F_NET_DEAL_COUNT_RATE,
    F_AMOUNT_PER_NET_DEAL,
    F_SETTLED_ROI,
    F_REFUND_EXEMPTION_RATE,
    F_REFUND_ORDER_EXEMPTION_RATE,
    F_COST_PER_SETTLED_DEAL,
    F_TRADE_AMOUNT_SETTLEMENT_RATE,
    F_ORDER_SETTLEMENT_RATE,
    F_AMOUNT_PER_SETTLED_DEAL,
    F_COST_PER_DEAL,
    F_AMOUNT_PER_DEAL,
)


def pdd_ad_extra_fields(items: list[dict[str, Any]]) -> dict[str, Any]:
    fields: dict[str, Any] = {}
    for field in PDD_AD_SUM_FIELDS:
        value = sum_numbers(items, field)
        if value:
            fields[field] = value
    for field in PDD_AD_RATIO_OR_AVERAGE_FIELDS:
        value = first_number(items, field)
        if value is not None:
            fields[field] = value
    return fields


def sample_ad_rows(rows: list[dict[str, Any]], limit: int = 10) -> list[dict[str, Any]]:
    sample_fields = [
        F_UNIQUE_KEY,
        F_PLATFORM,
        F_DATE,
        F_SPEND,
        F_PROMOTION_SPEND,
        F_ACTUAL_SPEND,
        F_DEAL_SPEND,
        F_TOTAL_SPEND,
        F_TRADE_AMOUNT,
        F_DEAL_AMOUNT,
        F_DEAL_COUNT,
        F_COST_PER_DEAL,
        F_AMOUNT_PER_DEAL,
        F_NET_TRADE_AMOUNT,
        F_NET_ACTUAL_ROI,
        F_NET_DEAL_COUNT,
        F_SETTLED_TRADE_AMOUNT,
        F_SETTLED_DEAL_COUNT,
        F_IMPRESSIONS,
        F_EXPOSURES,
        F_CLICKS,
        F_CLICK_RATE,
        F_CPC,
        F_ROI,
        F_PLATFORM_ROI,
        F_TRUE_ROI,
    ]
    return [
        {field: row.get(field) for field in sample_fields if row.get(field) not in (None, "")}
        for row in rows[:limit]
    ]


def parse_influencer_rows(platform: str, path: Path) -> list[dict[str, Any]]:
    if platform not in {"抖音", "视频号"}:
        return []
    if platform == "抖音" and classify_file(platform, path) != "influencer":
        return []
    if platform == "抖音":
        rows = doudian_commission_excel_rows(parse_doudian_commission_xlsx(path), path, [])
        for row in rows:
            actual_commission = number_value(row.get(I_ACTUAL_COMMISSION))
            estimated_commission = number_value(row.get(I_ESTIMATED_COMMISSION))
            commission = actual_commission if actual_commission and actual_commission > 0 else estimated_commission
            row.setdefault(I_INFLUENCER_ID, clean_text(row.get("抖音/火山号")))
            row.setdefault(I_INFLUENCER_NICK, clean_text(row.get("作者账号")))
            row.setdefault(I_COMMISSION_RATE, clean_text(row.get(I_COMMISSION_RATE_NUM)))
            row.setdefault(I_COMMISSION, commission)
            row.setdefault(I_COMMISSION_BASIS, "实际佣金支出" if actual_commission and actual_commission > 0 else "预估佣金支出")
        return collapse_influencer_rows(rows)
    rows = []
    for source in load_tabular(path):
        row = douyin_influencer_row(source, path) if platform == "抖音" else wechat_influencer_row(source, path)
        if row:
            rows.append(row)
    return collapse_influencer_rows(rows)


def douyin_influencer_row(row: dict[str, Any], path: Path) -> dict[str, Any] | None:
    order_no = clean_text(first_present(row, "订单id", "主订单编号", "订单ID", "订单号"))
    influencer_id = clean_text(first_present(row, "抖音/火山号", "达人ID"))
    influencer_nick = clean_text(first_present(row, "作者账号", "达人昵称"))
    actual_commission = first_number([row], "实际佣金支出", "达人实际承担优惠金额", "达人优惠")
    estimated_commission = first_number([row], "预估佣金支出")
    commission = actual_commission if actual_commission and actual_commission > 0 else estimated_commission
    if not order_no or not any((influencer_id, influencer_nick, commission)):
        return None
    return influencer_base(
        platform="抖音",
        source="抖音达人佣金Excel导入",
        order_no=order_no,
        created_at=normalize_datetime(first_present(row, "付款时间", "订单提交时间")),
        pay_at=normalize_datetime(first_present(row, "付款时间", "支付完成时间")),
        status=clean_text(first_present(row, "订单状态")),
        product_id=clean_text(first_present(row, "商品id", "商品ID")),
        product_name=clean_text(first_present(row, "商品名称", "选购商品")),
        quantity=number_value(first_present(row, "商品数量")),
        paid_amount=number_value(first_present(row, "支付金额", "订单应付金额")),
        influencer_id=influencer_id,
        influencer_nick=influencer_nick,
        commission_rate=clean_text(first_present(row, "佣金率")),
        commission=commission,
        commission_basis="实际佣金支出" if actual_commission and actual_commission > 0 else "预估佣金支出",
        commission_rate_number=number_value(first_present(row, "佣金率")),
        estimated_commission=estimated_commission,
        actual_commission=actual_commission,
        source_file=path,
        raw=row,
        shop_name=clean_text(first_present(row, "店铺名称")) or "抖音",
    )


def wechat_influencer_row(row: dict[str, Any], path: Path) -> dict[str, Any] | None:
    order_no = clean_text(first_present(row, "订单号"))
    influencer_nick = clean_text(first_present(row, "带货账号昵称"))
    commission = number_value(first_present(row, "带货费用"))
    mode = clean_text(first_present(row, "带货方式", "带货费用渠道"))
    if not order_no or not any((influencer_nick, commission, mode)):
        return None
    return influencer_base(
        platform="视频号",
        source="微信小店订单Excel导入",
        order_no=order_no,
        created_at=normalize_datetime(first_present(row, "订单下单时间")),
        pay_at=normalize_datetime(first_present(row, "支付时间")),
        status=clean_text(first_present(row, "订单状态")),
        product_id=clean_text(first_present(row, "商品编码(平台)", "商品ID")),
        product_name=clean_text(first_present(row, "商品名称")),
        quantity=number_value(first_present(row, "商品数量")),
        paid_amount=number_value(first_present(row, "订单实际支付金额")),
        influencer_id="",
        influencer_nick=influencer_nick,
        commission_rate=clean_text(first_present(row, "带货佣金率")),
        commission=commission,
        commission_basis=clean_text(first_present(row, "带货费用类型")) or "带货费用",
        commission_rate_number=number_value(first_present(row, "带货佣金率")),
        estimated_commission=commission,
        actual_commission=None,
        source_file=path,
        raw=redact_row(row),
        shop_name="视频号",
    )


def influencer_base(
    *,
    platform: str,
    source: str,
    order_no: str,
    created_at: str,
    pay_at: str,
    status: str,
    product_id: str,
    product_name: str,
    quantity: float | None,
    paid_amount: float | None,
    influencer_id: str,
    influencer_nick: str,
    commission_rate: str,
    commission: float | None,
    commission_basis: str,
    source_file: Path,
    raw: dict[str, Any],
    shop_name: str,
    commission_rate_number: float | None = None,
    estimated_commission: float | None = None,
    actual_commission: float | None = None,
) -> dict[str, Any]:
    return {
        F_UNIQUE_KEY: f"{platform}{order_no}",
        F_PLATFORM: platform,
        I_SOURCE: source,
        F_ORDER_NO: order_no,
        I_CREATED_AT: created_at,
        I_PAY_AT: pay_at,
        I_STATUS: status,
        F_PRODUCT_ID: product_id,
        F_PRODUCT_NAME: product_name,
        F_QUANTITY: quantity,
        F_PAID_AMOUNT: paid_amount,
        I_INFLUENCER_ID: influencer_id,
        I_INFLUENCER_NICK: influencer_nick,
        I_COMMISSION_RATE: commission_rate,
        I_COMMISSION: commission,
        I_COMMISSION_BASIS: commission_basis,
        I_COMMISSION_RATE_NUM: commission_rate_number,
        I_ESTIMATED_COMMISSION: estimated_commission,
        I_ACTUAL_COMMISSION: actual_commission,
        F_SHOP_ID: "",
        F_SHOP_NAME: shop_name,
        F_FETCHED_AT: datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        I_SOURCE_FILE: str(source_file),
        F_RAW: json.dumps(redact_row(raw), ensure_ascii=False, sort_keys=True, default=str),
    }


def collapse_influencer_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = row[F_UNIQUE_KEY]
        if key not in merged:
            merged[key] = dict(row)
            continue
        current = merged[key]
        current[F_PRODUCT_NAME] = join_unique(current.get(F_PRODUCT_NAME), row.get(F_PRODUCT_NAME))
        current[F_QUANTITY] = round((number_value(current.get(F_QUANTITY)) or 0) + (number_value(row.get(F_QUANTITY)) or 0), 4)
        current[I_COMMISSION] = round((number_value(current.get(I_COMMISSION)) or 0) + (number_value(row.get(I_COMMISSION)) or 0), 4)
        current[I_ESTIMATED_COMMISSION] = round((number_value(current.get(I_ESTIMATED_COMMISSION)) or 0) + (number_value(row.get(I_ESTIMATED_COMMISSION)) or 0), 4)
        current[I_ACTUAL_COMMISSION] = round((number_value(current.get(I_ACTUAL_COMMISSION)) or 0) + (number_value(row.get(I_ACTUAL_COMMISSION)) or 0), 4)
    return list(merged.values())


def reconcile_source_snapshots(
    snapshots: list[tuple[Path | str, list[dict[str, Any]]]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Union newest-first snapshots while enforcing one source row per unique key."""
    reconciled: dict[str, dict[str, Any]] = {}
    kept_source_by_key: dict[str, str] = {}
    overlapping_rows_ignored = 0
    conflicting_overlaps = 0
    conflict_samples: list[dict[str, Any]] = []
    volatile_fields = {F_FETCHED_AT, F_RAW, I_SOURCE_FILE}
    for source, rows in snapshots:
        source_name = str(source)
        for row in rows:
            unique_key = scalar_text(row.get(F_UNIQUE_KEY))
            if not unique_key:
                raise RuntimeError(f"Source row from {source} is missing {F_UNIQUE_KEY}")
            if unique_key in reconciled:
                overlapping_rows_ignored += 1
                kept = reconciled[unique_key]
                differing_fields = [
                    field
                    for field in sorted((set(kept) | set(row)) - volatile_fields)
                    if not field_value_equal(kept.get(field), row.get(field))
                ]
                if differing_fields:
                    conflicting_overlaps += 1
                    if len(conflict_samples) < 20:
                        conflict_samples.append(
                            {
                                F_UNIQUE_KEY: unique_key,
                                "kept_source": kept_source_by_key[unique_key],
                                "ignored_source": source_name,
                                "differing_fields": differing_fields,
                            }
                        )
                continue
            reconciled[unique_key] = row
            kept_source_by_key[unique_key] = source_name
    return list(reconciled.values()), {
        "snapshot_count": len(snapshots),
        "source_precedence": [str(source) for source, _rows in snapshots],
        "input_rows": sum(len(rows) for _source, rows in snapshots),
        "output_rows": len(reconciled),
        "overlapping_rows_ignored": overlapping_rows_ignored,
        "conflicting_overlaps": conflicting_overlaps,
        "conflict_samples": conflict_samples,
        "precedence": "mtime_ns_then_absolute_path_descending; newest_snapshot_wins",
    }


def _file_snapshot_date_scope(
    platform: str,
    rows: list[dict[str, Any]],
    source_info: dict[str, Any],
) -> tuple[str, str] | None:
    if source_info.get("source_type") != "file_snapshot":
        return None
    source_dates = [normalize_date(row.get(F_CREATED_AT)) for row in rows]
    missing_date_rows = sum(not value for value in source_dates)
    if missing_date_rows:
        raise RuntimeError(
            f"Cannot safely reconcile {platform} file snapshot: {missing_date_rows} source rows have no {F_CREATED_AT}"
        )
    if not source_dates:
        return None
    return min(source_dates), max(source_dates)


def _file_snapshot_impact_dates(
    order_rows_by_platform: dict[str, list[dict[str, Any]]],
    order_source_reconciliation: dict[str, dict[str, Any]],
) -> set[str]:
    impact_dates: set[str] = set()
    for platform, rows in order_rows_by_platform.items():
        scope = _file_snapshot_date_scope(
            platform,
            rows,
            order_source_reconciliation.get(platform) or {},
        )
        if not scope:
            continue
        current = datetime.strptime(scope[0], "%Y-%m-%d").date()
        end = datetime.strptime(scope[1], "%Y-%m-%d").date()
        while current <= end:
            impact_dates.add(current.isoformat())
            current += timedelta(days=1)
    return impact_dates


@contextmanager
def standard_import_lock() -> Iterable[None]:
    lock_path = Path(os.getenv("SHOPOPS_STANDARD_IMPORT_LOCK_PATH", "").strip() or ROOT / ".shopops-standard-import.lock")
    timeout_seconds = max(0, int(os.getenv("SHOPOPS_STANDARD_IMPORT_LOCK_TIMEOUT_SECONDS", "1800")))
    stale_seconds = max(timeout_seconds, int(os.getenv("SHOPOPS_STANDARD_IMPORT_LOCK_STALE_SECONDS", "7200")))
    token = f"{os.getpid()}:{time.time_ns()}"
    deadline = time.monotonic() + timeout_seconds
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    while True:
        try:
            descriptor = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            with os.fdopen(descriptor, "w", encoding="utf-8") as handle:
                handle.write(token)
            break
        except FileExistsError:
            try:
                age_seconds = time.time() - lock_path.stat().st_mtime
                if age_seconds > stale_seconds:
                    lock_path.unlink(missing_ok=True)
                    continue
            except FileNotFoundError:
                continue
            if time.monotonic() >= deadline:
                raise RuntimeError(f"Another standard import still owns lock {lock_path}")
            time.sleep(1)
    try:
        yield
    finally:
        try:
            if lock_path.read_text(encoding="utf-8") == token:
                lock_path.unlink(missing_ok=True)
        except FileNotFoundError:
            pass


def run_import(
    batch_dir: Path,
    dry_run: bool,
    evidence: Path,
    platforms: set[str] | None = None,
    kinds: set[str] | None = None,
    dates: set[str] | None = None,
    filter_ad_dates: bool = False,
    ensure_missing_ad_fields: bool = False,
    order_lookback_days: int = ORDER_ROLLING_LOOKBACK_DAYS,
) -> dict[str, Any]:
    if dry_run:
        return _run_import_unlocked(
            batch_dir=batch_dir,
            dry_run=True,
            evidence=evidence,
            platforms=platforms,
            kinds=kinds,
            dates=dates,
            filter_ad_dates=filter_ad_dates,
            ensure_missing_ad_fields=ensure_missing_ad_fields,
            order_lookback_days=order_lookback_days,
        )
    with standard_import_lock():
        return _run_import_unlocked(
            batch_dir=batch_dir,
            dry_run=False,
            evidence=evidence,
            platforms=platforms,
            kinds=kinds,
            dates=dates,
            filter_ad_dates=filter_ad_dates,
            ensure_missing_ad_fields=ensure_missing_ad_fields,
            order_lookback_days=order_lookback_days,
        )


def _run_import_unlocked(
    batch_dir: Path,
    dry_run: bool,
    evidence: Path,
    platforms: set[str] | None = None,
    kinds: set[str] | None = None,
    dates: set[str] | None = None,
    filter_ad_dates: bool = False,
    ensure_missing_ad_fields: bool = False,
    order_lookback_days: int = ORDER_ROLLING_LOOKBACK_DAYS,
) -> dict[str, Any]:
    _load_dotenv()
    settings = load_settings()
    discovered = discover_daily_files(batch_dir)
    selected_platforms = set(platforms or PLATFORMS)
    selected_kinds = set(kinds or {"orders", "ads", "influencer"})
    selected_dates = {normalize_date(date) for date in dates or set()}
    selected_dates.discard("")
    order_date_window = rolling_date_window(selected_dates, max(0, order_lookback_days))
    influencer_date_window = rolling_date_window(selected_dates, INFLUENCER_ROLLING_LOOKBACK_DAYS)
    order_rows_by_platform: dict[str, list[dict[str, Any]]] = {platform: [] for platform in PLATFORMS if platform in selected_platforms}
    ad_rows: list[dict[str, Any]] = []
    influencer_rows: list[dict[str, Any]] = []
    files: dict[str, Any] = {}
    source_reconciliation: dict[str, dict[str, Any]] = {"orders": {}, "ads": {}, "influencer": {}}

    for platform, platform_files in discovered.items():
        if platform not in selected_platforms:
            continue
        platform_info: dict[str, Any] = {}
        order_snapshots: list[tuple[Path | str, list[dict[str, Any]]]] = []
        for order_file in platform_files["orders"] if "orders" in selected_kinds else []:
            rows = parse_order_rows(platform, order_file)
            if order_date_window:
                rows = [row for row in rows if date_in_window(row.get(F_CREATED_AT), order_date_window)]
            order_snapshots.append((order_file, rows))
            platform_info.setdefault("orders", []).append({"file": str(order_file), "rows": len(rows)})
        if order_snapshots:
            order_rows_by_platform[platform], source_reconciliation["orders"][platform] = reconcile_source_snapshots(order_snapshots)
            source_reconciliation["orders"][platform]["source_type"] = "file_snapshot"
            source_reconciliation["orders"][platform]["scope_policy"] = "source_date_range"
        elif platform == "抖音" and "orders" in selected_kinds:
            fallback_dates = selected_dates if selected_dates and order_lookback_days == 0 else None
            rows, fallback_info = fetch_jushuitan_douyin_order_rows(settings, fallback_dates)
            order_rows_by_platform[platform], source_reconciliation["orders"][platform] = reconcile_source_snapshots(
                [(fallback_info.get("source") or "jushuitan", rows)]
            )
            source_reconciliation["orders"][platform]["source_type"] = "jushuitan_api"
            source_reconciliation["orders"][platform]["scope_policy"] = (
                "selected_dates" if fallback_dates else "upsert_and_verify_returned_rows_only"
            )
            platform_info.setdefault("orders", []).append(fallback_info)

        influencer_snapshots: list[tuple[Path | str, list[dict[str, Any]]]] = []
        for influencer_file in platform_files["influencer"] if "influencer" in selected_kinds else []:
            rows = parse_influencer_rows(platform, influencer_file)
            if influencer_date_window:
                rows = [
                    row
                    for row in rows
                    if date_in_window(row.get(I_CREATED_AT), influencer_date_window)
                    or date_in_window(row.get(I_PAY_AT), influencer_date_window)
                ]
            influencer_snapshots.append((influencer_file, rows))
            platform_info.setdefault("influencer", []).append({"file": str(influencer_file), "rows": len(rows)})
        if influencer_snapshots:
            platform_influencer_rows, source_reconciliation["influencer"][platform] = reconcile_source_snapshots(influencer_snapshots)
            influencer_rows.extend(platform_influencer_rows)

        ad_snapshots: list[tuple[Path | str, list[dict[str, Any]]]] = []
        for ad_file in platform_files["ads"] if "ads" in selected_kinds else []:
            rows = parse_ad_rows(platform, ad_file)
            if selected_dates and filter_ad_dates:
                rows = [row for row in rows if row.get(F_DATE) in selected_dates]
            ad_snapshots.append((ad_file, rows))
            platform_info.setdefault("ads", []).append({"file": str(ad_file), "rows": len(rows)})
        if ad_snapshots:
            platform_ad_rows, source_reconciliation["ads"][platform] = reconcile_source_snapshots(ad_snapshots)
            ad_rows.extend(platform_ad_rows)
        files[platform] = platform_info

    impact_dates = set(selected_dates)
    impact_dates.update(
        normalize_date(row.get(F_CREATED_AT))
        for rows in order_rows_by_platform.values()
        for row in rows
    )
    impact_dates.update(normalize_date(row.get(F_DATE)) for row in ad_rows)
    impact_dates.update(
        normalize_date(row.get(I_PAY_AT) or row.get(I_CREATED_AT))
        for row in influencer_rows
    )
    impact_dates.update(
        _file_snapshot_impact_dates(
            order_rows_by_platform,
            source_reconciliation["orders"],
        )
    )
    impact_dates.discard("")

    summary: dict[str, Any] = {
        "status": "dry_run" if dry_run else "started",
        "batch_dir": str(batch_dir),
        "feishu_base_url": f"https://my.feishu.cn/base/{settings.shopops_data_center_app_token or settings.feishu_app_token}",
        "field_policy": (
            f"create missing ad fields only when explicitly requested; existing records update changed cells only; file-based order snapshots reconcile exactly inside their source date range while preserving all records outside that range; exact-date Jushuitan imports prune only selected dates; orders use a rolling {max(0, order_lookback_days)}-day update window for this run; influencer commissions default to a rolling 90-day update window; orders may update order/import fields and product breakdown fields"
            if ensure_missing_ad_fields
            else f"existing Feishu fields only; never create or update table fields during daily import; existing records update changed cells only; file-based order snapshots reconcile exactly inside their source date range while preserving all records outside that range; exact-date Jushuitan imports prune only selected dates; orders use a rolling {max(0, order_lookback_days)}-day update window for this run; influencer commissions default to a rolling 90-day update window; orders may update order/import fields and product breakdown fields"
        ),
        "platform_filter": sorted(selected_platforms),
        "kind_filter": sorted(selected_kinds),
        "date_filter": sorted(selected_dates),
        "impact_dates": sorted(impact_dates),
        "order_date_window": order_date_window,
        "influencer_date_window": influencer_date_window,
        "ad_date_filter_applied": bool(selected_dates and filter_ad_dates),
        "unique_rules": {
            "orders": "platform_code + '_' + order_no; fallback match by order_no",
            "ads": "ads_platform_code_yyyy-mm-dd; fallback match by platform + date",
            "influencer": "Douyin influencer data must come from an explicit Douyin commission Excel, never from Douyin order exports; fallback match by platform + order_no",
        },
        "douyin_order_source_rule": f"Douyin order Excel/CSV is accepted only when its headers include 支付方式; otherwise orders fall back to Jushuitan API. File-based order imports use the configured {max(0, order_lookback_days)}-day order window; the Jushuitan fallback applies the selected date filter when the order window is 0 days.",
        "files": files,
        "source_reconciliation": source_reconciliation,
        "order_counts": {platform: len(rows) for platform, rows in order_rows_by_platform.items()},
        "ad_count": len(ad_rows),
        "influencer_count": len(influencer_rows),
        "ad_dates": sorted({row[F_DATE] for row in ad_rows}),
        "sample_ad_rows": sample_ad_rows(ad_rows),
        "accessory_counts": {
            platform: sum(1 for row in rows if row.get(F_ACCESSORY_FLAG) == "是")
            for platform, rows in order_rows_by_platform.items()
        },
        "sample_accessory_order_keys": {
            platform: [row[F_UNIQUE_KEY] for row in rows if row.get(F_ACCESSORY_FLAG) == "是"][:10]
            for platform, rows in order_rows_by_platform.items()
        },
        "sample_order_keys": {platform: [row[F_UNIQUE_KEY] for row in rows[:10]] for platform, rows in order_rows_by_platform.items()},
    }
    if dry_run:
        write_import_evidence(evidence, summary)
        return summary

    summary["progress"] = []
    progress_lock = threading.Lock()

    def progress(stage: str, detail: dict[str, Any] | None = None) -> None:
        with progress_lock:
            write_import_progress(evidence, summary, stage, detail)

    progress("parsed_source_files")
    client = FeishuDailyClient()
    progress("feishu_client_ready")
    product_table_id = os.getenv("SHOPOPS_PRODUCT_CATALOG_TABLE_ID", DEFAULT_PRODUCT_CATALOG_TABLE_ID).strip()
    progress("loading_product_rules", {"table_id": product_table_id})
    product_rules = client.product_rules(product_table_id)
    product_fields = product_field_names(product_rules)
    progress("loaded_product_rules", {"product_field_count": len(product_fields)})
    for platform, rows in order_rows_by_platform.items():
        order_rows_by_platform[platform] = add_product_breakdown_to_orders(rows, product_rules)
    writes: dict[str, Any] = {"orders": {}, "ads": {}, "influencer": {}}
    field_preflight: dict[str, Any] = {"orders": {}, "ads": {}, "influencer": {}}
    influencer_table_id = ""
    progress("field_preflight_started")
    for platform, rows in order_rows_by_platform.items():
        if not rows:
            continue
        table_id = os.getenv(ORDER_TABLE_ENV[platform], "").strip()
        if not table_id:
            raise RuntimeError(f"Missing {ORDER_TABLE_ENV[platform]}")
        missing_fields = missing_row_fields(
            client.field_names(table_id),
            rows,
            [F_UNIQUE_KEY, F_ORDER_NO, F_ACCESSORY_FLAG, *product_fields],
        )
        field_preflight["orders"][platform] = {"table_id": table_id, "missing_fields": missing_fields}
        if missing_fields:
            raise RuntimeError(f"Target order table {table_id} is missing existing fields required by this import: {missing_fields}")
        progress("field_preflight_orders_done", {"platform": platform, "row_count": len(rows)})
    if ad_rows:
        if not settings.shopops_ad_table_id:
            raise RuntimeError("Missing SHOPOPS_AD_TABLE_ID")
        missing_fields = missing_row_fields(client.field_names(settings.shopops_ad_table_id), ad_rows, [F_UNIQUE_KEY, F_PLATFORM, F_DATE])
        field_preflight["ads"] = {"table_id": settings.shopops_ad_table_id, "missing_fields": missing_fields}
        if missing_fields:
            raise RuntimeError(f"Target ad table {settings.shopops_ad_table_id} is missing existing fields required by this import: {missing_fields}")
        progress("field_preflight_ads_done", {"row_count": len(ad_rows)})
    if influencer_rows:
        influencer_table_id = os.getenv("SHOPOPS_DOUYIN_INFLUENCER_EXCEL_TABLE_ID", "").strip() or settings.table_douyin_influencer_commission
        if not influencer_table_id or not influencer_table_id.startswith("tbl"):
            raise RuntimeError("Missing SHOPOPS_DOUYIN_INFLUENCER_EXCEL_TABLE_ID or FEISHU_TABLE_DOUYIN_INFLUENCER_COMMISSION")
        missing_fields = missing_row_fields(client.field_names(influencer_table_id), influencer_rows, [F_UNIQUE_KEY, F_ORDER_NO])
        field_preflight["influencer"] = {"table_id": influencer_table_id, "missing_fields": missing_fields}
        if missing_fields:
            raise RuntimeError(f"Target influencer table {influencer_table_id} is missing existing fields required by this import: {missing_fields}")
        progress("field_preflight_influencer_done", {"row_count": len(influencer_rows)})
    summary["field_preflight"] = field_preflight
    progress("field_preflight_complete")

    def sync_order_platform(platform: str, rows: list[dict[str, Any]]) -> tuple[str, dict[str, Any]]:
        table_id = os.getenv(ORDER_TABLE_ENV[platform], "").strip()
        if not table_id:
            raise RuntimeError(f"Missing {ORDER_TABLE_ENV[platform]}")
        worker_client = FeishuDailyClient()
        progress("upsert_orders_started", {"platform": platform, "row_count": len(rows)})
        result = worker_client.upsert_rows(
            table_id=table_id,
            rows=rows,
            required_fields=[F_UNIQUE_KEY, F_ORDER_NO, F_ACCESSORY_FLAG, *product_fields],
            fallback_match_fields=(F_ORDER_NO,),
            allow_partial_fields=False,
            update_existing_fields={*ORDER_UPDATE_FIELDS, *product_fields},
            clear_empty_fields={*ORDER_UPDATE_FIELDS, *product_fields},
        )
        result["product_field_actions"] = {
            field: "validated_existing_no_field_changes" for field in product_fields
        }
        progress("upsert_orders_done", {"platform": platform, **result})
        source_info = source_reconciliation["orders"].get(platform) or {}
        source_type = source_info.get("source_type")
        if source_type == "file_snapshot":
            snapshot_scope = _file_snapshot_date_scope(platform, rows, source_info)
            if snapshot_scope is None:
                raise RuntimeError(f"Cannot safely reconcile empty {platform} file snapshot")
            snapshot_start_date, snapshot_end_date = snapshot_scope
            result["snapshot_scope"] = {
                "source_type": source_type,
                "date_field": F_CREATED_AT,
                "start_date": snapshot_start_date,
                "end_date": snapshot_end_date,
                "source_rows": len(rows),
                "outside_scope_policy": "preserve",
            }
            progress(
                "prune_stale_orders_started",
                {
                    "platform": platform,
                    "start_date": snapshot_start_date,
                    "end_date": snapshot_end_date,
                    "outside_scope_policy": "preserve",
                },
            )
            result["prune_stale_records"] = worker_client.prune_records_to_snapshot(
                table_id,
                source_rows=rows,
                fallback_match_fields=(F_ORDER_NO,),
                date_field=F_CREATED_AT,
                start_date=snapshot_start_date,
                end_date=snapshot_end_date,
            )
            progress(
                "prune_stale_orders_done",
                {"platform": platform, **result["prune_stale_records"]},
            )
        elif source_type == "jushuitan_api" and selected_dates and order_lookback_days == 0:
            progress("prune_stale_orders_started", {"platform": platform, "dates": sorted(selected_dates)})
            result["prune_stale_records"] = worker_client.prune_order_records_for_dates(
                table_id,
                source_rows=rows,
                dates=selected_dates,
            )
            progress(
                "prune_stale_orders_done",
                {"platform": platform, **result["prune_stale_records"]},
            )
        progress("readback_orders_started", {"platform": platform})
        verification = worker_client.verify_rows_by_unique_key(
            table_id,
            rows,
            compare_fields={*ORDER_UPDATE_FIELDS, *product_fields},
            fallback_match_fields=(F_ORDER_NO,),
            compare_empty_fields=True,
        )
        if verification["status"] != "success":
            progress("repair_orders_started", {"platform": platform, "verification": verification})
            result["repair_retry"] = worker_client.upsert_rows(
                table_id=table_id,
                rows=rows,
                required_fields=[F_UNIQUE_KEY, F_ORDER_NO, F_ACCESSORY_FLAG, *product_fields],
                fallback_match_fields=(F_ORDER_NO,),
                allow_partial_fields=False,
                update_existing_fields={*ORDER_UPDATE_FIELDS, *product_fields},
                clear_empty_fields={*ORDER_UPDATE_FIELDS, *product_fields},
            )
            verification = worker_client.verify_rows_by_unique_key(
                table_id,
                rows,
                compare_fields={*ORDER_UPDATE_FIELDS, *product_fields},
                fallback_match_fields=(F_ORDER_NO,),
                compare_empty_fields=True,
            )
            progress("repair_orders_done", {"platform": platform, "verification": verification})
        if source_type == "file_snapshot":
            scope = result["snapshot_scope"]
            scoped_count = worker_client.count_records_in_date_range(
                table_id,
                date_field=F_CREATED_AT,
                start_date=scope["start_date"],
                end_date=scope["end_date"],
            )
            result["snapshot_count_verification"] = {
                **scoped_count,
                "expected_in_scope": len(rows),
                "status": "success" if scoped_count["in_scope"] == len(rows) else "mismatch",
                "outside_scope_policy": "preserve",
            }
            if result["snapshot_count_verification"]["status"] != "success":
                verification = {
                    **verification,
                    "status": "mismatch",
                    "snapshot_count_verification": result["snapshot_count_verification"],
                }
        result["blank_identity_cleanup"] = worker_client.delete_blank_identity_records(
            table_id,
            key_fields=(F_ORDER_NO,),
            content_fields=(F_PLATFORM, F_DATA_SOURCE, F_CREATED_AT, F_PAID_AMOUNT, F_RAW),
        )
        result["global_uniqueness_verification"] = worker_client.verify_unique_identities(
            table_id,
            key_fields=(F_ORDER_NO,),
        )
        if result["global_uniqueness_verification"]["status"] != "success":
            verification = {
                **verification,
                "status": "mismatch",
                "global_uniqueness_verification": result["global_uniqueness_verification"],
            }
        result["verification"] = verification
        result["readback_count"] = verification["readback_count"]
        result["missing_unique_keys"] = verification["missing_unique_keys"]
        result["duplicate_unique_keys"] = verification["duplicate_unique_keys"]
        result["mismatched_rows"] = verification["mismatched_rows"]
        progress("readback_orders_done", {"platform": platform, **verification})
        return platform, result

    order_tasks = [(platform, rows) for platform, rows in order_rows_by_platform.items() if rows]
    order_results: dict[str, dict[str, Any]] = {}
    if len(order_tasks) > 1:
        with ThreadPoolExecutor(max_workers=min(len(order_tasks), 4), thread_name_prefix="feishu-orders") as executor:
            future_map = {executor.submit(sync_order_platform, platform, rows): platform for platform, rows in order_tasks}
            for future in as_completed(future_map):
                platform, result = future.result()
                order_results[platform] = result
    else:
        for platform, rows in order_tasks:
            platform, result = sync_order_platform(platform, rows)
            order_results[platform] = result
    for platform, _rows in order_tasks:
        writes["orders"][platform] = order_results[platform]

    if ad_rows:
        if not settings.shopops_ad_table_id:
            raise RuntimeError("Missing SHOPOPS_AD_TABLE_ID")
        created_ad_fields = []
        if ensure_missing_ad_fields:
            created_ad_fields = client.ensure_missing_fields_for_rows(settings.shopops_ad_table_id, ad_rows, AD_FIELD_TYPES)
        progress("upsert_ads_started", {"row_count": len(ad_rows)})
        writes["ads"] = client.upsert_rows(
            table_id=settings.shopops_ad_table_id,
            rows=ad_rows,
            required_fields=[F_UNIQUE_KEY, F_PLATFORM, F_DATE],
            fallback_match_fields=(F_PLATFORM, F_DATE),
            allow_partial_fields=False,
        )
        writes["ads"]["created_missing_fields"] = created_ad_fields
        progress("upsert_ads_done", writes["ads"])
        progress("canonicalize_ads_started")
        writes["ads"]["canonicalize_unique_keys"] = client.canonicalize_ad_unique_keys(settings.shopops_ad_table_id)
        progress("canonicalize_ads_done", writes["ads"]["canonicalize_unique_keys"])
        progress("readback_ads_started")
        verification = client.verify_rows_by_unique_key(
            settings.shopops_ad_table_id,
            ad_rows,
            fallback_match_fields=(F_PLATFORM, F_DATE),
        )
        if verification["status"] != "success":
            progress("repair_ads_started", {"verification": verification})
            writes["ads"]["repair_retry"] = client.upsert_rows(
                table_id=settings.shopops_ad_table_id,
                rows=ad_rows,
                required_fields=[F_UNIQUE_KEY, F_PLATFORM, F_DATE],
                fallback_match_fields=(F_PLATFORM, F_DATE),
                allow_partial_fields=False,
            )
            verification = client.verify_rows_by_unique_key(
                settings.shopops_ad_table_id,
                ad_rows,
                fallback_match_fields=(F_PLATFORM, F_DATE),
            )
            progress("repair_ads_done", {"verification": verification})
        writes["ads"]["verification"] = verification
        writes["ads"]["readback_count"] = verification["readback_count"]
        writes["ads"]["missing_unique_keys"] = verification["missing_unique_keys"]
        writes["ads"]["duplicate_unique_keys"] = verification["duplicate_unique_keys"]
        writes["ads"]["mismatched_rows"] = verification["mismatched_rows"]
        progress("readback_ads_done", verification)

    if influencer_rows:
        table_id = influencer_table_id
        progress("upsert_influencer_started", {"row_count": len(influencer_rows)})
        writes["influencer"] = client.upsert_rows(
            table_id=table_id,
            rows=influencer_rows,
            required_fields=[F_UNIQUE_KEY, F_ORDER_NO],
            fallback_match_fields=(F_PLATFORM, F_ORDER_NO),
            allow_partial_fields=False,
        )
        writes["influencer"]["dedupe_policy"] = {
            "action": "repair_matching_incoming_keys",
            "identity": [F_PLATFORM, F_ORDER_NO],
            "deleted_duplicate_records": writes["influencer"]["deleted_duplicate_records"],
        }
        progress("upsert_influencer_done", writes["influencer"])
        progress("readback_influencer_started")
        verification = client.verify_rows_by_unique_key(
            table_id,
            influencer_rows,
            fallback_match_fields=(F_PLATFORM, F_ORDER_NO),
        )
        if verification["status"] != "success":
            progress("repair_influencer_started", {"verification": verification})
            writes["influencer"]["repair_retry"] = client.upsert_rows(
                table_id=table_id,
                rows=influencer_rows,
                required_fields=[F_UNIQUE_KEY, F_ORDER_NO],
                fallback_match_fields=(F_PLATFORM, F_ORDER_NO),
                allow_partial_fields=False,
            )
            verification = client.verify_rows_by_unique_key(
                table_id,
                influencer_rows,
                fallback_match_fields=(F_PLATFORM, F_ORDER_NO),
            )
            progress("repair_influencer_done", {"verification": verification})
        writes["influencer"]["blank_identity_cleanup"] = client.delete_blank_identity_records(
            table_id,
            key_fields=(F_PLATFORM, F_ORDER_NO),
            content_fields=(F_PLATFORM, I_SOURCE, I_CREATED_AT, F_PAID_AMOUNT, I_INFLUENCER_ID, F_RAW),
        )
        writes["influencer"]["global_uniqueness_verification"] = client.verify_unique_identities(
            table_id,
            key_fields=(F_PLATFORM, F_ORDER_NO),
        )
        if writes["influencer"]["global_uniqueness_verification"]["status"] != "success":
            verification = {
                **verification,
                "status": "mismatch",
                "global_uniqueness_verification": writes["influencer"]["global_uniqueness_verification"],
            }
        writes["influencer"]["verification"] = verification
        writes["influencer"]["readback_count"] = verification["readback_count"]
        writes["influencer"]["missing_unique_keys"] = verification["missing_unique_keys"]
        writes["influencer"]["duplicate_unique_keys"] = verification["duplicate_unique_keys"]
        writes["influencer"]["mismatched_rows"] = verification["mismatched_rows"]
        progress("readback_influencer_done", verification)

    summary["field_preflight"] = field_preflight
    summary["writes"] = writes
    verification_failures = []
    for platform, item in writes["orders"].items():
        verification = item.get("verification") or {}
        if verification.get("status") not in (None, "success"):
            verification_failures.append(
                {
                    "section": "orders",
                    "platform": platform,
                    "table_id": field_preflight["orders"].get(platform, {}).get("table_id"),
                    "verification": verification,
                }
            )
    for section in ("ads", "influencer"):
        item = writes.get(section) or {}
        verification = item.get("verification") or {}
        if verification.get("status") not in (None, "success"):
            verification_failures.append(
                {
                    "section": section,
                    "table_id": field_preflight.get(section, {}).get("table_id"),
                    "verification": verification,
                }
            )
    summary["verification_failures"] = verification_failures
    summary["status"] = "source_verified" if not verification_failures else "readback_mismatch"
    progress("source_verification_complete", {"status": summary["status"]})
    return summary


def write_import_evidence(evidence: Path, summary: dict[str, Any]) -> None:
    evidence.parent.mkdir(parents=True, exist_ok=True)
    summary["evidence_path"] = str(evidence.resolve())
    evidence.write_text(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True, default=str), encoding="utf-8")


def write_import_progress(
    evidence: Path,
    summary: dict[str, Any],
    stage: str,
    detail: dict[str, Any] | None = None,
) -> None:
    item = {
        "stage": stage,
        "at": datetime.now().isoformat(timespec="seconds"),
    }
    if detail:
        item["detail"] = detail
    summary.setdefault("progress", []).append(item)
    write_import_evidence(evidence, summary)


def decode_process_output(value: str | bytes | None) -> str:
    if value is None:
        return ""
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return value


def run_formula_dynamic_summary(
    *,
    evidence_dir: Path,
    timeout_seconds: int,
    impact_dates: set[str] | None = None,
) -> dict[str, Any]:
    """Reconcile formula dimensions and verify every date touched by this import."""
    normalized_impact_dates = sorted({normalize_date(value) for value in impact_dates or set()} - {""})
    command = [
        sys.executable,
        str(ROOT / "scripts" / "bootstrap_formula_dynamic_summary.py"),
        "--days-ahead",
        "0",
        "--evidence-dir",
        str(evidence_dir),
    ]
    if normalized_impact_dates:
        for impact_date in normalized_impact_dates:
            command.extend(["--impact-date", impact_date])
    summary_table_id = os.getenv("SHOPOPS_FORMULA_SUMMARY_TABLE_ID", "").strip()
    if summary_table_id:
        command.extend(["--summary-table-id", summary_table_id])
    total_summary_table_id = os.getenv("SHOPOPS_FORMULA_TOTAL_SUMMARY_TABLE_ID", "").strip()
    if total_summary_table_id:
        command.extend(["--total-summary-table-id", total_summary_table_id])

    env = os.environ.copy()
    env.setdefault("PYTHONIOENCODING", "utf-8")
    env.setdefault("PYTHONUTF8", "1")
    started = datetime.now()
    stages: list[dict[str, Any]] = []

    def run_stage(name: str, stage_command: list[str]) -> dict[str, Any]:
        try:
            completed = subprocess.run(
                stage_command,
                cwd=ROOT,
                capture_output=True,
                timeout=timeout_seconds,
                env=env,
            )
        except subprocess.TimeoutExpired as exc:
            return {
                "name": name,
                "status": "timed_out",
                "command": stage_command,
                "returncode": 124,
                "timed_out": True,
                "stdout_tail": decode_process_output(exc.stdout)[-12000:],
                "stderr_tail": decode_process_output(exc.stderr)[-12000:],
            }

        stdout = decode_process_output(completed.stdout)
        stderr = decode_process_output(completed.stderr)
        parsed: dict[str, Any] | None = None
        try:
            parsed = json.loads(stdout)
        except (TypeError, json.JSONDecodeError):
            parsed = None
        return {
            "name": name,
            "status": "success" if completed.returncode == 0 and parsed else "failed",
            "command": stage_command,
            "returncode": completed.returncode,
            "timed_out": False,
            "summary": parsed,
            "stdout_tail": stdout[-12000:],
            "stderr_tail": stderr[-12000:],
        }

    bootstrap_stage = run_stage("bootstrap", command)
    stages.append(bootstrap_stage)
    if bootstrap_stage["status"] != "success":
        return {
            "status": bootstrap_stage["status"],
            "command": command,
            "returncode": bootstrap_stage["returncode"],
            "timed_out": bootstrap_stage["timed_out"],
            "timeout_seconds": timeout_seconds,
            "started_at": started.isoformat(timespec="seconds"),
            "finished_at": datetime.now().isoformat(timespec="seconds"),
            "summary": bootstrap_stage.get("summary"),
            "stages": stages,
            "stdout_tail": bootstrap_stage["stdout_tail"],
            "stderr_tail": bootstrap_stage["stderr_tail"],
        }

    bootstrap_summary = bootstrap_stage["summary"] or {}
    resolved_summary_table_id = summary_table_id or str(
        ((bootstrap_summary.get("summary_table") or {}).get("table_id") or "")
    ).strip()
    reconciliation_commands = [
        (
            "product_detail_formulas",
            [
                sys.executable,
                str(ROOT / "scripts" / "repair_formula_summary_product_detail_formulas.py"),
                "--target-table-id",
                resolved_summary_table_id,
                "--evidence-dir",
                str(evidence_dir),
            ],
        ),
        (
            "product_order_sales",
            [
                sys.executable,
                str(ROOT / "scripts" / "repair_formula_summary_product_order_sales.py"),
                "--target-table-id",
                resolved_summary_table_id,
                "--evidence-dir",
                str(evidence_dir),
                *[
                    value
                    for impact_date in normalized_impact_dates
                    for value in ("--impact-date", impact_date)
                ],
            ],
        ),
    ]
    def run_reconciliation_stages(prefix: str = "") -> dict[str, Any] | None:
        for name, stage_command in reconciliation_commands:
            if not resolved_summary_table_id:
                return {
                    "status": "failed",
                    "command": command,
                    "returncode": 1,
                    "timed_out": False,
                    "timeout_seconds": timeout_seconds,
                    "started_at": started.isoformat(timespec="seconds"),
                    "finished_at": datetime.now().isoformat(timespec="seconds"),
                    "summary": bootstrap_summary,
                    "stages": stages,
                    "stdout_tail": "",
                    "stderr_tail": "Formula summary table id was not returned by bootstrap",
                }
            stage = run_stage(f"{prefix}{name}", stage_command)
            stages.append(stage)
            if stage["status"] != "success":
                return {
                    "status": stage["status"],
                    "command": command,
                    "returncode": stage["returncode"],
                    "timed_out": stage["timed_out"],
                    "timeout_seconds": timeout_seconds,
                    "started_at": started.isoformat(timespec="seconds"),
                    "finished_at": datetime.now().isoformat(timespec="seconds"),
                    "summary": bootstrap_summary,
                    "stages": stages,
                    "stdout_tail": stage["stdout_tail"],
                    "stderr_tail": stage["stderr_tail"],
                }
        return None

    reconciliation_failure = run_reconciliation_stages()
    if reconciliation_failure:
        return reconciliation_failure

    def run_verifiers(attempt: int) -> list[dict[str, Any]]:
        if not normalized_impact_dates:
            return []
        range_start = datetime.strptime(normalized_impact_dates[0], "%Y-%m-%d").date()
        range_end = datetime.strptime(normalized_impact_dates[-1], "%Y-%m-%d").date()
        verifier_stages: list[dict[str, Any]] = []
        chunk_start = range_start
        while chunk_start <= range_end:
            chunk_end = min(range_end, chunk_start + timedelta(days=13))
            start_date = chunk_start.isoformat()
            end_date = chunk_end.isoformat()
            evidence_path = evidence_dir / (
                f"source-summary-verification-{start_date.replace('-', '')}-{end_date.replace('-', '')}-attempt-{attempt}.json"
            )
            # The verifier emits one comparison row per date/platform/product.
            # Fourteen-day windows preserve that per-date gate while avoiding
            # one full source-table read per day and oversized filter formulas.
            stage = run_stage(
                f"verify_{start_date}_{end_date}_attempt_{attempt}",
                [
                    sys.executable,
                    str(ROOT / "scripts" / "verify_formula_dynamic_summary.py"),
                    "--start-date",
                    start_date,
                    "--end-date",
                    end_date,
                    "--evidence",
                    str(evidence_path),
                ],
            )
            stages.append(stage)
            verifier_stages.append(stage)
            chunk_start = chunk_end + timedelta(days=1)
        return verifier_stages

    verifier_stages = run_verifiers(attempt=1) if normalized_impact_dates else []
    if verifier_stages and any(stage["status"] != "success" for stage in verifier_stages):
        # A verifier is read-only.  One idempotent bootstrap retry repairs
        # structural rows and refreshes formula definitions before re-verifying.
        # Formula updates are forced only after the first independent readback
        # fails, avoiding routine full-table recalculation on successful runs.
        retry_command = [*command, "--force-formula-updates", "--force-summary-formula-updates"]
        retry_bootstrap = run_stage("bootstrap_dimension_and_formula_retry", retry_command)
        stages.append(retry_bootstrap)
        if retry_bootstrap["status"] != "success":
            return {
                "status": retry_bootstrap["status"],
                "command": command,
                "returncode": retry_bootstrap["returncode"],
                "timed_out": retry_bootstrap["timed_out"],
                "timeout_seconds": timeout_seconds,
                "started_at": started.isoformat(timespec="seconds"),
                "finished_at": datetime.now().isoformat(timespec="seconds"),
                "summary": retry_bootstrap.get("summary") or bootstrap_summary,
                "stages": stages,
                "stdout_tail": retry_bootstrap["stdout_tail"],
                "stderr_tail": retry_bootstrap["stderr_tail"],
            }
        reconciliation_failure = run_reconciliation_stages(prefix="retry_")
        if reconciliation_failure:
            return reconciliation_failure
        verifier_stages = run_verifiers(attempt=2)
    if verifier_stages and any(stage["status"] != "success" for stage in verifier_stages):
        failed_verifiers = [stage for stage in verifier_stages if stage["status"] != "success"]
        return {
            "status": "verification_failed",
            "command": command,
            "returncode": 4,
            "timed_out": any(stage["timed_out"] for stage in failed_verifiers),
            "timeout_seconds": timeout_seconds,
            "started_at": started.isoformat(timespec="seconds"),
            "finished_at": datetime.now().isoformat(timespec="seconds"),
            "summary": bootstrap_summary,
            "impact_dates": normalized_impact_dates,
            "stages": stages,
            "stdout_tail": failed_verifiers[-1]["stdout_tail"],
            "stderr_tail": failed_verifiers[-1]["stderr_tail"],
        }

    return {
        "status": "success",
        "command": command,
        "returncode": 0,
        "timed_out": False,
        "timeout_seconds": timeout_seconds,
        "started_at": started.isoformat(timespec="seconds"),
        "finished_at": datetime.now().isoformat(timespec="seconds"),
        "summary": bootstrap_summary,
        "impact_dates": normalized_impact_dates,
        "stages": stages,
        "stdout_tail": stages[-1]["stdout_tail"],
        "stderr_tail": stages[-1]["stderr_tail"],
    }

def normalize_platform(value: Any) -> str:
    text = clean_text(value).lower()
    if "抖音" in text or "douyin" in text or "doudian" in text:
        return "抖音"
    if "拼多多" in text or "pdd" in text or "pinduoduo" in text:
        return "拼多多"
    if "视频号" in text or "微信" in text or "wechat" in text:
        return "视频号"
    if "天猫" in text or "tmall" in text:
        return "天猫"
    return clean_text(value)


ORDER_CREATED_AT_ALIASES: dict[str, tuple[str, ...]] = {
    "天猫": ("订单创建时间", "创建时间", "下单时间", "订单下单时间", "订单付款时间", "支付时间"),
    "抖音": ("订单提交时间", "下单时间", "订单下单时间", "创建时间", "支付完成时间", "付款时间", "支付时间"),
    "拼多多": ("订单成交时间", "下单时间", "订单创建时间", "创建时间", "支付时间"),
    "视频号": ("订单下单时间", "下单时间", "订单创建时间", "创建时间", "支付时间"),
}


def order_created_at(platform: str, row: dict[str, Any], order_no: str) -> str:
    value = normalize_datetime(first_present(row, *ORDER_CREATED_AT_ALIASES.get(platform, ("创建时间", "下单时间"))))
    if value:
        return value
    if platform == "拼多多":
        return pdd_date_from_order_no(order_no)
    return ""


NON_SOLD_STATUS_KEYWORDS = (
    "退款",
    "交易关闭",
    "已关闭",
    "已取消",
    "订单关闭",
    "待付款",
    "等待买家付款",
    "未付款",
    "Cancelled",
    "cancelled",
    "CANCELLED",
)
NON_SOLD_PRODUCT_KEYWORDS = ("补收差价", "差价专用", "购买前须联系客服", "联系客服确认")
ACCESSORY_PRODUCT_KEYWORDS = ("配件",)


def actual_sold_quantity(
    *,
    quantity: float | None,
    product: str,
    unit_price: float | None,
    refund_amount: float | None,
    trade_status: str,
    fulfill_status: str,
) -> float | None:
    if refund_amount and refund_amount > 0:
        return 0
    text = f"{trade_status}/{fulfill_status}".replace("无售后或售后取消", "")
    if any(keyword in text for keyword in NON_SOLD_STATUS_KEYWORDS):
        return 0
    if unit_price == 0 or any(keyword in product for keyword in NON_SOLD_PRODUCT_KEYWORDS) or is_accessory_product(product):
        return 0
    return quantity


def is_accessory_product(product: str) -> bool:
    return any(keyword in product for keyword in ACCESSORY_PRODUCT_KEYWORDS)


def pdd_date_from_order_no(order_no: str) -> str:
    prefix = clean_text(order_no).split("-", 1)[0]
    if len(prefix) != 6 or not prefix.isdigit():
        return ""
    month = int(prefix[2:4])
    day = int(prefix[4:6])
    if not (1 <= month <= 12 and 1 <= day <= 31):
        return ""
    return f"20{prefix[:2]}-{month:02d}-{day:02d}"


def order_unique_key(platform: str, order_no: str) -> str:
    return f"{PLATFORM_CODES[platform]}_{clean_text(order_no)}"


def ad_unique_key(platform: str, date_text: str) -> str:
    return f"ads_{PLATFORM_CODES[platform]}_{date_text}"


def clean_header(value: Any) -> str:
    text = "" if value is None or is_nan(value) else str(value).strip()
    return re.sub(r"[\ue000-\uf8ff].*$", "", text).strip()


def clean_text(value: Any) -> str:
    if value in (None, "-", "--") or is_nan(value):
        return ""
    return str(value).strip().strip("\t").strip("'")


def scalar_text(value: Any) -> str:
    if isinstance(value, list):
        return "".join(str(item.get("text") if isinstance(item, dict) else item) for item in value).strip()
    return clean_text(value)


def field_value_equal(left: Any, right: Any) -> bool:
    if left in (None, "") and right in (None, ""):
        return True
    left_number = number_value(left)
    right_number = number_value(right)
    if left_number is not None and right_number is not None:
        return abs(left_number - right_number) < 0.000001
    return scalar_text(left) == scalar_text(right)


def missing_row_fields(existing_fields: set[str], rows: list[dict[str, Any]], required_fields: list[str]) -> list[str]:
    used_fields = {
        key
        for row in rows
        for key, value in row.items()
        if value not in (None, "")
    }
    return sorted(({*required_fields, *used_fields} - existing_fields))


def first_present(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in row and row[key] not in (None, "") and not is_nan(row[key]):
            return row[key]
    return None


def number_value(value: Any) -> float | None:
    text = clean_text(value).replace(",", "").replace("元", "")
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1]
    try:
        return round(float(text), 6)
    except ValueError:
        return None


def sum_numbers(rows: list[dict[str, Any]], *keys: str) -> float:
    total = 0.0
    for row in rows:
        for key in keys:
            value = number_value(row.get(key))
            if value is not None:
                total += value
                break
    return round(total, 6)


def first_number(rows: list[dict[str, Any]], *keys: str) -> float | None:
    for row in rows:
        for key in keys:
            value = number_value(row.get(key))
            if value is not None:
                return value
    return None


def ratio(numerator: float | int | None, denominator: float | int | None) -> float | None:
    if numerator is None or denominator in (None, 0):
        return None
    return round(float(numerator) / float(denominator), 6)


def refund_from_status(explicit_refund: float | None, paid_amount: float | None, status: str) -> float:
    if explicit_refund is not None:
        return explicit_refund
    return paid_amount or 0 if "退款成功" in status else 0


def normalize_datetime(value: Any) -> str:
    text = clean_text(value).replace("/", "-")
    if not text:
        return ""
    for candidate in (text, text[:19], text[:16], text[:10]):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(candidate, fmt)
                return parsed.strftime("%Y-%m-%d" if fmt == "%Y-%m-%d" else "%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
    return text


def normalize_date(value: Any) -> str:
    text = normalize_datetime(value)
    return text[:10] if text else ""


def rolling_date_window(selected_dates: set[str], lookback_days: int) -> dict[str, Any] | None:
    parsed_dates = []
    for value in selected_dates:
        try:
            parsed_dates.append(datetime.strptime(value, "%Y-%m-%d").date())
        except ValueError:
            continue
    if not parsed_dates:
        return None
    end_date = max(parsed_dates)
    start_date = end_date - timedelta(days=lookback_days)
    return {
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "lookback_days": lookback_days,
    }


def date_in_window(value: Any, window: dict[str, Any] | None) -> bool:
    if not window:
        return True
    date_text = normalize_date(value)
    return bool(date_text) and str(window["start_date"]) <= date_text <= str(window["end_date"])


def redact_row(row: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in row.items():
        if any(part in str(key) for part in SENSITIVE_KEY_PARTS):
            result[key] = "[REDACTED]"
        else:
            result[key] = value
    return result


def join_unique(left: Any, right: Any, separator: str = "; ") -> str:
    values: list[str] = []
    for value in (clean_text(left), clean_text(right)):
        if not value:
            continue
        for part in value.split(separator):
            item = part.strip()
            if item and item not in values:
                values.append(item)
    return separator.join(values)


def is_nan(value: Any) -> bool:
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def chunks(values: list[dict[str, Any]], size: int) -> list[list[dict[str, Any]]]:
    return [values[index : index + size] for index in range(0, len(values), size)]


def main() -> int:
    parser = argparse.ArgumentParser(description="Import one ShopOps daily folder into existing Feishu Bitable tables.")
    parser.add_argument("--batch-dir", default=r"D:\lyh\ShopOps\0608", help="Daily folder such as D:\\lyh\\ShopOps\\0608.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and validate locally without writing Feishu.")
    parser.add_argument("--evidence", default="", help="Evidence JSON path.")
    parser.add_argument("--platform", action="append", choices=PLATFORMS, help="Only import one platform; repeat for multiple platforms.")
    parser.add_argument("--kind", action="append", choices=("orders", "ads", "influencer"), help="Only import one data kind; repeat for multiple kinds.")
    parser.add_argument("--date", action="append", help="Only import one normalized date (YYYY-MM-DD); repeat for multiple dates.")
    parser.add_argument("--filter-ad-dates", action="store_true", help="Also apply --date filters to ad files. By default ad imports use every date present in the source files.")
    parser.add_argument("--ensure-missing-ad-fields", action="store_true", help="Create missing Feishu ad table fields that are present in imported rows.")
    parser.add_argument(
        "--order-lookback-days",
        type=int,
        default=ORDER_ROLLING_LOOKBACK_DAYS,
        help="Order date lookback around --date. Default keeps the daily historical 90-day update window; hourly jobs pass 0 for date-only incremental imports.",
    )
    parser.add_argument(
        "--formula-summary-timeout-seconds",
        type=int,
        default=1800,
        help="Timeout for refreshing the formula dynamic summary after a successful source-table import.",
    )
    args = parser.parse_args()

    batch_dir = Path(args.batch_dir)
    date_dir = batch_dir.name
    evidence = Path(args.evidence) if args.evidence else Path("docs/live-evidence") / f"daily-import-{date_dir}.json"
    try:
        summary = run_import(
            batch_dir=batch_dir,
            dry_run=args.dry_run,
            evidence=evidence,
            platforms=set(args.platform or []),
            kinds=set(args.kind or []),
            dates=set(args.date or []),
            filter_ad_dates=args.filter_ad_dates,
            ensure_missing_ad_fields=args.ensure_missing_ad_fields,
            order_lookback_days=args.order_lookback_days,
        )
        if args.dry_run:
            summary["formula_summary"] = {
                "status": "skipped",
                "reason": "dry_run_does_not_write_feishu_source_or_summary_tables",
            }
        elif summary.get("status") == "source_verified":
            formula_summary = run_formula_dynamic_summary(
                evidence_dir=Path("docs/live-evidence/formula-dynamic-summary"),
                timeout_seconds=args.formula_summary_timeout_seconds,
                impact_dates=set(summary.get("impact_dates") or summary.get("date_filter") or []),
            )
            summary["formula_summary"] = formula_summary
            summary["status"] = "success" if formula_summary.get("status") == "success" else "formula_summary_failed"
        write_import_evidence(evidence, summary)
    except Exception as exc:
        summary = {
            "status": "failed",
            "batch_dir": str(batch_dir),
            "platform_filter": sorted(set(args.platform or [])),
            "kind_filter": sorted(set(args.kind or [])),
            "date_filter": sorted(set(args.date or [])),
            "error": {
                "type": type(exc).__name__,
                "message": str(exc),
                "traceback_tail": traceback.format_exc()[-6000:],
            },
        }
        write_import_evidence(evidence, summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True, default=str))
    return 0 if summary["status"] in {"success", "dry_run"} else 4


if __name__ == "__main__":
    raise SystemExit(main())
