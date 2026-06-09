from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import time
from collections import Counter, defaultdict
from dataclasses import replace
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from shopops.collectors.jushuitan_order_api import JushuitanOrderApiCollector, first_value
from shopops.config import Settings, _load_dotenv, load_settings
from shopops.services.product_breakdown import (
    DEFAULT_PRODUCT_CATALOG_TABLE_ID,
    ProductRule,
    effective_sales_amount,
    product_breakdown_values,
    product_field_names,
    product_rules_from_records,
    summary_product_formula_fields,
)
from shopops.storage.feishu_bootstrap import FEISHU_BASE_URL, FeishuOpenApiClient, PlatformTableSpec, merge_env_file


TEXT_FIELD = 1
NUMBER_FIELD = 2
FORMULA_FIELD = 20

TOTAL_PLATFORM = "全平台总计"
PLATFORMS = ("天猫", "抖音", "拼多多", "视频号")
API_ORDER_PLATFORMS = {"抖音", "视频号"}
LOCAL_ORDER_PLATFORMS = set(PLATFORMS)
PLATFORM_CODES = {
    "天猫": "tmall",
    "抖音": "douyin",
    "拼多多": "pdd",
    "视频号": "shipinhao",
}
SOURCE_ORDER_TABLE_ID = "tblbPjczqMElVU9a"
SUMMARY_TABLE_ID = "tblepMIg19Ov1kSw"

TARGET_TABLES = {
    "天猫": {"name": "订单明细-天猫", "env": "SHOPOPS_ORDER_TABLE_TMALL_ID"},
    "抖音": {"name": "订单明细-抖音", "env": "SHOPOPS_ORDER_TABLE_DOUYIN_ID"},
    "拼多多": {"name": "订单明细-拼多多", "env": "SHOPOPS_ORDER_TABLE_PINDUODUO_ID"},
    "视频号": {"name": "订单明细-视频号", "env": "SHOPOPS_ORDER_TABLE_WECHAT_CHANNELS_ID"},
}

STALE_IMPORT_KEY_PREFIXES = {
    "天猫": ("tmall_export_", "taobao_", "jushuitan_"),
    "抖音": ("douyin_export_", "jushuitan_"),
    "拼多多": ("pdd_export_", "jushuitan_"),
    "视频号": ("wechat_channels_export_", "wechat_channels_", "jushuitan_"),
}


def normalize_platform_formula(value_expr: str) -> str:
    return (
        f'IF({value_expr}.CONTAIN("抖音"),"抖音",'
        f'IF({value_expr}.CONTAIN("拼多多")||{value_expr}.CONTAIN("PDD"),"拼多多",'
        f'IF({value_expr}.CONTAIN("视频号")||{value_expr}.CONTAIN("微信"),"视频号",'
        f'IF({value_expr}.CONTAIN("天猫")||{value_expr}.CONTAIN("淘宝")||{value_expr}.CONTAIN("千牛"),"天猫",{value_expr}))))'
    )

F_UNIQUE_KEY = "unique_key"
F_PLATFORM = "平台"
F_DATA_SOURCE = "数据来源"
F_SHOP_ID = "店铺ID"
F_SHOP_NAME = "店铺名称"
F_FETCHED_AT = "采集时间"
F_ORDER_NO = "订单号"
F_CREATED_AT = "创建时间"
F_BUYER_NICK = "买家昵称"
F_PRODUCT_NAME = "商品名称"
F_UNIT_PRICE = "单价"
F_QUANTITY = "数量"
F_FULFILL_STATUS = "履约/售后状态"
F_TRADE_STATUS = "交易状态"
F_PAID_AMOUNT = "实收款"
F_REFUND_AMOUNT = "退款金额"
F_PRODUCT_COST = "商品成本"
F_FREIGHT_COST = "运费成本"
F_PLATFORM_FEE = "平台扣点"
F_OTHER_FEE = "其他费用"
F_OPERATION = "操作信息"
F_PAGE_URL = "页面URL"
F_SCREENSHOT = "页面截图"
F_RAW = "原始数据"

UNIFIED_FIELDS = [
    (F_UNIQUE_KEY, TEXT_FIELD),
    (F_PLATFORM, TEXT_FIELD),
    (F_DATA_SOURCE, TEXT_FIELD),
    (F_SHOP_ID, TEXT_FIELD),
    (F_SHOP_NAME, TEXT_FIELD),
    (F_FETCHED_AT, TEXT_FIELD),
    (F_ORDER_NO, TEXT_FIELD),
    (F_CREATED_AT, TEXT_FIELD),
    (F_BUYER_NICK, TEXT_FIELD),
    (F_PRODUCT_NAME, TEXT_FIELD),
    (F_UNIT_PRICE, NUMBER_FIELD),
    (F_QUANTITY, NUMBER_FIELD),
    (F_FULFILL_STATUS, TEXT_FIELD),
    (F_TRADE_STATUS, TEXT_FIELD),
    (F_PAID_AMOUNT, NUMBER_FIELD),
    (F_REFUND_AMOUNT, NUMBER_FIELD),
    (F_PRODUCT_COST, NUMBER_FIELD),
    (F_FREIGHT_COST, NUMBER_FIELD),
    (F_PLATFORM_FEE, NUMBER_FIELD),
    (F_OTHER_FEE, NUMBER_FIELD),
    (F_OPERATION, TEXT_FIELD),
    (F_PAGE_URL, TEXT_FIELD),
    (F_SCREENSHOT, TEXT_FIELD),
    (F_RAW, TEXT_FIELD),
]

FORMULA_FIELDS = {
    "公式_统计日期": {"formatter": "", "expression": 'LEFT([创建时间],10)'},
    "公式_汇总平台": {"formatter": "", "expression": normalize_platform_formula("[平台]")},
    "公式_销售额": {"formatter": "0.00", "expression": "IFBLANK([实收款],0)"},
    "公式_退款金额": {"formatter": "0.00", "expression": "IFBLANK([退款金额],0)"},
    "公式_有效销售额": {
        "formatter": "0.00",
        "expression": "IF(IFBLANK([实收款],0)-IFBLANK([退款金额],0)<0,0,IFBLANK([实收款],0)-IFBLANK([退款金额],0))",
    },
    "公式_商品成本": {"formatter": "0.00", "expression": "IFBLANK([商品成本],0)"},
    "公式_运费成本": {"formatter": "0.00", "expression": "IFBLANK([运费成本],0)"},
    "公式_平台扣点": {"formatter": "0.00", "expression": "IFBLANK([平台扣点],0)"},
    "公式_其他费用": {"formatter": "0.00", "expression": "IFBLANK([其他费用],0)"},
}

ORDER_SUMMARY_FIELDS = {
    "订单数": {"formatter": "0", "source_field": "unique_key", "op": "COUNTA"},
    "实际卖出数量": {"formatter": "0", "source_field": "数量", "op": "SUM"},
    "销售额": {"formatter": "0.00", "source_field": "公式_销售额", "op": "SUM"},
    "退款金额": {"formatter": "0.00", "source_field": "公式_退款金额", "op": "SUM"},
    "有效销售额": {"formatter": "0.00", "source_field": "公式_有效销售额", "op": "SUM"},
    "商品成本": {"formatter": "0.00", "source_field": "公式_商品成本", "op": "SUM"},
    "运费成本": {"formatter": "0.00", "source_field": "公式_运费成本", "op": "SUM"},
    "平台扣点": {"formatter": "0.00", "source_field": "公式_平台扣点", "op": "SUM"},
    "其他费用": {"formatter": "0.00", "source_field": "公式_其他费用", "op": "SUM"},
}

SENSITIVE_KEY_PARTS = (
    "phone",
    "mobile",
    "tel",
    "address",
    "addr",
    "receiver",
    "consignee",
    "收件",
    "收货",
    "手机",
    "电话",
    "地址",
    "消费者资料",
    "详细地址",
)

class FeishuClient:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings
        self.app_token = settings.shopops_data_center_app_token or settings.feishu_app_token
        if not self.app_token:
            raise RuntimeError("Missing FEISHU_APP_TOKEN or SHOPOPS_DATA_CENTER_APP_TOKEN")
        self.auth = FeishuOpenApiClient(settings.feishu_app_id, settings.feishu_app_secret)

    def request(
        self,
        method: str,
        path: str,
        payload: dict[str, Any] | None = None,
        params: dict[str, Any] | None = None,
        allow_duplicate_field: bool = False,
    ) -> dict[str, Any]:
        response = requests.request(
            method,
            f"{FEISHU_BASE_URL}{path}",
            headers=self.auth.headers(),
            json=payload,
            params=params,
            timeout=30,
        )
        try:
            body = response.json()
        except ValueError as exc:
            raise RuntimeError(f"Feishu API returned non-JSON HTTP {response.status_code}: {response.text[:500]}") from exc
        if allow_duplicate_field and body.get("code") == 1254014:
            return {}
        if response.status_code >= 400 or body.get("code") != 0:
            raise RuntimeError(f"Feishu API {method} {path} failed HTTP {response.status_code}: {body}")
        return body.get("data") or {}

    def list_tables(self) -> list[dict[str, Any]]:
        data = self.request("GET", f"/bitable/v1/apps/{self.app_token}/tables", params={"page_size": 100})
        return list(data.get("items") or [])

    def ensure_order_tables(self, env_path: Path) -> dict[str, str]:
        existing_by_name = {str(item.get("name")): item for item in self.list_tables() if item.get("name")}
        result: dict[str, str] = {}
        env_updates: dict[str, str] = {}
        for platform, config in TARGET_TABLES.items():
            table_name = str(config["name"])
            existing = existing_by_name.get(table_name)
            if existing:
                table_id = str(existing.get("table_id"))
                reused = True
            else:
                table_id = self.create_order_table(table_name)
                reused = False
            self.ensure_unified_order_fields(table_id)
            result[platform] = table_id
            env_updates[str(config["env"])] = table_id
            env_updates[f'{config["env"]}_NAME'] = table_name
            log(f"table platform={platform} table_id={table_id} reused={reused}")
        merge_env_file(env_path, env_updates)
        return result

    def create_order_table(self, table_name: str) -> str:
        spec = PlatformTableSpec(
            "SHOPOPS_PLATFORM_ORDER_TABLE",
            table_name,
            table_name,
            [{"field_name": name, "type": field_type} for name, field_type in UNIFIED_FIELDS],
        )
        data = self.auth.create_table(self.app_token, spec)
        table_id = str(data.get("table_id") or "")
        if not table_id:
            raise RuntimeError(f"Create table {table_name} did not return table_id")
        return table_id

    def field_index(self, table_id: str) -> dict[str, dict[str, Any]]:
        fields: dict[str, dict[str, Any]] = {}
        page_token = None
        while True:
            params: dict[str, Any] = {"page_size": 100}
            if page_token:
                params["page_token"] = page_token
            data = self.request("GET", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/fields", params=params)
            for item in data.get("items", []) or []:
                if item.get("field_name"):
                    fields[str(item["field_name"])] = item
            if not data.get("has_more"):
                return fields
            page_token = data.get("page_token")

    def ensure_unified_order_fields(self, table_id: str) -> None:
        existing = self.field_index(table_id)
        for name, field_type in UNIFIED_FIELDS:
            if name in existing:
                continue
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/fields",
                {"field_name": name, "type": field_type},
                allow_duplicate_field=True,
            )
            existing = self.field_index(table_id)
        for name, config in FORMULA_FIELDS.items():
            self.ensure_formula_field(table_id, name, config["expression"], config["formatter"])

    def product_rules(self, product_table_id: str) -> list[ProductRule]:
        return product_rules_from_records(list(self.list_records(product_table_id)))

    def ensure_product_breakdown_fields(self, table_id: str, rules: list[ProductRule]) -> list[str]:
        existing = self.field_index(table_id)
        result: list[str] = []
        for name in product_field_names(rules):
            current = existing.get(name)
            if not current:
                self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/fields", {"field_name": name, "type": NUMBER_FIELD})
                existing[name] = {"field_name": name, "type": NUMBER_FIELD}
            elif int(current.get("type") or 0) != NUMBER_FIELD:
                self.request("DELETE", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/fields/{current['field_id']}")
                self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/fields", {"field_name": name, "type": NUMBER_FIELD})
                existing[name] = {"field_name": name, "type": NUMBER_FIELD}
            result.append(name)
        return result

    def ensure_formula_field(self, table_id: str, name: str, expression: str, formatter: str) -> None:
        existing = self.field_index(table_id)
        payload = {
            "field_name": name,
            "type": FORMULA_FIELD,
            "property": {"formatter": formatter, "formula_expression": expression},
        }
        current = existing.get(name)
        if not current:
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/fields", payload)
            return
        if int(current.get("type") or 0) != FORMULA_FIELD:
            raise RuntimeError(f"Field {name} exists in table {table_id}, but it is not a formula field")
        self.request("PUT", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/fields/{current['field_id']}", payload)

    def list_records(self, table_id: str, page_size: int = 500) -> Iterable[dict[str, Any]]:
        page_token = None
        page = 0
        while True:
            page += 1
            params: dict[str, Any] = {"page_size": page_size}
            if page_token:
                params["page_token"] = page_token
            data = self.request("GET", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records", params=params)
            items = data.get("items") or []
            log(f"list_records table={table_id} page={page} count={len(items)}")
            yield from items
            if not data.get("has_more"):
                return
            page_token = data.get("page_token")

    def record_index(self, table_id: str) -> dict[str, str]:
        index: dict[str, str] = {}
        for item in self.list_records(table_id):
            fields = item.get("fields") or {}
            unique_key = fields.get(F_UNIQUE_KEY)
            if unique_key:
                index[str(unique_key)] = str(item.get("record_id"))
        return index

    def upsert_records(self, table_id: str, rows: list[dict[str, Any]]) -> dict[str, int]:
        if not rows:
            return {"created": 0, "updated": 0, "saved": 0}
        index = self.record_index(table_id)
        to_create: list[dict[str, Any]] = []
        to_update: list[dict[str, Any]] = []
        for row in rows:
            unique_key = str(row.get(F_UNIQUE_KEY) or "")
            if not unique_key:
                continue
            payload = {"fields": clean_feishu_fields(row)}
            record_id = index.get(unique_key)
            if record_id:
                payload["record_id"] = record_id
                to_update.append(payload)
            else:
                to_create.append(payload)
        for chunk in chunks(to_create, 500):
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_create", {"records": chunk})
        for chunk in chunks(to_update, 500):
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_update", {"records": chunk})
        return {"created": len(to_create), "updated": len(to_update), "saved": len(to_create) + len(to_update)}

    def delete_records(self, table_id: str, record_ids: list[str]) -> int:
        for chunk in chunks(record_ids, 500):
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_delete",
                {"records": chunk},
            )
        return len(record_ids)

    def delete_records_by_unique_key_prefixes(self, table_id: str, prefixes: tuple[str, ...]) -> int:
        if not prefixes:
            return 0
        record_ids: list[str] = []
        for record in self.list_records(table_id):
            fields = record.get("fields") or {}
            unique_key = scalar_text(fields.get(F_UNIQUE_KEY))
            if any(unique_key.startswith(prefix) for prefix in prefixes):
                record_id = record.get("record_id")
                if record_id:
                    record_ids.append(str(record_id))
        return self.delete_records(table_id, record_ids)

    def update_formula_summary(
        self,
        table_ids: dict[str, str],
        summary_table_id: str,
        product_rules: list[ProductRule] | None = None,
    ) -> None:
        order_table_names = [TARGET_TABLES[platform]["name"] for platform in PLATFORMS]
        for field_name, config in ORDER_SUMMARY_FIELDS.items():
            expression = build_order_summary_expression(order_table_names, config["source_field"], config["op"])
            self.ensure_formula_field(summary_table_id, field_name, expression, config["formatter"])
        for field_name, config in summary_product_formula_fields(order_table_names, product_rules or []).items():
            self.ensure_formula_field(summary_table_id, field_name, config["expression"], config["formatter"])

    def update_ad_platform_formula(self, ad_table_id: str | None) -> None:
        if not ad_table_id:
            return
        try:
            self.ensure_formula_field(ad_table_id, "公式_汇总平台", normalize_platform_formula("[平台]"), "")
        except Exception as exc:
            log(f"warning: failed to update ad platform formula table={ad_table_id}: {exc}")

    def normalize_existing_summary_platforms(self, summary_table_id: str) -> int:
        updates: list[dict[str, Any]] = []
        for record in self.list_records(summary_table_id):
            fields = record.get("fields") or {}
            platform = scalar_text(fields.get("平台"))
            if platform not in {"淘宝", "千牛淘宝"}:
                continue
            stat_date = scalar_text(fields.get("统计日期"))
            if not stat_date:
                continue
            updates.append(
                {
                    "record_id": str(record.get("record_id")),
                    "fields": {
                        "unique_key": f"{stat_date}-天猫",
                        "统计日期": stat_date,
                        "平台": "天猫",
                    },
                }
            )
        for chunk in chunks(updates, 500):
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{summary_table_id}/records/batch_update", {"records": chunk})
        return len(updates)

    def upsert_summary_dimension_rows(self, summary_table_id: str, dates: set[date]) -> dict[str, int]:
        rows: list[dict[str, Any]] = []
        for day in sorted(dates):
            day_text = day.isoformat()
            for platform in (*PLATFORMS, TOTAL_PLATFORM):
                rows.append({"unique_key": f"{day_text}-{platform}", "统计日期": day_text, "平台": platform})
        return self.upsert_records(summary_table_id, rows)

    def table_count(self, table_id: str) -> int:
        return sum(1 for _ in self.list_records(table_id))


def build_order_summary_expression(order_table_names: list[str], source_field: str, op: str) -> str:
    parts = []
    for table_name in order_table_names:
        filtered = (
            f"[{table_name}].FILTER("
            "CurrentValue.[公式_统计日期]=[统计日期]&&"
            f'([平台]="{TOTAL_PLATFORM}"||CurrentValue.[平台]=[平台])'
            ")"
        )
        parts.append(f"{filtered}.[{source_field}].{op}()")
    return "+".join(parts)


def split_source_table(client: FeishuClient, source_table_id: str, target_ids: dict[str, str]) -> tuple[dict[str, Any], set[date]]:
    rows_by_platform: dict[str, list[dict[str, Any]]] = defaultdict(list)
    skipped: Counter[str] = Counter()
    dates: set[date] = set()
    for record in client.list_records(source_table_id):
        fields = record.get("fields") or {}
        platform = normalize_platform(fields.get(F_PLATFORM))
        if platform not in target_ids:
            skipped[platform] += 1
            continue
        row = unified_row_from_source(fields, platform)
        rows_by_platform[platform].append(row)
        add_row_date(dates, row)

    upsert: dict[str, Any] = {}
    for platform in PLATFORMS:
        upsert[platform] = client.upsert_records(target_ids[platform], rows_by_platform.get(platform, []))
    return {
        "source_table_id": source_table_id,
        "split_counts": {platform: len(rows_by_platform.get(platform, [])) for platform in PLATFORMS},
        "skipped_platforms": dict(skipped),
        "upsert": upsert,
    }, dates


def unified_row_from_source(fields: dict[str, Any], platform: str) -> dict[str, Any]:
    order_no = scalar_text(fields.get(F_ORDER_NO))
    paid_amount = number(fields.get(F_PAID_AMOUNT)) or number(fields.get("公式_销售额"))
    fulfill_status = scalar_text(fields.get(F_FULFILL_STATUS))
    refund_amount = refund_amount_from_fulfill_status(
        explicit_refund=number(fields.get(F_REFUND_AMOUNT)) or number(fields.get("公式_退款金额")),
        paid_amount=paid_amount,
        fulfill_status=fulfill_status,
    )
    row = {
        F_UNIQUE_KEY: canonical_unique_key(platform, order_no) or scalar_text(fields.get(F_UNIQUE_KEY)),
        F_PLATFORM: platform,
        F_DATA_SOURCE: scalar_text(fields.get(F_DATA_SOURCE)),
        F_SHOP_ID: scalar_text(fields.get(F_SHOP_ID)),
        F_SHOP_NAME: scalar_text(fields.get(F_SHOP_NAME)),
        F_FETCHED_AT: scalar_text(fields.get(F_FETCHED_AT)),
        F_ORDER_NO: order_no,
        F_CREATED_AT: scalar_text(fields.get(F_CREATED_AT)),
        F_BUYER_NICK: scalar_text(fields.get(F_BUYER_NICK)),
        F_PRODUCT_NAME: scalar_text(fields.get(F_PRODUCT_NAME)),
        F_UNIT_PRICE: number(fields.get(F_UNIT_PRICE)),
        F_QUANTITY: number(fields.get(F_QUANTITY)),
        F_FULFILL_STATUS: fulfill_status,
        F_TRADE_STATUS: scalar_text(fields.get(F_TRADE_STATUS)),
        F_PAID_AMOUNT: paid_amount,
        F_REFUND_AMOUNT: refund_amount,
        F_PRODUCT_COST: number(fields.get(F_PRODUCT_COST)) or number(fields.get("公式_商品成本")) or 0,
        F_FREIGHT_COST: number(fields.get(F_FREIGHT_COST)) or number(fields.get("公式_运费成本")) or 0,
        F_PLATFORM_FEE: number(fields.get(F_PLATFORM_FEE)) or number(fields.get("公式_平台扣点")) or 0,
        F_OTHER_FEE: number(fields.get(F_OTHER_FEE)) or number(fields.get("公式_其他费用")) or 0,
        F_OPERATION: scalar_text(fields.get(F_OPERATION)),
        F_PAGE_URL: scalar_text(fields.get(F_PAGE_URL)),
        F_SCREENSHOT: scalar_text(fields.get(F_SCREENSHOT)),
    }
    row[F_RAW] = merge_raw_payload(fields.get(F_RAW), source_extra_fields(fields))
    return row


def source_extra_fields(fields: dict[str, Any]) -> dict[str, Any]:
    unified = {name for name, _ in UNIFIED_FIELDS}
    formulas = set(FORMULA_FIELDS)
    return {key: value for key, value in fields.items() if key not in unified and key not in formulas}


def merge_raw_payload(raw_value: Any, extras: dict[str, Any]) -> str:
    raw_text = scalar_text(raw_value)
    payload: dict[str, Any] = {}
    if raw_text:
        try:
            parsed = json.loads(raw_text)
            if isinstance(parsed, dict):
                payload.update(parsed)
            else:
                payload["raw"] = parsed
        except json.JSONDecodeError:
            payload["raw"] = raw_text
    if extras:
        payload["__split_source_extra_fields__"] = sanitize_raw(extras)
    return json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)


def import_latest_local_files(
    data_root: Path,
    allowed_platforms: set[str] | None = None,
) -> tuple[dict[str, Any], dict[str, list[dict[str, Any]]], set[date]]:
    discovered = discover_latest_order_files(data_root)
    rows_by_platform: dict[str, list[dict[str, Any]]] = {platform: [] for platform in PLATFORMS}
    dates: set[date] = set()
    imports: dict[str, Any] = {}
    for platform, path in discovered.items():
        if allowed_platforms is not None and platform not in allowed_platforms:
            imports[platform] = {
                "source_file": str(path),
                "row_count": 0,
                "source_row_count": 0,
                "merged_duplicate_rows": 0,
                "status": "skipped_fixed_source_policy",
            }
            continue
        source_rows = load_local_order_rows(platform, path)
        rows, merge_info = merge_rows_by_unique_key(source_rows)
        rows_by_platform[platform].extend(rows)
        for row in rows:
            add_row_date(dates, row)
        imports[platform] = {
            "source_file": str(path),
            "row_count": len(rows),
            "source_row_count": len(source_rows),
            "merged_duplicate_rows": merge_info["merged_duplicate_rows"],
            "duplicate_order_keys": merge_info["duplicate_order_keys"],
            "status": "imported",
        }
    for platform in PLATFORMS:
        imports.setdefault(
            platform,
            {
                "source_file": None,
                "row_count": 0,
                "source_row_count": 0,
                "merged_duplicate_rows": 0,
                "duplicate_order_keys": [],
                "status": "no_file",
            },
        )
    return imports, rows_by_platform, dates


def discover_latest_order_files(data_root: Path) -> dict[str, Path]:
    result: dict[str, Path] = {}
    if not data_root.exists():
        return result
    for platform_dir in data_root.iterdir():
        if not platform_dir.is_dir():
            continue
        platform = normalize_platform(platform_dir.name)
        if platform not in PLATFORMS:
            continue
        candidates = []
        for path in platform_dir.rglob("*"):
            if not path.is_file() or path.suffix.lower() not in {".csv", ".xlsx", ".xls"}:
                continue
            name = path.name
            if "商品推广" in name or "推广" in name or "账户_分天数据" in name:
                continue
            candidates.append(path)
        if candidates:
            result[platform] = max(candidates, key=lambda item: item.stat().st_mtime)
    return result


def load_local_order_rows(platform: str, path: Path) -> list[dict[str, Any]]:
    source_rows = read_tabular_file(path)
    if platform == "抖音":
        return [row for row in (douyin_export_row(item, path) for item in source_rows) if row]
    if platform == "拼多多":
        return [row for row in (pdd_export_row(item, path) for item in source_rows) if row]
    if platform == "视频号":
        return [row for row in (wechat_channels_export_row(item, path) for item in source_rows) if row]
    if platform == "天猫":
        return [row for row in (tmall_export_row(item, path) for item in source_rows) if row]
    return []


def merge_rows_by_unique_key(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    duplicate_keys: list[str] = []
    for row in rows:
        unique_key = scalar_text(row.get(F_UNIQUE_KEY))
        if not unique_key:
            continue
        current = merged.get(unique_key)
        if current is None:
            merged[unique_key] = dict(row)
            continue
        duplicate_keys.append(unique_key)
        merge_duplicate_order_row(current, row)
    return list(merged.values()), {
        "merged_duplicate_rows": len(rows) - len(merged),
        "duplicate_order_keys": sorted(set(duplicate_keys))[:50],
    }


def merge_duplicate_order_row(target: dict[str, Any], source: dict[str, Any]) -> None:
    target[F_PRODUCT_NAME] = join_unique_text(target.get(F_PRODUCT_NAME), source.get(F_PRODUCT_NAME))
    for field in (F_QUANTITY, F_PAID_AMOUNT, F_REFUND_AMOUNT, F_PRODUCT_COST, F_FREIGHT_COST, F_PLATFORM_FEE, F_OTHER_FEE):
        target[field] = round((number(target.get(field)) or 0) + (number(source.get(field)) or 0), 2)
    target[F_FULFILL_STATUS] = join_unique_text(target.get(F_FULFILL_STATUS), source.get(F_FULFILL_STATUS), separator="/")
    target[F_TRADE_STATUS] = join_unique_text(target.get(F_TRADE_STATUS), source.get(F_TRADE_STATUS), separator="/")
    target[F_OPERATION] = join_unique_text(target.get(F_OPERATION), source.get(F_OPERATION), separator="/")
    target[F_RAW] = json.dumps(
        {
            "merged": True,
            "rows": [raw_payload(target.get(F_RAW)), raw_payload(source.get(F_RAW))],
        },
        ensure_ascii=False,
        sort_keys=True,
        default=str,
    )


def raw_payload(value: Any) -> Any:
    text = scalar_text(value)
    if not text:
        return None
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return text


def join_unique_text(left: Any, right: Any, separator: str = "; ") -> str:
    values: list[str] = []
    for item in (scalar_text(left), scalar_text(right)):
        if not item:
            continue
        for part in item.split(separator):
            part = part.strip()
            if part and part not in values:
                values.append(part)
    return separator.join(values)


def read_tabular_file(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        return read_csv_file(path)
    return read_excel_file(path)


def read_csv_file(path: Path) -> list[dict[str, Any]]:
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            with path.open("r", encoding=encoding, newline="") as fh:
                reader = csv.DictReader(fh)
                return [{str(k or "").strip(): v for k, v in row.items()} for row in reader if any(v for v in row.values())]
        except UnicodeError:
            continue
    raise UnicodeError(f"Cannot decode CSV file {path}")


def read_excel_file(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows


def douyin_export_row(source: dict[str, Any], path: Path) -> dict[str, Any] | None:
    order_no = clean_text(first_present(source, "主订单编号", "子订单编号", "订单id", "订单ID"))
    if not order_no:
        return None
    sub_order_no = clean_text(first_present(source, "子订单编号")) or order_no
    quantity = number(first_present(source, "商品数量", "数量"))
    paid_amount = number(first_present(source, "订单应付金额", "实收款", "支付金额"))
    refund_amount = douyin_refund_amount(source, paid_amount)
    unit_price = number(first_present(source, "商品单价", "单价"))
    return base_row(
        unique_key=canonical_unique_key("抖音", order_no),
        platform="抖音",
        data_source="抖店订单导出",
        order_no=order_no,
        created_at=order_created_at("抖音", source, order_no),
        product_name=clean_text(first_present(source, "选购商品", "商品名称")),
        quantity=quantity,
        unit_price=unit_price,
        paid_amount=paid_amount,
        refund_amount=refund_amount,
        freight_cost=number(first_present(source, "运费")) or 0,
        platform_fee=number(first_present(source, "手续费")) or 0,
        fulfill_status=clean_text(first_present(source, "售后状态", "承诺发货时间")),
        trade_status=clean_text(first_present(source, "订单状态")),
        operation="抖店订单导出导入",
        raw={"source_file": str(path), "row": sanitize_raw(source)},
    )


def douyin_refund_amount(source: dict[str, Any], paid_amount: float | None) -> float:
    explicit_refund = number(
        first_present(
            source,
            "退款金额",
            "已退款金额",
            "售后退款金额",
            "订单退款金额",
            "商品已退款金额",
        )
    )
    return refund_amount_from_fulfill_status(
        explicit_refund=explicit_refund,
        paid_amount=paid_amount,
        fulfill_status=clean_text(first_present(source, "履约/售后状态", "售后状态")),
    )


def refund_amount_from_fulfill_status(
    *,
    explicit_refund: float | None,
    paid_amount: float | None,
    fulfill_status: str,
) -> float:
    if "退款成功" in fulfill_status and (explicit_refund is None or explicit_refund == 0):
        return paid_amount or 0
    return explicit_refund or 0


def pdd_export_row(source: dict[str, Any], path: Path) -> dict[str, Any] | None:
    order_no = clean_text(first_present(source, "订单号", "订单编号", "订单ID"))
    if not order_no:
        return None
    quantity = number(first_present(source, "商品数量(件)", "商品数量", "数量"))
    paid_amount = number(first_present(source, "商家实收金额(元)", "用户实付金额(元)", "商品总价(元)", "实收款"))
    unit_price = None
    if paid_amount is not None and quantity not in (None, 0):
        unit_price = round(paid_amount / quantity, 2)
    order_status = clean_text(first_present(source, "订单状态"))
    aftersale_status = clean_text(first_present(source, "售后状态"))
    refund_amount = number(
        first_present(
            source,
            "退款金额",
            "已退款金额",
            "售后退款金额",
            "订单退款金额",
            "商品已退款金额",
        )
    )
    refund_amount = refund_amount_from_fulfill_status(
        explicit_refund=refund_amount,
        paid_amount=paid_amount,
        fulfill_status="/".join(item for item in (order_status, aftersale_status) if item),
    )
    return base_row(
        unique_key=canonical_unique_key("拼多多", order_no),
        platform="拼多多",
        data_source="拼多多订单导出",
        order_no=order_no,
        created_at=order_created_at("拼多多", source, order_no),
        product_name=clean_text(first_present(source, "商品", "商品名称")),
        quantity=quantity,
        unit_price=unit_price,
        paid_amount=paid_amount,
        refund_amount=refund_amount,
        freight_cost=number(first_present(source, "邮费(元)", "运费")) or 0,
        platform_fee=0,
        fulfill_status="/".join(item for item in (order_status, aftersale_status) if item),
        trade_status=order_status,
        operation="拼多多订单导出导入",
        raw={"source_file": str(path), "row": sanitize_raw(source)},
    )


def wechat_channels_export_row(source: dict[str, Any], path: Path) -> dict[str, Any] | None:
    order_no = clean_text(first_present(source, "订单号"))
    if not order_no:
        return None
    return base_row(
        unique_key=canonical_unique_key("视频号", order_no),
        platform="视频号",
        data_source="微信小店订单导出",
        order_no=order_no,
        created_at=order_created_at("视频号", source, order_no),
        product_name=clean_text(first_present(source, "商品名称")),
        quantity=number(first_present(source, "商品数量")),
        unit_price=number(first_present(source, "商品实际价格(单件)", "商品价格(单件)")),
        paid_amount=number(first_present(source, "订单实际收款金额", "订单实际支付金额", "商品实际价格(总共)")),
        refund_amount=number(first_present(source, "商品已退款金额")) or 0,
        freight_cost=number(first_present(source, "订单运费", "商品平均运费")) or 0,
        platform_fee=number(first_present(source, "技术服务费")) or 0,
        other_fee=number(first_present(source, "运费险预计投保费用")) or 0,
        fulfill_status=clean_text(first_present(source, "商品发货", "商品售后")),
        trade_status=clean_text(first_present(source, "订单状态")),
        operation="微信小店订单导出导入",
        raw={"source_file": str(path), "row": sanitize_raw(source)},
    )


def tmall_export_row(source: dict[str, Any], path: Path) -> dict[str, Any] | None:
    order_no = clean_text(first_present(source, "订单编号", "订单号", "主订单编号"))
    if not order_no:
        return None
    quantity = number(first_present(source, "宝贝总数量", "数量", "商品数量", "购买数量"))
    paid_amount = number(first_present(source, "买家实付金额", "实收款", "买家实付", "订单金额", "支付金额", "总金额"))
    refund_amount = number(first_present(source, "退款金额")) or 0
    if refund_amount > 0 and paid_amount is not None and paid_amount < refund_amount:
        paid_amount = round(paid_amount + refund_amount, 2)
    unit_price = number(first_present(source, "单价", "商品单价"))
    unit_price_basis = first_present(source, "买家应付货款", "总金额(旧版)", "买家实付金额", "实收款", "总金额")
    basis_amount = number(unit_price_basis)
    if (basis_amount is None or basis_amount == 0) and refund_amount > 0:
        basis_amount = refund_amount
    if unit_price is None and basis_amount is not None and quantity not in (None, 0):
        unit_price = round(basis_amount / quantity, 2)
    return base_row(
        unique_key=canonical_unique_key("天猫", order_no),
        platform="天猫",
        data_source="天猫订单导出",
        order_no=order_no,
        created_at=order_created_at("天猫", source, order_no),
        product_name=clean_text(first_present(source, "商品标题", "商品名称", "宝贝标题", "商品")),
        quantity=quantity,
        unit_price=unit_price,
        paid_amount=paid_amount,
        refund_amount=refund_amount,
        freight_cost=number(first_present(source, "买家应付邮费", "运费")) or 0,
        platform_fee=number(first_present(source, "卖家服务费")) or 0,
        fulfill_status=clean_text(first_present(source, "履约/售后状态", "售后状态")),
        trade_status=clean_text(first_present(source, "交易状态", "订单状态")),
        operation="天猫订单导出导入",
        raw={"source_file": str(path), "row": sanitize_raw(source)},
        shop_id=clean_text(first_present(source, "店铺ID")),
        shop_name=clean_text(first_present(source, "店铺名称")) or "天猫",
    )


def base_row(
    *,
    unique_key: str,
    platform: str,
    data_source: str,
    order_no: str,
    created_at: str,
    product_name: str,
    quantity: float | None,
    unit_price: float | None,
    paid_amount: float | None,
    refund_amount: float | None,
    freight_cost: float | None,
    platform_fee: float | None,
    fulfill_status: str,
    trade_status: str,
    operation: str,
    raw: dict[str, Any],
    product_cost: float | None = 0,
    other_fee: float | None = 0,
    shop_id: str = "",
    shop_name: str | None = None,
) -> dict[str, Any]:
    quantity = actual_sold_quantity(
        quantity=quantity,
        product_name=product_name,
        unit_price=unit_price,
        refund_amount=refund_amount,
        trade_status=trade_status,
        fulfill_status=fulfill_status,
    )
    return {
        F_UNIQUE_KEY: unique_key,
        F_PLATFORM: platform,
        F_DATA_SOURCE: data_source,
        F_SHOP_ID: shop_id,
        F_SHOP_NAME: shop_name or platform,
        F_FETCHED_AT: datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        F_ORDER_NO: order_no,
        F_CREATED_AT: created_at,
        F_BUYER_NICK: "",
        F_PRODUCT_NAME: product_name,
        F_UNIT_PRICE: unit_price,
        F_QUANTITY: quantity,
        F_FULFILL_STATUS: fulfill_status,
        F_TRADE_STATUS: trade_status,
        F_PAID_AMOUNT: paid_amount,
        F_REFUND_AMOUNT: refund_amount,
        F_PRODUCT_COST: product_cost,
        F_FREIGHT_COST: freight_cost,
        F_PLATFORM_FEE: platform_fee,
        F_OTHER_FEE: other_fee,
        F_OPERATION: operation,
        F_PAGE_URL: "",
        F_SCREENSHOT: "",
        F_RAW: json.dumps(raw, ensure_ascii=False, sort_keys=True, default=str),
    }


def fetch_api_orders(
    settings: Settings,
    allowed_platforms: set[str] | None = None,
) -> tuple[dict[str, Any], dict[str, list[dict[str, Any]]], set[date]]:
    targets = {
        "天猫": ("taobao", "JUSHUITAN_SHOP_ID_TMALL"),
        "抖音": ("doudian", "JUSHUITAN_SHOP_ID_DOUYIN"),
        "拼多多": ("pinduoduo", "JUSHUITAN_SHOP_ID_PINDUODUO"),
        "视频号": ("wechat_channels", "JUSHUITAN_SHOP_ID_WECHAT_CHANNELS"),
    }
    rows_by_platform: dict[str, list[dict[str, Any]]] = {platform: [] for platform in PLATFORMS}
    dates: set[date] = set()
    results: dict[str, Any] = {}
    for platform, (platform_code, env_name) in targets.items():
        if allowed_platforms is not None and platform not in allowed_platforms:
            results[platform] = {"status": "skipped_local_file_policy", "env": env_name}
            continue
        shop_id = os.getenv(env_name, "").strip()
        if not shop_id:
            results[platform] = {"status": "skipped_missing_env", "env": env_name}
            continue
        platform_settings = replace(
            settings,
            order_source="jushuitan",
            use_mock_collectors=False,
            shop_platform=platform_code,
            shop_id=shop_id,
            shop_name=platform,
            jushuitan_shop_ids=shop_id,
        )
        result = JushuitanOrderApiCollector(platform_settings).fetch_today()
        results[platform] = {
            "success": result.success,
            "shop_id": shop_id,
            "order_count": result.order_count,
            "total_amount": result.total_amount,
            "error_code": result.error_code,
            "error_message": result.error_message,
        }
        if not result.success:
            continue
        for order in result.orders or []:
            row = row_from_jushuitan_order(platform, order)
            rows_by_platform[platform].append(row)
            add_row_date(dates, row)
    return results, rows_by_platform, dates


def row_from_jushuitan_order(platform: str, order: dict[str, Any]) -> dict[str, Any]:
    raw = order.get("raw") if isinstance(order.get("raw"), dict) else {}
    order_no = jushuitan_platform_order_no(order, raw)
    product_name, quantity, unit_price = product_summary(raw)
    paid_amount = number(order.get("paid_amount"))
    if unit_price is None and paid_amount is not None and quantity not in (None, 0):
        unit_price = round(paid_amount / quantity, 2)
    return base_row(
        unique_key=canonical_unique_key(platform, order_no) or str(order.get("unique_key") or ""),
        platform=platform,
        data_source="聚水潭真实接口",
        order_no=order_no,
        created_at=str(order.get("created_at") or ""),
        product_name=product_name,
        quantity=quantity,
        unit_price=unit_price,
        paid_amount=paid_amount,
        refund_amount=0,
        freight_cost=0,
        platform_fee=0,
        fulfill_status=str(first_value(raw, "shipping_status", "send_status", "logistics_status") or ""),
        trade_status=str(order.get("order_status") or ""),
        operation="聚水潭API真实拉取",
        raw={"provider": "jushuitan", "row": sanitize_raw(raw)},
    ) | {
        F_SHOP_ID: str(order.get("shop_id") or ""),
        F_SHOP_NAME: str(order.get("shop_name") or platform),
        F_FETCHED_AT: str(order.get("fetched_at") or datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        F_BUYER_NICK: str(first_value(raw, "buyer_nick", "buyer_account", "buyer_id") or ""),
    }


def jushuitan_platform_order_no(order: dict[str, Any], raw: dict[str, Any]) -> str:
    return clean_text(
        first_value(
            raw,
            "so_id",
            "shop_order_id",
            "outer_so_id",
            "platform_order_no",
            "platform_order_id",
            "online_order_no",
            "tid",
            "order_sn",
            "order_no",
            "order_id",
            "o_id",
        )
        or order.get("order_id")
        or ""
    )


def product_summary(raw: dict[str, Any]) -> tuple[str, float | None, float | None]:
    items = first_list(raw, ("items", "order_items", "products", "skus", "details", "drp_co_id_froms"))
    if not items:
        return (
            str(first_value(raw, "product_name", "item_name", "sku_name", "goods_name") or ""),
            number(first_value(raw, "qty", "quantity", "num")),
            number(first_value(raw, "price", "unit_price", "sale_price")),
        )
    names: list[str] = []
    total_qty = 0.0
    first_price: float | None = None
    for item in items:
        if not isinstance(item, dict):
            continue
        name = first_value(item, "name", "product_name", "item_name", "sku_name", "goods_name")
        if name:
            names.append(str(name))
        total_qty += number(first_value(item, "qty", "quantity", "num")) or 0
        if first_price is None:
            first_price = number(first_value(item, "price", "unit_price", "sale_price"))
    return "; ".join(names), total_qty or None, first_price


def first_list(payload: dict[str, Any], keys: tuple[str, ...]) -> list[Any] | None:
    for key in keys:
        value = payload.get(key)
        if isinstance(value, list):
            return value
    return None


def upsert_rows_by_platform(
    client: FeishuClient,
    target_ids: dict[str, str],
    rows_by_platform: dict[str, list[dict[str, Any]]],
    product_rules: list[ProductRule] | None = None,
) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for platform in PLATFORMS:
        rows = rows_by_platform.get(platform) or []
        apply_product_breakdown_values(rows, product_rules or [])
        result[platform] = client.upsert_records(target_ids[platform], rows)
    return result


def apply_product_breakdown_values(rows: list[dict[str, Any]], product_rules: list[ProductRule]) -> None:
    if not product_rules:
        return
    for row in rows:
        row.update(
            product_breakdown_values(
                product_rules,
                product_name=row.get(F_PRODUCT_NAME),
                actual_quantity=row.get(F_QUANTITY),
                valid_sales=effective_sales_amount(row.get(F_PAID_AMOUNT), row.get(F_REFUND_AMOUNT)),
            )
        )


def cleanup_stale_import_keys(client: FeishuClient, target_ids: dict[str, str]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for platform in PLATFORMS:
        prefixes = STALE_IMPORT_KEY_PREFIXES.get(platform, ())
        deleted = client.delete_records_by_unique_key_prefixes(target_ids[platform], prefixes)
        result[platform] = {"prefixes": list(prefixes), "deleted": deleted}
    return result


def cleanup_non_primary_order_records(
    client: FeishuClient,
    target_ids: dict[str, str],
    primary_rows_by_platform: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for platform in PLATFORMS:
        allowed_keys = {
            scalar_text(row.get(F_UNIQUE_KEY))
            for row in primary_rows_by_platform.get(platform, [])
            if scalar_text(row.get(F_UNIQUE_KEY))
        }
        seen_kept_keys: set[str] = set()
        delete_record_ids: list[str] = []
        reasons: Counter[str] = Counter()
        scanned = 0
        for record in client.list_records(target_ids[platform]):
            scanned += 1
            fields = record.get("fields") or {}
            unique_key = scalar_text(fields.get(F_UNIQUE_KEY))
            order_no = clean_order_no(fields.get(F_ORDER_NO))
            canonical_key = canonical_unique_key(platform, order_no)
            reason = ""
            if not unique_key or not order_no:
                reason = "missing_unique_key_or_order_no"
            elif unique_key not in allowed_keys:
                reason = "not_in_latest_local_export"
            elif unique_key != canonical_key:
                reason = "noncanonical_unique_key"
            elif unique_key in seen_kept_keys:
                reason = "duplicate_canonical_unique_key"
            if reason:
                record_id = record.get("record_id")
                if record_id:
                    delete_record_ids.append(str(record_id))
                    reasons[reason] += 1
            else:
                seen_kept_keys.add(unique_key)
        deleted = client.delete_records(target_ids[platform], delete_record_ids)
        result[platform] = {
            "scanned": scanned,
            "latest_local_orders": len(allowed_keys),
            "kept": len(seen_kept_keys),
            "deleted": deleted,
            "delete_reasons": dict(reasons),
        }
    return result


def audit_order_duplicates(client: FeishuClient, target_ids: dict[str, str]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for platform in PLATFORMS:
        order_counts: Counter[str] = Counter()
        key_counts: Counter[str] = Counter()
        noncanonical = 0
        total = 0
        samples: list[dict[str, Any]] = []
        for record in client.list_records(target_ids[platform]):
            total += 1
            fields = record.get("fields") or {}
            order_no = clean_order_no(fields.get(F_ORDER_NO))
            unique_key = scalar_text(fields.get(F_UNIQUE_KEY))
            if order_no:
                order_counts[order_no] += 1
            if unique_key:
                key_counts[unique_key] += 1
            if order_no and unique_key and unique_key != canonical_unique_key(platform, order_no):
                noncanonical += 1
                if len(samples) < 10:
                    samples.append(
                        {
                            "record_id": record.get("record_id"),
                            "unique_key": unique_key,
                            "order_no": order_no,
                            "expected_unique_key": canonical_unique_key(platform, order_no),
                        }
                    )
        duplicate_orders = {key: count for key, count in order_counts.items() if count > 1}
        duplicate_keys = {key: count for key, count in key_counts.items() if count > 1}
        result[platform] = {
            "total_records": total,
            "unique_orders": len(order_counts),
            "duplicate_order_count": len(duplicate_orders),
            "duplicate_order_extra_rows": sum(count - 1 for count in duplicate_orders.values()),
            "duplicate_unique_key_count": len(duplicate_keys),
            "duplicate_unique_key_extra_rows": sum(count - 1 for count in duplicate_keys.values()),
            "noncanonical_unique_key_rows": noncanonical,
            "noncanonical_samples": samples,
            "duplicate_order_samples": dict(list(duplicate_orders.items())[:10]),
            "duplicate_unique_key_samples": dict(list(duplicate_keys.items())[:10]),
        }
    return result


def verify_tables(client: FeishuClient, target_ids: dict[str, str], summary_table_id: str) -> dict[str, Any]:
    expected_fields = [name for name, _ in UNIFIED_FIELDS] + list(FORMULA_FIELDS)
    field_signatures: dict[str, list[str]] = {}
    counts: dict[str, int] = {}
    for platform, table_id in target_ids.items():
        field_names = list(client.field_index(table_id))
        field_signatures[platform] = field_names
        counts[platform] = client.table_count(table_id)
    missing_fields = {
        platform: sorted(set(expected_fields) - set(names))
        for platform, names in field_signatures.items()
    }
    extra_visible_fields = {
        platform: sorted(set(names) - set(expected_fields))
        for platform, names in field_signatures.items()
    }
    summary_sample = [
        record.get("fields") or {}
        for record in client.list_records(summary_table_id)
        if scalar_text((record.get("fields") or {}).get("统计日期")) >= date.today().isoformat()
    ][:20]
    return {
        "target_counts": counts,
        "field_sets_identical": len({tuple(sorted(names)) for names in field_signatures.values()}) == 1,
        "missing_fields": missing_fields,
        "extra_visible_fields": extra_visible_fields,
        "summary_today_or_later_sample": summary_sample,
    }


def clean_feishu_fields(row: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in row.items() if value not in (None, "")}


def normalize_platform(value: Any) -> str:
    text = scalar_text(value)
    lower = text.lower()
    if "抖音" in text or "douyin" in lower or "doudian" in lower:
        return "抖音"
    if "拼多多" in text or "pdd" in lower or "pinduoduo" in lower:
        return "拼多多"
    if "视频号" in text or "微信" in text or "wechat" in lower:
        return "视频号"
    if "天猫" in text or "淘宝" in text or "千牛" in text or "tmall" in lower or "taobao" in lower:
        return "天猫"
    return text or "未知平台"


def canonical_unique_key(platform: str, order_no: Any) -> str:
    cleaned = clean_order_no(order_no)
    if not cleaned:
        return ""
    code = PLATFORM_CODES.get(platform) or platform.lower()
    return f"{code}_{cleaned}"


def clean_order_no(value: Any) -> str:
    return scalar_text(value).replace("\t", "").strip().strip("'")


def scalar_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        parts: list[str] = []
        for item in value:
            if isinstance(item, dict) and "text" in item:
                parts.append(str(item.get("text") or ""))
            else:
                parts.append(str(item))
        return "".join(parts).strip()
    return str(value).strip()


def number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    text = scalar_text(value)
    if text in {"", "-", "--"}:
        return None
    text = text.replace(",", "").replace("元", "").replace("\t", "").strip()
    if text.endswith("%"):
        text = text[:-1]
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def first_present(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


ORDER_CREATED_AT_ALIASES: dict[str, tuple[str, ...]] = {
    "天猫": ("订单创建时间", "创建时间", "下单时间", "订单下单时间", "订单付款时间", "支付时间"),
    "抖音": ("订单提交时间", "下单时间", "订单下单时间", "创建时间", "支付完成时间", "付款时间", "支付时间"),
    "拼多多": ("订单成交时间", "下单时间", "订单创建时间", "创建时间", "支付时间"),
    "视频号": ("订单下单时间", "下单时间", "订单创建时间", "创建时间", "支付时间"),
}


def order_created_at(platform: str, row: dict[str, Any], order_no: str) -> str:
    value = normalize_datetime_text(first_present(row, *ORDER_CREATED_AT_ALIASES.get(platform, ("创建时间", "下单时间"))))
    if value:
        return value
    if platform == "拼多多":
        return pdd_date_from_order_no(order_no)
    return ""


NON_SOLD_STATUS_KEYWORDS = ("退款", "交易关闭", "已关闭", "已取消", "订单关闭", "待付款", "等待买家付款", "未付款")
NON_SOLD_PRODUCT_KEYWORDS = ("补收差价", "差价专用", "购买前须联系客服", "联系客服确认")


def actual_sold_quantity(
    *,
    quantity: float | None,
    product_name: str,
    unit_price: float | None,
    refund_amount: float | None,
    trade_status: str,
    fulfill_status: str,
) -> float | None:
    if refund_amount and refund_amount > 0:
        return 0
    text = f"{trade_status}/{fulfill_status}".replace("无售后或售后取消", "")
    if any(keyword in text for keyword in NON_SOLD_STATUS_KEYWORDS):
        return 0
    if unit_price == 0 or any(keyword in product_name for keyword in NON_SOLD_PRODUCT_KEYWORDS):
        return 0
    return quantity


def pdd_date_from_order_no(order_no: str) -> str:
    prefix = clean_text(order_no).split("-", 1)[0]
    if len(prefix) != 6 or not prefix.isdigit():
        return ""
    month = int(prefix[2:4])
    day = int(prefix[4:6])
    if not (1 <= month <= 12 and 1 <= day <= 31):
        return ""
    return f"20{prefix[:2]}-{month:02d}-{day:02d}"


def clean_text(value: Any) -> str:
    if value in (None, "-", "--"):
        return ""
    return str(value).strip().strip("\t").rstrip("\t")


def sanitize_raw(value: Any) -> Any:
    if isinstance(value, dict):
        result: dict[str, Any] = {}
        for key, item in value.items():
            lowered = str(key).lower()
            if any(part in lowered or part in str(key) for part in SENSITIVE_KEY_PARTS):
                result[key] = "[REDACTED]"
            else:
                result[key] = sanitize_raw(item)
        return result
    if isinstance(value, list):
        return [sanitize_raw(item) for item in value]
    return value


def add_row_date(dates: set[date], row: dict[str, Any]) -> None:
    for key in (F_CREATED_AT, F_FETCHED_AT):
        parsed = parse_date(row.get(key))
        if parsed:
            dates.add(parsed)
            return


def parse_date(value: Any) -> date | None:
    text = scalar_text(value)
    if not text:
        return None
    text = text.replace("/", "-")
    for candidate in (text, text[:19], text[:16], text[:10]):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                continue
    return None


def normalize_datetime_text(value: Any) -> str:
    text = scalar_text(value).strip()
    if not text:
        return ""
    normalized = text.replace("/", "-")
    for candidate in (normalized, normalized[:19], normalized[:16], normalized[:10]):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(candidate, fmt)
                if fmt == "%Y-%m-%d":
                    return parsed.strftime("%Y-%m-%d")
                return parsed.strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
    return text


def chunks(values: list[dict[str, Any]], size: int) -> list[list[dict[str, Any]]]:
    return [values[index : index + size] for index in range(0, len(values), size)]


def log(message: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {message}", flush=True)


def main() -> int:
    parser = argparse.ArgumentParser(description="Split ShopOps order rows into four unified Feishu platform tables.")
    parser.add_argument("--source-table-id", default=os.getenv("SHOPOPS_ORDER_TABLE_ID") or SOURCE_ORDER_TABLE_ID)
    parser.add_argument("--summary-table-id", default=os.getenv("SHOPOPS_FORMULA_SUMMARY_TABLE_ID") or SUMMARY_TABLE_ID)
    parser.add_argument("--data-root", default=r"D:\lyh\ShopOps")
    parser.add_argument("--env-path", default=".env")
    parser.add_argument("--evidence-dir", default="docs/live-evidence/platform-order-split")
    parser.add_argument("--skip-source-split", action="store_true")
    parser.add_argument("--skip-local-import", action="store_true")
    parser.add_argument("--skip-api-fetch", action="store_true")
    parser.add_argument("--skip-source-policy-cleanup", action="store_true")
    parser.add_argument(
        "--write-api-supplement",
        action="store_true",
        help="Write API rows into order tables. Default is comparison-only because local exports are the primary source.",
    )
    parser.add_argument(
        "--no-local-fallback",
        action="store_true",
        help="Deprecated; local import now follows the fixed source policy.",
    )
    args = parser.parse_args()

    _load_dotenv()
    settings = load_settings()
    client = FeishuClient(settings)
    evidence_dir = Path(args.evidence_dir)
    evidence_dir.mkdir(parents=True, exist_ok=True)

    target_ids = client.ensure_order_tables(Path(args.env_path))
    product_table_id = os.getenv("SHOPOPS_PRODUCT_CATALOG_TABLE_ID", DEFAULT_PRODUCT_CATALOG_TABLE_ID).strip()
    product_rules = client.product_rules(product_table_id)
    product_order_fields = {
        platform: client.ensure_product_breakdown_fields(table_id, product_rules)
        for platform, table_id in target_ids.items()
    }
    all_dates: set[date] = {date.today()}
    result: dict[str, Any] = {
        "status": "running",
        "app_token": client.app_token,
        "source_table_id": args.source_table_id,
        "summary_table_id": args.summary_table_id,
        "target_tables": {
            platform: {"table_id": table_id, "name": TARGET_TABLES[platform]["name"]}
            for platform, table_id in target_ids.items()
        },
        "source_policy": {
            "mode": "local_export_primary",
            "api_platforms": sorted(API_ORDER_PLATFORMS),
            "local_file_platforms": sorted(LOCAL_ORDER_PLATFORMS),
            "api_write": bool(args.write_api_supplement),
        },
        "product_breakdown": {
            "product_table_id": product_table_id,
            "products": [rule.name for rule in product_rules],
            "order_formula_fields": product_order_fields,
        },
    }

    if args.skip_source_policy_cleanup:
        result["source_policy_cleanup"] = {"status": "skipped"}
    else:
        result["source_policy_cleanup"] = cleanup_stale_import_keys(client, target_ids)

    if args.skip_source_split:
        result["source_split"] = {"status": "skipped"}
    else:
        split_result, split_dates = split_source_table(client, args.source_table_id, target_ids)
        all_dates |= split_dates
        result["source_split"] = split_result

    if args.skip_api_fetch:
        result["api_fetch"] = {"status": "skipped"}
    else:
        api_fetch, api_rows, api_dates = fetch_api_orders(settings, API_ORDER_PLATFORMS)
        if args.write_api_supplement:
            api_upsert: dict[str, Any] = upsert_rows_by_platform(client, target_ids, api_rows, product_rules)
            all_dates |= api_dates
        else:
            api_upsert = {"status": "skipped_local_export_primary"}
        result["api_fetch"] = {"platforms": api_fetch, "upsert": api_upsert}

    if args.skip_local_import:
        result["local_import"] = {"status": "skipped"}
    else:
        local_imports, local_rows, local_dates = import_latest_local_files(Path(args.data_root), LOCAL_ORDER_PLATFORMS)
        local_upsert = upsert_rows_by_platform(client, target_ids, local_rows, product_rules)
        all_dates |= local_dates
        result["local_import"] = {
            "policy": "local_export_primary",
            "local_file_platforms": sorted(LOCAL_ORDER_PLATFORMS),
            "files": local_imports,
            "upsert": local_upsert,
        }
        result["local_primary_cleanup"] = cleanup_non_primary_order_records(client, target_ids, local_rows)

    normalized_rows = client.normalize_existing_summary_platforms(args.summary_table_id)
    client.update_ad_platform_formula(settings.shopops_ad_table_id)
    client.update_formula_summary(target_ids, args.summary_table_id, product_rules)
    dimension_upsert = client.upsert_summary_dimension_rows(args.summary_table_id, all_dates)
    time.sleep(8)
    result["summary_update"] = {
        "normalized_old_taobao_rows": normalized_rows,
        "dimension_dates": len(all_dates),
        "dimension_upsert": dimension_upsert,
        "order_formula_fields": list(ORDER_SUMMARY_FIELDS),
    }
    result["verification"] = verify_tables(client, target_ids, args.summary_table_id)
    result["order_duplicate_audit"] = audit_order_duplicates(client, target_ids)
    result["status"] = "success"
    evidence_path = evidence_dir / "platform-order-split-result.json"
    evidence_path.write_text(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True, default=str), encoding="utf-8")
    result["evidence_path"] = str(evidence_path.resolve())
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
