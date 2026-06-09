from __future__ import annotations

import argparse
import json
import os
import sys
from collections import Counter
from dataclasses import replace
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from scripts.run_jushuitan_orders_to_feishu import FeishuOrderWriter, feishu_order_fields
from shopops.collectors.jushuitan_order_api import JushuitanOrderApiCollector
from shopops.config import Settings, load_settings
from shopops.storage.feishu_bootstrap import FEISHU_BASE_URL, FeishuOpenApiClient


TEXT_FIELD = 1
NUMBER_FIELD = 2

TARGET_TABLE_ENV = "FEISHU_TABLE_DOUYIN_INFLUENCER_COMMISSION"
WECHAT_SHOP_ENV = "JUSHUITAN_SHOP_ID_WECHAT_CHANNELS"

F_UNIQUE_KEY = "unique_key"
F_PLATFORM = "\u5e73\u53f0"
F_SOURCE = "\u6570\u636e\u6765\u6e90"
F_ORDER_NO = "\u8ba2\u5355\u53f7"
F_ORDER_CREATED_AT = "\u8ba2\u5355\u4e0b\u5355\u65f6\u95f4"
F_PAY_AT = "\u652f\u4ed8\u65f6\u95f4"
F_ORDER_STATUS = "\u8ba2\u5355\u72b6\u6001"
F_TRADE_NO = "\u4ea4\u6613\u5355\u53f7"
F_PRODUCT_NAME = "\u5546\u54c1\u540d\u79f0"
F_PLATFORM_PRODUCT_CODE = "\u5546\u54c1\u7f16\u7801(\u5e73\u53f0)"
F_CUSTOM_PRODUCT_CODE = "\u5546\u54c1\u7f16\u7801(\u81ea\u5b9a\u4e49)"
F_CUSTOM_SKU_CODE = "SKU\u7f16\u7801(\u81ea\u5b9a\u4e49)"
F_PRODUCT_ATTR = "\u5546\u54c1\u5c5e\u6027"
F_QUANTITY = "\u5546\u54c1\u6570\u91cf"
F_PRODUCT_TOTAL = "\u5546\u54c1\u603b\u4ef7"
F_PAID_AMOUNT = "\u8ba2\u5355\u5b9e\u9645\u652f\u4ed8\u91d1\u989d"
F_RECEIVED_AMOUNT = "\u8ba2\u5355\u5b9e\u9645\u6536\u6b3e\u91d1\u989d"
F_SERVICE_FEE = "\u6280\u672f\u670d\u52a1\u8d39"
F_SHIPPING_INSURANCE = "\u8fd0\u8d39\u9669\u9884\u8ba1\u6295\u4fdd\u8d39\u7528"
F_INFLUENCER_MODE = "\u5e26\u8d27\u65b9\u5f0f"
F_INFLUENCER_ACCOUNT_TYPE = "\u5e26\u8d27\u8d26\u53f7\u7c7b\u578b"
F_INFLUENCER_NICK = "\u5e26\u8d27\u8d26\u53f7\u6635\u79f0"
F_INFLUENCER_FEE_CHANNEL = "\u5e26\u8d27\u8d39\u7528\u6e20\u9053"
F_INFLUENCER_FEE_TYPE = "\u5e26\u8d27\u8d39\u7528\u7c7b\u578b"
F_INFLUENCER_FEE = "\u5e26\u8d27\u8d39\u7528"
F_COMMISSION_RATE = "\u5e26\u8d27\u4f63\u91d1\u7387"
F_JST_MATCH_STATUS = "\u805a\u6c34\u6f6d\u5339\u914d\u72b6\u6001"
F_JST_ORDER_NO = "\u805a\u6c34\u6f6d\u8ba2\u5355\u53f7"
F_JST_ORDER_STATUS = "\u805a\u6c34\u6f6d\u8ba2\u5355\u72b6\u6001"
F_JST_SHOP_ID = "\u805a\u6c34\u6f6d\u5e97\u94faID"
F_FETCHED_AT = "\u91c7\u96c6\u65f6\u95f4"
F_SOURCE_FILE = "\u6765\u6e90\u6587\u4ef6"
F_RAW = "\u539f\u59cb\u6570\u636e"

TARGET_FIELDS = [
    (F_UNIQUE_KEY, TEXT_FIELD),
    (F_PLATFORM, TEXT_FIELD),
    (F_SOURCE, TEXT_FIELD),
    (F_ORDER_NO, TEXT_FIELD),
    (F_ORDER_CREATED_AT, TEXT_FIELD),
    (F_PAY_AT, TEXT_FIELD),
    (F_ORDER_STATUS, TEXT_FIELD),
    (F_TRADE_NO, TEXT_FIELD),
    (F_PRODUCT_NAME, TEXT_FIELD),
    (F_PLATFORM_PRODUCT_CODE, TEXT_FIELD),
    (F_CUSTOM_PRODUCT_CODE, TEXT_FIELD),
    (F_CUSTOM_SKU_CODE, TEXT_FIELD),
    (F_PRODUCT_ATTR, TEXT_FIELD),
    (F_QUANTITY, NUMBER_FIELD),
    (F_PRODUCT_TOTAL, NUMBER_FIELD),
    (F_PAID_AMOUNT, NUMBER_FIELD),
    (F_RECEIVED_AMOUNT, NUMBER_FIELD),
    (F_SERVICE_FEE, NUMBER_FIELD),
    (F_SHIPPING_INSURANCE, NUMBER_FIELD),
    (F_INFLUENCER_MODE, TEXT_FIELD),
    (F_INFLUENCER_ACCOUNT_TYPE, TEXT_FIELD),
    (F_INFLUENCER_NICK, TEXT_FIELD),
    (F_INFLUENCER_FEE_CHANNEL, TEXT_FIELD),
    (F_INFLUENCER_FEE_TYPE, TEXT_FIELD),
    (F_INFLUENCER_FEE, NUMBER_FIELD),
    (F_COMMISSION_RATE, TEXT_FIELD),
    (F_JST_MATCH_STATUS, TEXT_FIELD),
    (F_JST_ORDER_NO, TEXT_FIELD),
    (F_JST_ORDER_STATUS, TEXT_FIELD),
    (F_JST_SHOP_ID, TEXT_FIELD),
    (F_FETCHED_AT, TEXT_FIELD),
    (F_SOURCE_FILE, TEXT_FIELD),
    (F_RAW, TEXT_FIELD),
]

SENSITIVE_COLUMNS = {
    "\u6536\u4ef6\u4eba\u59d3\u540d",
    "\u6536\u4ef6\u4eba\u5730\u5740",
    "\u6536\u4ef6\u4eba\u624b\u673a",
}


class FeishuTableClient:
    def __init__(self, settings: Settings, table_id: str) -> None:
        self.settings = settings
        self.app_token = settings.shopops_data_center_app_token or settings.feishu_app_token
        self.table_id = table_id
        self.client = FeishuOpenApiClient(settings.feishu_app_id, settings.feishu_app_secret)

    def request(
        self,
        method: str,
        path: str,
        payload: dict[str, Any] | None = None,
        params: dict[str, Any] | None = None,
        allow_duplicate: bool = False,
    ) -> dict[str, Any]:
        response = requests.request(
            method,
            f"{FEISHU_BASE_URL}{path}",
            headers=self.client.headers(),
            json=payload,
            params=params,
            timeout=30,
        )
        body = response.json()
        if allow_duplicate and body.get("code") == 1254014:
            return {}
        if response.status_code >= 400 or body.get("code") != 0:
            raise RuntimeError(f"Feishu API {method} {path} failed HTTP {response.status_code}: {body}")
        return body.get("data") or {}

    def ensure_fields(self) -> list[str]:
        existing = self.field_names()
        created: list[str] = []
        for field_name, field_type in TARGET_FIELDS:
            if field_name in existing:
                continue
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/fields",
                {"field_name": field_name, "type": field_type},
                allow_duplicate=True,
            )
            existing.add(field_name)
            created.append(field_name)
        return created

    def field_names(self) -> set[str]:
        names: set[str] = set()
        page_token = None
        while True:
            params: dict[str, Any] = {"page_size": 100}
            if page_token:
                params["page_token"] = page_token
            data = self.request(
                "GET",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/fields",
                params=params,
            )
            for item in data.get("items", []) or []:
                if item.get("field_name"):
                    names.add(str(item["field_name"]))
            if not data.get("has_more"):
                return names
            page_token = data.get("page_token")

    def existing_records_by_unique_key(self) -> dict[str, str]:
        records: dict[str, str] = {}
        page_token = None
        page = 0
        while True:
            page += 1
            params: dict[str, Any] = {"page_size": 500}
            if page_token:
                params["page_token"] = page_token
            data = self.request(
                "GET",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records",
                params=params,
            )
            for item in data.get("items", []) or []:
                fields = item.get("fields") or {}
                unique_key = fields.get(F_UNIQUE_KEY)
                if unique_key:
                    records[str(unique_key)] = str(item.get("record_id"))
            log(f"loaded existing target records page={page} total_keys={len(records)}")
            if not data.get("has_more"):
                return records
            page_token = data.get("page_token")

    def upsert_rows(self, rows: list[dict[str, Any]]) -> dict[str, int]:
        existing = self.existing_records_by_unique_key()
        to_create: list[dict[str, Any]] = []
        to_update: list[dict[str, Any]] = []
        for row in rows:
            clean = {key: value for key, value in row.items() if value not in (None, "")}
            record_id = existing.get(str(row[F_UNIQUE_KEY]))
            if record_id:
                to_update.append({"record_id": record_id, "fields": clean})
            else:
                to_create.append({"fields": clean})

        for chunk in chunks(to_create, 500):
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records/batch_create",
                {"records": chunk},
            )
        for chunk in chunks(to_update, 500):
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records/batch_update",
                {"records": chunk},
            )
        return {"created": len(to_create), "updated": len(to_update), "saved": len(to_create) + len(to_update)}

    def readback_unique_keys(self, unique_keys: set[str]) -> dict[str, dict[str, Any]]:
        found: dict[str, dict[str, Any]] = {}
        page_token = None
        page = 0
        while True:
            page += 1
            params: dict[str, Any] = {"page_size": 500}
            if page_token:
                params["page_token"] = page_token
            data = self.request(
                "GET",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records",
                params=params,
            )
            for item in data.get("items", []) or []:
                fields = item.get("fields") or {}
                unique_key = fields.get(F_UNIQUE_KEY)
                if unique_key in unique_keys:
                    found[str(unique_key)] = fields
            log(f"readback target records page={page} found={len(found)}/{len(unique_keys)}")
            if not data.get("has_more"):
                return found
            page_token = data.get("page_token")


def settings_for_wechat_channels(settings: Settings) -> Settings:
    shop_id = os.getenv(WECHAT_SHOP_ENV, "").strip()
    return replace(
        settings,
        order_source="jushuitan",
        use_mock_collectors=False,
        shop_platform="wechat_channels",
        shop_id=shop_id,
        shop_name="\u89c6\u9891\u53f7",
        jushuitan_shop_ids=shop_id,
    )


def fetch_wechat_channels_orders(settings: Settings) -> tuple[Any, int]:
    platform_settings = settings_for_wechat_channels(settings)
    log("fetching Jushuitan wechat_channels orders")
    result = JushuitanOrderApiCollector(platform_settings).fetch_today()
    saved_count = 0
    if result.success and result.orders:
        log(f"writing {len(result.orders)} Jushuitan orders to Feishu order table")
        writer = FeishuOrderWriter(settings)
        writer.ensure_fields()
        rows = [feishu_order_fields("\u89c6\u9891\u53f7", order) for order in result.orders]
        saved_count = writer.write_orders(rows)
    log(f"Jushuitan done success={result.success} orders={result.order_count} saved={saved_count}")
    return result, saved_count


def load_excel_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows


def influencer_rows_from_excel(excel_rows: list[dict[str, Any]], source_file: Path, jst_orders: list[dict[str, Any]]) -> list[dict[str, Any]]:
    fetched_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    jst_by_order_no: dict[str, dict[str, Any]] = {}
    for order in jst_orders:
        order_no = str(order.get("order_id") or "")
        if order_no:
            jst_by_order_no[order_no] = order

    order_counter = Counter(str(value(row, F_ORDER_NO) or "") for row in excel_rows)
    sequence_by_order: Counter[str] = Counter()
    rows: list[dict[str, Any]] = []
    for row in excel_rows:
        order_no = str(value(row, F_ORDER_NO) or "").strip()
        if not order_no:
            continue
        sequence_by_order[order_no] += 1
        suffix = "" if order_counter[order_no] == 1 else f"_{sequence_by_order[order_no]}"
        unique_key = f"shipin_{order_no}{suffix}"
        jst_order = jst_by_order_no.get(order_no)
        rows.append(
            {
                F_UNIQUE_KEY: unique_key,
                F_PLATFORM: "\u89c6\u9891\u53f7",
                F_SOURCE: "\u5fae\u4fe1\u5c0f\u5e97\u5bfc\u51faExcel",
                F_ORDER_NO: order_no,
                F_ORDER_CREATED_AT: text_value(row, F_ORDER_CREATED_AT),
                F_PAY_AT: text_value(row, F_PAY_AT),
                F_ORDER_STATUS: text_value(row, F_ORDER_STATUS),
                F_TRADE_NO: text_value(row, F_TRADE_NO),
                F_PRODUCT_NAME: text_value(row, F_PRODUCT_NAME),
                F_PLATFORM_PRODUCT_CODE: text_value(row, F_PLATFORM_PRODUCT_CODE),
                F_CUSTOM_PRODUCT_CODE: text_value(row, F_CUSTOM_PRODUCT_CODE),
                F_CUSTOM_SKU_CODE: text_value(row, F_CUSTOM_SKU_CODE),
                F_PRODUCT_ATTR: text_value(row, F_PRODUCT_ATTR),
                F_QUANTITY: number_value(row, F_QUANTITY),
                F_PRODUCT_TOTAL: number_value(row, F_PRODUCT_TOTAL),
                F_PAID_AMOUNT: number_value(row, F_PAID_AMOUNT),
                F_RECEIVED_AMOUNT: number_value(row, F_RECEIVED_AMOUNT),
                F_SERVICE_FEE: number_value(row, F_SERVICE_FEE),
                F_SHIPPING_INSURANCE: number_value(row, F_SHIPPING_INSURANCE),
                F_INFLUENCER_MODE: text_value(row, F_INFLUENCER_MODE),
                F_INFLUENCER_ACCOUNT_TYPE: text_value(row, F_INFLUENCER_ACCOUNT_TYPE),
                F_INFLUENCER_NICK: text_value(row, F_INFLUENCER_NICK),
                F_INFLUENCER_FEE_CHANNEL: text_value(row, F_INFLUENCER_FEE_CHANNEL),
                F_INFLUENCER_FEE_TYPE: text_value(row, F_INFLUENCER_FEE_TYPE),
                F_INFLUENCER_FEE: number_value(row, F_INFLUENCER_FEE),
                F_COMMISSION_RATE: text_value(row, F_COMMISSION_RATE),
                F_JST_MATCH_STATUS: "\u5df2\u5339\u914d" if jst_order else "\u672a\u5339\u914d",
                F_JST_ORDER_NO: str(jst_order.get("order_id") or "") if jst_order else "",
                F_JST_ORDER_STATUS: str(jst_order.get("order_status") or "") if jst_order else "",
                F_JST_SHOP_ID: str(jst_order.get("shop_id") or "") if jst_order else "",
                F_FETCHED_AT: fetched_at,
                F_SOURCE_FILE: str(source_file),
                F_RAW: json.dumps(redact_excel_row(row), ensure_ascii=False, sort_keys=True, default=str),
            }
        )
    return rows


def value(row: dict[str, Any], key: str) -> Any:
    return row.get(key)


def text_value(row: dict[str, Any], key: str) -> str:
    raw = value(row, key)
    if raw in (None, "-"):
        return ""
    return str(raw)


def number_value(row: dict[str, Any], key: str) -> float | None:
    raw = value(row, key)
    if raw in (None, "", "-"):
        return None
    return round(float(raw), 2)


def redact_excel_row(row: dict[str, Any]) -> dict[str, Any]:
    redacted: dict[str, Any] = {}
    for key, item in row.items():
        if key in SENSITIVE_COLUMNS:
            redacted[key] = "[REDACTED]"
        else:
            redacted[key] = item
    return redacted


def chunks(values: list[dict[str, Any]], size: int) -> list[list[dict[str, Any]]]:
    return [values[index : index + size] for index in range(0, len(values), size)]


def log(message: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {message}", flush=True)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--target-table", default="")
    parser.add_argument("--skip-jushuitan", action="store_true")
    parser.add_argument("--include-no-influencer", action="store_true")
    args = parser.parse_args()

    settings = load_settings()
    target_table = args.target_table or os.getenv(TARGET_TABLE_ENV, "").strip()
    if not target_table:
        raise RuntimeError(f"Missing {TARGET_TABLE_ENV} or --target-table")
    if not os.getenv(WECHAT_SHOP_ENV, "").strip():
        raise RuntimeError(f"Missing {WECHAT_SHOP_ENV}")

    jst_result = None
    jst_saved_count = 0
    if args.skip_jushuitan:
        log("skipping Jushuitan fetch")
        jst_orders: list[dict[str, Any]] = []
    else:
        jst_result, jst_saved_count = fetch_wechat_channels_orders(settings)
        if not jst_result.success:
            raise RuntimeError(f"Jushuitan fetch failed: {jst_result.error_code} {jst_result.error_message}")
        jst_orders = jst_result.orders or []

    excel_path = Path(args.excel)
    log(f"loading Excel {excel_path}")
    excel_rows = load_excel_rows(excel_path)
    no_influencer_rows = [row for row in excel_rows if not is_influencer_excel_row(row)]
    import_excel_rows = excel_rows if args.include_no_influencer else [row for row in excel_rows if is_influencer_excel_row(row)]
    log(
        f"Excel rows={len(excel_rows)} influencer_rows={len(import_excel_rows)} "
        f"no_influencer_rows={len(no_influencer_rows)}"
    )
    influencer_rows = influencer_rows_from_excel(import_excel_rows, excel_path, jst_orders)

    client = FeishuTableClient(settings, target_table)
    log(f"ensuring target table fields table={target_table}")
    created_fields = client.ensure_fields()
    log(f"upserting {len(influencer_rows)} influencer rows")
    upsert_summary = client.upsert_rows(influencer_rows)
    log("reading back written unique keys")
    readback = client.readback_unique_keys({row[F_UNIQUE_KEY] for row in influencer_rows})
    missing_keys = sorted(set(row[F_UNIQUE_KEY] for row in influencer_rows) - set(readback))
    mismatched: list[str] = []
    for row in influencer_rows:
        fields = readback.get(row[F_UNIQUE_KEY])
        if not fields:
            continue
        if str(fields.get(F_ORDER_NO) or "") != str(row[F_ORDER_NO]):
            mismatched.append(row[F_UNIQUE_KEY])

    summary = {
        "status": "success" if not missing_keys and not mismatched else "readback_mismatch",
        "target_table": target_table,
        "excel_file": str(excel_path),
        "excel_row_count": len(excel_rows),
        "skipped_no_influencer_count": 0 if args.include_no_influencer else len(no_influencer_rows),
        "skipped_no_influencer_order_nos": [str(value(row, F_ORDER_NO) or "") for row in no_influencer_rows],
        "influencer_row_count": len(influencer_rows),
        "unique_key_prefix": "shipin_",
        "created_fields": created_fields,
        "upsert": upsert_summary,
        "readback_count": len(readback),
        "missing_unique_keys": missing_keys,
        "mismatched_unique_keys": mismatched,
        "jushuitan": {
            "fetched": not args.skip_jushuitan,
            "success": bool(jst_result.success) if jst_result else None,
            "shop_id": os.getenv(WECHAT_SHOP_ENV, "").strip(),
            "order_count": jst_result.order_count if jst_result else None,
            "saved_to_order_table_count": jst_saved_count,
            "total_amount": jst_result.total_amount if jst_result else None,
            "error_code": jst_result.error_code if jst_result else None,
            "error_message": jst_result.error_message if jst_result else None,
        },
        "influencer_nicks": dict(Counter(row.get(F_INFLUENCER_NICK) or "" for row in influencer_rows)),
        "jushuitan_match_status": dict(Counter(row.get(F_JST_MATCH_STATUS) or "" for row in influencer_rows)),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if summary["status"] == "success" else 4


def is_influencer_excel_row(row: dict[str, Any]) -> bool:
    mode = text_value(row, F_INFLUENCER_MODE).strip()
    nick = text_value(row, F_INFLUENCER_NICK).strip()
    account_type = text_value(row, F_INFLUENCER_ACCOUNT_TYPE).strip()
    channel = text_value(row, F_INFLUENCER_FEE_CHANNEL).strip()
    return any(value not in ("", "-") for value in (mode, nick, account_type, channel))


if __name__ == "__main__":
    raise SystemExit(main())
