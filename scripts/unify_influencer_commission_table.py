from __future__ import annotations

import argparse
import json
import os
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from scripts.write_douyin_influencer_excel_to_feishu import parse_doudian_xlsx
from shopops.config import load_settings
from shopops.storage.feishu_bootstrap import FEISHU_BASE_URL, FeishuOpenApiClient


TEXT_FIELD = 1
NUMBER_FIELD = 2

F_UNIQUE_KEY = "unique_key"
F_PLATFORM = "平台"
F_SOURCE = "数据来源"
F_ORDER_NO = "订单号"
F_CREATED_AT = "下单时间"
F_PAY_AT = "支付时间"
F_ORDER_STATUS = "订单状态"
F_PRODUCT_ID = "商品ID"
F_PRODUCT_NAME = "商品名称"
F_QUANTITY = "商品数量"
F_PAID_AMOUNT = "支付金额"
F_INFLUENCER = "带货达人"
F_INFLUENCER_ID = "带货达人ID"
F_COMMISSION_RATE = "带货佣金率"
F_COMMISSION = "带货佣金"
F_COMMISSION_BASIS = "带货佣金口径"
F_SHOP_ID = "店铺ID"
F_SHOP_NAME = "店铺名称"
F_SOURCE_FILE = "来源文件"
F_RAW = "原始数据"
F_FETCHED_AT = "采集时间"

CANONICAL_FIELDS = [
    (F_UNIQUE_KEY, TEXT_FIELD),
    (F_PLATFORM, TEXT_FIELD),
    (F_SOURCE, TEXT_FIELD),
    (F_ORDER_NO, TEXT_FIELD),
    (F_CREATED_AT, TEXT_FIELD),
    (F_PAY_AT, TEXT_FIELD),
    (F_ORDER_STATUS, TEXT_FIELD),
    (F_PRODUCT_ID, TEXT_FIELD),
    (F_PRODUCT_NAME, TEXT_FIELD),
    (F_QUANTITY, NUMBER_FIELD),
    (F_PAID_AMOUNT, NUMBER_FIELD),
    (F_INFLUENCER, TEXT_FIELD),
    (F_INFLUENCER_ID, TEXT_FIELD),
    (F_COMMISSION_RATE, TEXT_FIELD),
    (F_COMMISSION, NUMBER_FIELD),
    (F_COMMISSION_BASIS, TEXT_FIELD),
    (F_SHOP_ID, TEXT_FIELD),
    (F_SHOP_NAME, TEXT_FIELD),
    (F_FETCHED_AT, TEXT_FIELD),
    (F_SOURCE_FILE, TEXT_FIELD),
    (F_RAW, TEXT_FIELD),
]

DUPLICATE_FIELDS_TO_DELETE = {
    "达人ID",
    "达人昵称",
    "佣金率",
    "预估佣金",
    "结算佣金",
    "订单下单时间",
    "带货账号昵称",
    "带货费用",
    "作者账号",
    "抖音/火山号",
    "预估佣金支出",
    "实际佣金支出",
    "订单实际支付金额",
}

SENSITIVE_WECHAT_FIELDS = {"收件人姓名", "收件人地址", "收件人手机"}


class FeishuTableClient:
    def __init__(self, table_id: str) -> None:
        settings = load_settings()
        self.app_token = settings.shopops_data_center_app_token or settings.feishu_app_token
        self.table_id = table_id
        self.session = requests.Session()
        self.session.trust_env = False
        os.environ["NO_PROXY"] = "open.feishu.cn"
        os.environ["no_proxy"] = "open.feishu.cn"
        self.auth = FeishuOpenApiClient(settings.feishu_app_id, settings.feishu_app_secret)

    def request(
        self,
        method: str,
        path: str,
        payload: dict[str, Any] | None = None,
        params: dict[str, Any] | None = None,
        allow_duplicate: bool = False,
    ) -> dict[str, Any]:
        response = self.session.request(
            method,
            f"{FEISHU_BASE_URL}{path}",
            headers=self.auth.headers(),
            json=payload,
            params=params,
            timeout=30,
        )
        body = response.json()
        if allow_duplicate and body.get("code") == 1254014:
            return {}
        if response.status_code >= 400 or body.get("code") != 0:
            raise RuntimeError(f"Feishu API {method} {path} failed HTTP {response.status_code}: {body}")
        return body.get("data") or {}

    def fields(self) -> list[dict[str, Any]]:
        items: list[dict[str, Any]] = []
        page_token = None
        while True:
            params: dict[str, Any] = {"page_size": 100}
            if page_token:
                params["page_token"] = page_token
            data = self.request("GET", f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/fields", params=params)
            items.extend(data.get("items") or [])
            if not data.get("has_more"):
                return items
            page_token = data.get("page_token")

    def ensure_fields(self) -> list[str]:
        existing = {field.get("field_name"): field for field in self.fields()}
        created: list[str] = []
        for field_name, field_type in CANONICAL_FIELDS:
            if field_name in existing:
                continue
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/fields",
                {"field_name": field_name, "type": field_type},
                allow_duplicate=True,
            )
            created.append(field_name)
        return created

    def existing_record_ids(self) -> dict[str, str]:
        records: dict[str, str] = {}
        page_token = None
        while True:
            params: dict[str, Any] = {"page_size": 500}
            if page_token:
                params["page_token"] = page_token
            data = self.request("GET", f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records", params=params)
            for item in data.get("items") or []:
                fields = item.get("fields") or {}
                unique_key = fields.get(F_UNIQUE_KEY)
                if unique_key:
                    records[str(unique_key)] = str(item.get("record_id"))
            if not data.get("has_more"):
                return records
            page_token = data.get("page_token")

    def upsert_rows(self, rows: list[dict[str, Any]]) -> dict[str, int]:
        existing = self.existing_record_ids()
        create_rows: list[dict[str, Any]] = []
        update_rows: list[dict[str, Any]] = []
        for row in rows:
            clean = {key: value for key, value in row.items() if value not in (None, "")}
            record_id = existing.get(str(row[F_UNIQUE_KEY]))
            if record_id:
                update_rows.append({"record_id": record_id, "fields": clean})
            else:
                create_rows.append({"fields": clean})
        for chunk in chunks(create_rows, 500):
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records/batch_create", {"records": chunk})
        for chunk in chunks(update_rows, 500):
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records/batch_update", {"records": chunk})
        return {"created": len(create_rows), "updated": len(update_rows), "saved": len(create_rows) + len(update_rows)}

    def readback(self, unique_keys: set[str]) -> dict[str, dict[str, Any]]:
        found: dict[str, dict[str, Any]] = {}
        page_token = None
        while True:
            params: dict[str, Any] = {"page_size": 500}
            if page_token:
                params["page_token"] = page_token
            data = self.request("GET", f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records", params=params)
            for item in data.get("items") or []:
                fields = item.get("fields") or {}
                unique_key = fields.get(F_UNIQUE_KEY)
                if unique_key in unique_keys:
                    found[str(unique_key)] = fields
            if not data.get("has_more"):
                return found
            page_token = data.get("page_token")

    def delete_duplicate_fields(self) -> list[str]:
        deleted: list[str] = []
        for field in self.fields():
            field_name = str(field.get("field_name") or "")
            field_id = str(field.get("field_id") or "")
            if field_name not in DUPLICATE_FIELDS_TO_DELETE or not field_id:
                continue
            self.request("DELETE", f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/fields/{field_id}")
            deleted.append(field_name)
        return deleted


def wechat_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    fetched_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    order_counts: Counter[str] = Counter()
    source_rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        source = dict(zip(headers, values))
        if any(value not in (None, "") for value in source.values()):
            source_rows.append(source)
            order_counts[str(source.get("订单号") or "").strip()] += 1
    order_sequence: Counter[str] = Counter()
    for source in source_rows:
        order_no = text(source.get("订单号"))
        influencer = text(source.get("带货账号昵称"))
        if not order_no or not influencer:
            continue
        order_sequence[order_no] += 1
        suffix = "" if order_counts[order_no] == 1 else f"_{order_sequence[order_no]}"
        rows.append(
            {
                F_UNIQUE_KEY: f"shipin_{order_no}{suffix}",
                F_PLATFORM: "视频号",
                F_SOURCE: "微信小店导出Excel",
                F_ORDER_NO: order_no,
                F_CREATED_AT: text(source.get("订单下单时间")),
                F_PAY_AT: text(source.get("支付时间")),
                F_ORDER_STATUS: text(source.get("订单状态")),
                F_PRODUCT_ID: text(source.get("商品编码(平台)")),
                F_PRODUCT_NAME: text(source.get("商品名称")),
                F_QUANTITY: number(source.get("商品数量")),
                F_PAID_AMOUNT: number(source.get("订单实际支付金额")),
                F_INFLUENCER: influencer,
                F_INFLUENCER_ID: "",
                F_COMMISSION_RATE: text(source.get("带货佣金率")),
                F_COMMISSION: number(source.get("带货费用")),
                F_COMMISSION_BASIS: text(source.get("带货费用类型")) or "带货费用",
                F_SHOP_ID: "",
                F_SHOP_NAME: "视频号",
                F_FETCHED_AT: fetched_at,
                F_SOURCE_FILE: str(path),
                F_RAW: json.dumps(redact_wechat(source), ensure_ascii=False, sort_keys=True, default=str),
            }
        )
    return rows


def douyin_rows(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    fetched_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for source in parse_doudian_xlsx(path):
        order_no = text(source.get("订单id"))
        influencer = text(source.get("作者账号"))
        if not order_no or not influencer:
            continue
        actual_commission = number(source.get("实际佣金支出"))
        estimated_commission = number(source.get("预估佣金支出"))
        commission = actual_commission if actual_commission and actual_commission > 0 else estimated_commission
        basis = "实际佣金支出" if actual_commission and actual_commission > 0 else "预估佣金支出"
        rows.append(
            {
                F_UNIQUE_KEY: f"douyin_{order_no}",
                F_PLATFORM: "抖音",
                F_SOURCE: "抖店导出Excel",
                F_ORDER_NO: order_no,
                F_CREATED_AT: "",
                F_PAY_AT: text(source.get("付款时间")),
                F_ORDER_STATUS: text(source.get("订单状态")),
                F_PRODUCT_ID: text(source.get("商品id")),
                F_PRODUCT_NAME: text(source.get("商品名称")),
                F_QUANTITY: number(source.get("商品数量")),
                F_PAID_AMOUNT: number(source.get("支付金额")),
                F_INFLUENCER: influencer,
                F_INFLUENCER_ID: text(source.get("抖音/火山号")),
                F_COMMISSION_RATE: text(source.get("佣金率")),
                F_COMMISSION: commission,
                F_COMMISSION_BASIS: basis,
                F_SHOP_ID: text(source.get("店铺id")),
                F_SHOP_NAME: text(source.get("店铺名称")),
                F_FETCHED_AT: fetched_at,
                F_SOURCE_FILE: str(path),
                F_RAW: json.dumps(source, ensure_ascii=False, sort_keys=True),
            }
        )
    return rows


def validate_key_fields(rows: list[dict[str, Any]], readback: dict[str, dict[str, Any]]) -> dict[str, Any]:
    required = [F_ORDER_NO, F_INFLUENCER, F_COMMISSION_RATE, F_COMMISSION]
    source_missing = {
        field: [row[F_UNIQUE_KEY] for row in rows if row.get(field) in (None, "")]
        for field in required
    }
    readback_missing = {
        field: [key for key, fields in readback.items() if fields.get(field) in (None, "")]
        for field in required
    }
    mismatched_order = [
        row[F_UNIQUE_KEY]
        for row in rows
        if row[F_UNIQUE_KEY] in readback and str(readback[row[F_UNIQUE_KEY]].get(F_ORDER_NO) or "") != str(row[F_ORDER_NO])
    ]
    return {
        "source_missing": {field: values[:20] for field, values in source_missing.items() if values},
        "readback_missing": {field: values[:20] for field, values in readback_missing.items() if values},
        "missing_unique_keys": sorted(set(row[F_UNIQUE_KEY] for row in rows) - set(readback))[:50],
        "mismatched_order_unique_keys": mismatched_order[:50],
    }


def text(value: Any) -> str:
    if value in (None, "-", ""):
        return ""
    return str(value).strip().rstrip("\t")


def number(value: Any) -> float | None:
    value_text = text(value).replace(",", "")
    if not value_text:
        return None
    if value_text.endswith("%"):
        value_text = value_text[:-1]
    return round(float(value_text), 4)


def redact_wechat(row: dict[str, Any]) -> dict[str, Any]:
    return {key: ("[REDACTED]" if key in SENSITIVE_WECHAT_FIELDS else value) for key, value in row.items()}


def chunks(values: list[dict[str, Any]], size: int) -> list[list[dict[str, Any]]]:
    return [values[index : index + size] for index in range(0, len(values), size)]


def find_default_wechat_file() -> Path:
    matches = [
        path
        for path in (Path.home() / "Downloads").rglob("*.xlsx")
        if "wx8248933f80e464e7" in str(path) and "2026年06月06日20时51分55秒" in str(path)
    ]
    if not matches:
        raise FileNotFoundError("Could not find the requested WeChat Channels workbook in Downloads")
    return sorted(matches, key=lambda item: item.stat().st_mtime, reverse=True)[0]


def find_default_douyin_file() -> Path:
    matches = [
        path
        for path in (Path.home() / "Downloads").glob("*.xlsx")
        if path.name == "抖音.xlsx" or path.name.encode("unicode_escape").decode() == r"\u6296\u97f3.xlsx"
    ]
    if not matches:
        raise FileNotFoundError("Could not find 抖音.xlsx in Downloads")
    return matches[0]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--wechat-excel", default="")
    parser.add_argument("--douyin-excel", default="")
    parser.add_argument("--target-table", default="")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--keep-duplicate-fields", action="store_true")
    parser.add_argument("--evidence", default="docs/live-evidence/unified-influencer-commission-result.json")
    args = parser.parse_args()

    settings = load_settings()
    table_id = args.target_table or settings.table_douyin_influencer_commission
    wechat_path = Path(args.wechat_excel) if args.wechat_excel else find_default_wechat_file()
    douyin_path = Path(args.douyin_excel) if args.douyin_excel else find_default_douyin_file()
    rows = wechat_rows(wechat_path) + douyin_rows(douyin_path)
    unique_key_counts = Counter(row[F_UNIQUE_KEY] for row in rows)
    duplicate_unique_keys = [key for key, count in unique_key_counts.items() if count > 1]
    source_validation = validate_key_fields(rows, {})
    summary: dict[str, Any] = {
        "status": "dry_run" if args.dry_run else "started",
        "target_table": table_id,
        "wechat_excel": str(wechat_path),
        "douyin_excel": str(douyin_path),
        "source_row_count": len(rows),
        "platform_counts": dict(Counter(row[F_PLATFORM] for row in rows)),
        "duplicate_unique_keys": duplicate_unique_keys[:50],
        "source_validation": source_validation["source_missing"],
    }
    if args.dry_run:
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0 if not duplicate_unique_keys and not source_validation["source_missing"] else 4

    client = FeishuTableClient(table_id)
    created_fields = client.ensure_fields()
    upsert = client.upsert_rows(rows)
    readback = client.readback({row[F_UNIQUE_KEY] for row in rows})
    validation = validate_key_fields(rows, readback)
    deleted_fields: list[str] = []
    if not args.keep_duplicate_fields and not any(validation.values()):
        deleted_fields = client.delete_duplicate_fields()
    remaining_duplicate_fields = [
        field.get("field_name")
        for field in client.fields()
        if field.get("field_name") in DUPLICATE_FIELDS_TO_DELETE
    ]
    summary.update(
        {
            "status": "success" if not any(validation.values()) and not remaining_duplicate_fields else "needs_review",
            "created_fields": created_fields,
            "upsert": upsert,
            "readback_count": len(readback),
            "validation": validation,
            "deleted_duplicate_fields": deleted_fields,
            "remaining_duplicate_fields": remaining_duplicate_fields,
        }
    )
    evidence_path = Path(args.evidence)
    evidence_path.parent.mkdir(parents=True, exist_ok=True)
    evidence_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if summary["status"] == "success" else 4


if __name__ == "__main__":
    raise SystemExit(main())
