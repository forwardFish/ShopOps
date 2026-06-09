from __future__ import annotations

import argparse
import json
import os
import sys
from collections import Counter
from dataclasses import replace
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from scripts.run_jushuitan_orders_to_feishu import FeishuOrderWriter, feishu_order_fields
from shopops.collectors.jushuitan_order_api import JushuitanOrderApiCollector
from shopops.config import Settings, load_settings
from shopops.storage.feishu_bootstrap import FEISHU_BASE_URL, FeishuOpenApiClient


TEXT_FIELD = 1
NUMBER_FIELD = 2

TARGET_TABLE_ENV = "FEISHU_TABLE_DOUYIN_INFLUENCER_COMMISSION"
DOUYIN_SHOP_ENV = "JUSHUITAN_SHOP_ID_DOUYIN"

F_UNIQUE_KEY = "unique_key"
F_PLATFORM = "\u5e73\u53f0"
F_SOURCE = "\u6570\u636e\u6765\u6e90"
F_ORDER_NO = "\u8ba2\u5355\u53f7"
F_ORDER_STATUS = "\u8ba2\u5355\u72b6\u6001"
F_PAY_AT = "\u652f\u4ed8\u65f6\u95f4"
F_RECEIVED_AT = "\u6536\u8d27\u65f6\u95f4"
F_SETTLED_AT = "\u8ba2\u5355\u7ed3\u7b97\u65f6\u95f4"
F_PRODUCT_ID = "\u5546\u54c1ID"
F_PRODUCT_NAME = "\u5546\u54c1\u540d\u79f0"
F_QUANTITY = "\u5546\u54c1\u6570\u91cf"
F_PAID_AMOUNT = "\u652f\u4ed8\u91d1\u989d"
F_COMMISSION_RATE = "\u4f63\u91d1\u7387"
F_ESTIMATED_COMMISSION = "\u9884\u4f30\u4f63\u91d1\u652f\u51fa"
F_SETTLEMENT_AMOUNT = "\u7ed3\u7b97\u91d1\u989d"
F_ACTUAL_COMMISSION = "\u5b9e\u9645\u4f63\u91d1\u652f\u51fa"
F_AUTHOR_ACCOUNT = "\u4f5c\u8005\u8d26\u53f7"
F_DOUYIN_ID = "\u6296\u97f3/\u706b\u5c71\u53f7"
F_SHOP_ID = "\u5e97\u94faID"
F_SHOP_NAME = "\u5e97\u94fa\u540d\u79f0"
F_COMMISSION_TYPE = "\u4f63\u91d1\u7c7b\u578b"
F_ORDER_SOURCE = "\u8ba2\u5355\u6765\u6e90"
F_APP_CHANNEL = "APP\u6e20\u9053"
F_TRAFFIC_DETAIL_SOURCE = "\u6d41\u91cf\u7ec6\u5206\u6765\u6e90"
F_TRAFFIC_SOURCE = "\u6d41\u91cf\u6765\u6e90"
F_ORDER_TYPE = "\u8ba2\u5355\u7c7b\u578b"
F_JST_MATCH_STATUS = "\u805a\u6c34\u6f6d\u5339\u914d\u72b6\u6001"
F_JST_ORDER_NO = "\u805a\u6c34\u6f6d\u8ba2\u5355\u53f7"
F_JST_ORDER_STATUS = "\u805a\u6c34\u6f6d\u8ba2\u5355\u72b6\u6001"
F_JST_SHOP_ID = "\u805a\u6c34\u6f6d\u5e97\u94faID"
F_FETCHED_AT = "\u91c7\u96c6\u65f6\u95f4"
F_SOURCE_FILE = "\u6765\u6e90\u6587\u4ef6"
F_RAW = "\u539f\u59cb\u6570\u636e"

TARGET_FIELDS = [
    (F_UNIQUE_KEY, TEXT_FIELD),
    (F_PLATFORM, TEXT_FIELD),
    (F_SOURCE, TEXT_FIELD),
    (F_ORDER_NO, TEXT_FIELD),
    (F_ORDER_STATUS, TEXT_FIELD),
    (F_PAY_AT, TEXT_FIELD),
    (F_RECEIVED_AT, TEXT_FIELD),
    (F_SETTLED_AT, TEXT_FIELD),
    (F_PRODUCT_ID, TEXT_FIELD),
    (F_PRODUCT_NAME, TEXT_FIELD),
    (F_QUANTITY, NUMBER_FIELD),
    (F_PAID_AMOUNT, NUMBER_FIELD),
    (F_COMMISSION_RATE, NUMBER_FIELD),
    (F_ESTIMATED_COMMISSION, NUMBER_FIELD),
    (F_SETTLEMENT_AMOUNT, NUMBER_FIELD),
    (F_ACTUAL_COMMISSION, NUMBER_FIELD),
    (F_AUTHOR_ACCOUNT, TEXT_FIELD),
    (F_DOUYIN_ID, TEXT_FIELD),
    (F_SHOP_ID, TEXT_FIELD),
    (F_SHOP_NAME, TEXT_FIELD),
    (F_COMMISSION_TYPE, TEXT_FIELD),
    (F_ORDER_SOURCE, TEXT_FIELD),
    (F_APP_CHANNEL, TEXT_FIELD),
    (F_TRAFFIC_DETAIL_SOURCE, TEXT_FIELD),
    (F_TRAFFIC_SOURCE, TEXT_FIELD),
    (F_ORDER_TYPE, TEXT_FIELD),
    (F_JST_MATCH_STATUS, TEXT_FIELD),
    (F_JST_ORDER_NO, TEXT_FIELD),
    (F_JST_ORDER_STATUS, TEXT_FIELD),
    (F_JST_SHOP_ID, TEXT_FIELD),
    (F_FETCHED_AT, TEXT_FIELD),
    (F_SOURCE_FILE, TEXT_FIELD),
    (F_RAW, TEXT_FIELD),
]


class FeishuTableClient:
    def __init__(self, settings: Settings, table_id: str) -> None:
        self.settings = settings
        self.app_token = settings.shopops_data_center_app_token or settings.feishu_app_token
        self.table_id = table_id
        self.client = FeishuOpenApiClient(settings.feishu_app_id, settings.feishu_app_secret)
        self.session = requests.Session()
        self.session.trust_env = False
        os.environ["NO_PROXY"] = "open.feishu.cn"
        os.environ["no_proxy"] = "open.feishu.cn"

    def request(
        self,
        method: str,
        path: str,
        payload: dict[str, Any] | None = None,
        params: dict[str, Any] | None = None,
        allow_duplicate: bool = False,
    ) -> dict[str, Any]:
        response = self.session.request(
            method,
            f"{FEISHU_BASE_URL}{path}",
            headers=self.client.headers(),
            json=payload,
            params=params,
            timeout=30,
        )
        body = response.json()
        if allow_duplicate and body.get("code") == 1254014:
            return {}
        if response.status_code >= 400 or body.get("code") != 0:
            raise RuntimeError(f"Feishu API {method} {path} failed HTTP {response.status_code}: {body}")
        return body.get("data") or {}

    def ensure_fields(self) -> list[str]:
        existing = self.field_names()
        created: list[str] = []
        for field_name, field_type in TARGET_FIELDS:
            if field_name in existing:
                continue
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/fields",
                {"field_name": field_name, "type": field_type},
                allow_duplicate=True,
            )
            existing.add(field_name)
            created.append(field_name)
        return created

    def field_names(self) -> set[str]:
        names: set[str] = set()
        page_token = None
        while True:
            params: dict[str, Any] = {"page_size": 100}
            if page_token:
                params["page_token"] = page_token
            data = self.request(
                "GET",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/fields",
                params=params,
            )
            for item in data.get("items", []) or []:
                if item.get("field_name"):
                    names.add(str(item["field_name"]))
            if not data.get("has_more"):
                return names
            page_token = data.get("page_token")

    def iter_records(self, field_names: list[str] | None = None) -> Any:
        page_token = None
        while True:
            params: dict[str, Any] = {"page_size": 500}
            if page_token:
                params["page_token"] = page_token
            if field_names:
                params["field_names"] = json.dumps(field_names, ensure_ascii=False)
            data = self.request(
                "GET",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records",
                params=params,
            )
            yield from data.get("items", []) or []
            if not data.get("has_more"):
                return
            page_token = data.get("page_token")

    def delete_platform_records(self, platform: str) -> dict[str, Any]:
        record_ids: list[str] = []
        for record in self.iter_records([F_PLATFORM]):
            fields = record.get("fields") or {}
            if scalar_text(fields.get(F_PLATFORM)) == platform:
                record_id = str(record.get("record_id") or "")
                if record_id:
                    record_ids.append(record_id)
        for chunk in chunks([{"record_id": record_id} for record_id in record_ids], 500):
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records/batch_delete",
                {"records": [item["record_id"] for item in chunk]},
            )
        return {"platform": platform, "deleted_records": len(record_ids)}

    def duplicate_platform_order_keys(self, platform: str) -> dict[str, Any]:
        counts: Counter[str] = Counter()
        for record in self.iter_records([F_PLATFORM, F_ORDER_NO]):
            fields = record.get("fields") or {}
            if scalar_text(fields.get(F_PLATFORM)) != platform:
                continue
            order_no = scalar_text(fields.get(F_ORDER_NO))
            if order_no:
                counts[f"{platform}|{order_no}"] += 1
        duplicates = {key: count for key, count in counts.items() if count > 1}
        return {
            "platform": platform,
            "unique_order_count": len(counts),
            "duplicate_key_count": len(duplicates),
            "duplicate_extra_rows": sum(count - 1 for count in duplicates.values()),
            "sample_duplicate_keys": list(duplicates)[:20],
        }

    def existing_records_by_unique_key(self) -> dict[str, str]:
        records: dict[str, str] = {}
        page_token = None
        page = 0
        while True:
            page += 1
            params: dict[str, Any] = {"page_size": 500}
            if page_token:
                params["page_token"] = page_token
            data = self.request(
                "GET",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records",
                params=params,
            )
            for item in data.get("items", []) or []:
                fields = item.get("fields") or {}
                unique_key = fields.get(F_UNIQUE_KEY)
                if unique_key:
                    records[str(unique_key)] = str(item.get("record_id"))
            log(f"loaded existing target records page={page} total_keys={len(records)}")
            if not data.get("has_more"):
                return records
            page_token = data.get("page_token")

    def upsert_rows(self, rows: list[dict[str, Any]]) -> dict[str, int]:
        existing = self.existing_records_by_unique_key()
        to_create: list[dict[str, Any]] = []
        to_update: list[dict[str, Any]] = []
        for row in rows:
            clean = {key: value for key, value in row.items() if value not in (None, "")}
            record_id = existing.get(str(row[F_UNIQUE_KEY]))
            if record_id:
                to_update.append({"record_id": record_id, "fields": clean})
            else:
                to_create.append({"fields": clean})

        for chunk in chunks(to_create, 500):
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records/batch_create",
                {"records": chunk},
            )
        for chunk in chunks(to_update, 500):
            self.request(
                "POST",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records/batch_update",
                {"records": chunk},
            )
        return {"created": len(to_create), "updated": len(to_update), "saved": len(to_create) + len(to_update)}

    def readback_unique_keys(self, unique_keys: set[str]) -> dict[str, dict[str, Any]]:
        found: dict[str, dict[str, Any]] = {}
        page_token = None
        page = 0
        while True:
            page += 1
            params: dict[str, Any] = {"page_size": 500}
            if page_token:
                params["page_token"] = page_token
            data = self.request(
                "GET",
                f"/bitable/v1/apps/{self.app_token}/tables/{self.table_id}/records",
                params=params,
            )
            for item in data.get("items", []) or []:
                fields = item.get("fields") or {}
                unique_key = fields.get(F_UNIQUE_KEY)
                if unique_key in unique_keys:
                    found[str(unique_key)] = fields
            log(f"readback target records page={page} found={len(found)}/{len(unique_keys)}")
            if not data.get("has_more"):
                return found
            page_token = data.get("page_token")


def settings_for_douyin(settings: Settings) -> Settings:
    shop_id = os.getenv(DOUYIN_SHOP_ENV, "").strip()
    return replace(
        settings,
        order_source="jushuitan",
        use_mock_collectors=False,
        shop_platform="doudian",
        shop_id=shop_id,
        shop_name="\u6296\u97f3",
        jushuitan_shop_ids=shop_id,
    )


def fetch_douyin_orders(settings: Settings) -> tuple[Any, int]:
    platform_settings = settings_for_douyin(settings)
    log("fetching Jushuitan douyin orders")
    result = JushuitanOrderApiCollector(platform_settings).fetch_today()
    saved_count = 0
    if result.success and result.orders:
        log(f"writing {len(result.orders)} Jushuitan douyin orders to Feishu order table")
        writer = FeishuOrderWriter(settings)
        writer.ensure_fields()
        rows = [feishu_order_fields("\u6296\u97f3", order) for order in result.orders]
        saved_count = writer.write_orders(rows)
    log(f"Jushuitan douyin done success={result.success} orders={result.order_count} saved={saved_count}")
    return result, saved_count


def parse_doudian_xlsx(path: Path) -> list[dict[str, str]]:
    rows = openpyxl_sheet_rows(path)
    if not rows:
        return []
    headers = rows[0]
    data_rows = [row for row in rows[1:] if any(value.strip() for value in row)]
    return [dict(zip(headers, row)) for row in data_rows]


def openpyxl_sheet_rows(path: Path) -> list[list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    reset = getattr(sheet, "reset_dimensions", None)
    if callable(reset):
        reset()
    rows: list[list[str]] = []
    for values in sheet.iter_rows(values_only=True):
        rows.append([clean_cell(value) for value in values])
    return rows


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().rstrip("\t")


def doudian_influencer_rows(excel_rows: list[dict[str, str]], source_file: Path, jst_orders: list[dict[str, Any]]) -> list[dict[str, Any]]:
    fetched_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    jst_by_order_no = {str(order.get("order_id") or ""): order for order in jst_orders if order.get("order_id")}
    rows: list[dict[str, Any]] = []
    for row in excel_rows:
        order_no = text_value(row, "\u8ba2\u5355id")
        if not order_no:
            continue
        jst_order = jst_by_order_no.get(order_no)
        rows.append(
            {
                F_UNIQUE_KEY: f"douyin_{order_no}",
                F_PLATFORM: "\u6296\u97f3",
                F_SOURCE: "\u6296\u5e97\u5bfc\u51faExcel",
                F_ORDER_NO: order_no,
                F_ORDER_STATUS: text_value(row, "\u8ba2\u5355\u72b6\u6001"),
                F_PAY_AT: text_value(row, "\u4ed8\u6b3e\u65f6\u95f4"),
                F_RECEIVED_AT: text_value(row, "\u6536\u8d27\u65f6\u95f4"),
                F_SETTLED_AT: text_value(row, "\u8ba2\u5355\u7ed3\u7b97\u65f6\u95f4"),
                F_PRODUCT_ID: text_value(row, "\u5546\u54c1id"),
                F_PRODUCT_NAME: text_value(row, "\u5546\u54c1\u540d\u79f0"),
                F_QUANTITY: number_value(row, "\u5546\u54c1\u6570\u91cf"),
                F_PAID_AMOUNT: number_value(row, "\u652f\u4ed8\u91d1\u989d"),
                F_COMMISSION_RATE: percent_number_value(row, "\u4f63\u91d1\u7387"),
                F_ESTIMATED_COMMISSION: number_value(row, "\u9884\u4f30\u4f63\u91d1\u652f\u51fa"),
                F_SETTLEMENT_AMOUNT: number_value(row, "\u7ed3\u7b97\u91d1\u989d"),
                F_ACTUAL_COMMISSION: number_value(row, "\u5b9e\u9645\u4f63\u91d1\u652f\u51fa"),
                F_AUTHOR_ACCOUNT: text_value(row, "\u4f5c\u8005\u8d26\u53f7"),
                F_DOUYIN_ID: text_value(row, "\u6296\u97f3/\u706b\u5c71\u53f7"),
                F_SHOP_ID: text_value(row, "\u5e97\u94faid"),
                F_SHOP_NAME: text_value(row, "\u5e97\u94fa\u540d\u79f0"),
                F_COMMISSION_TYPE: text_value(row, "\u4f63\u91d1\u7c7b\u578b"),
                F_ORDER_SOURCE: text_value(row, "\u8ba2\u5355\u6765\u6e90"),
                F_APP_CHANNEL: text_value(row, "\u8ba2\u5355\u6765\u6e90"),
                F_TRAFFIC_DETAIL_SOURCE: text_value(row, "\u6d41\u91cf\u7ec6\u5206\u6765\u6e90"),
                F_TRAFFIC_SOURCE: text_value(row, "\u6d41\u91cf\u6765\u6e90"),
                F_ORDER_TYPE: text_value(row, "\u8ba2\u5355\u7c7b\u578b"),
                F_JST_MATCH_STATUS: "\u5df2\u5339\u914d" if jst_order else "\u672a\u5339\u914d",
                F_JST_ORDER_NO: str(jst_order.get("order_id") or "") if jst_order else "",
                F_JST_ORDER_STATUS: str(jst_order.get("order_status") or "") if jst_order else "",
                F_JST_SHOP_ID: str(jst_order.get("shop_id") or "") if jst_order else "",
                F_FETCHED_AT: fetched_at,
                F_SOURCE_FILE: str(source_file),
                F_RAW: json.dumps(row, ensure_ascii=False, sort_keys=True),
            }
        )
    return rows


def is_influencer_row(row: dict[str, str]) -> bool:
    return text_value(row, "\u4f5c\u8005\u8d26\u53f7") not in ("", "-")


def text_value(row: dict[str, str], key: str) -> str:
    value = row.get(key, "")
    if value in (None, "-"):
        return ""
    return str(value).strip().rstrip("\t")


def scalar_text(value: Any) -> str:
    if isinstance(value, list):
        return "".join(str(item.get("text") if isinstance(item, dict) else item) for item in value).strip()
    return clean_cell(value)


def number_value(row: dict[str, str], key: str) -> float | None:
    value = text_value(row, key)
    if value == "":
        return None
    return round(float(value), 2)


def percent_number_value(row: dict[str, str], key: str) -> float | None:
    value = text_value(row, key)
    if value == "":
        return None
    return round(float(value.rstrip("%")), 4)


def chunks(values: list[dict[str, Any]], size: int) -> list[list[dict[str, Any]]]:
    return [values[index : index + size] for index in range(0, len(values), size)]


def log(message: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {message}", flush=True)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--target-table", default="")
    parser.add_argument("--skip-jushuitan", action="store_true")
    parser.add_argument("--no-delete-existing-douyin", action="store_true")
    parser.add_argument("--include-no-author", action="store_true", help="Deprecated; all Excel rows with an order id are imported.")
    parser.add_argument("--evidence", default="")
    args = parser.parse_args()

    settings = load_settings()
    target_table = args.target_table or os.getenv(TARGET_TABLE_ENV, "").strip()
    if not target_table:
        raise RuntimeError(f"Missing {TARGET_TABLE_ENV} or --target-table")
    if not args.skip_jushuitan and not os.getenv(DOUYIN_SHOP_ENV, "").strip():
        raise RuntimeError(f"Missing {DOUYIN_SHOP_ENV}")

    jst_result = None
    jst_saved_count = 0
    if args.skip_jushuitan:
        log("skipping Jushuitan douyin fetch")
        jst_orders: list[dict[str, Any]] = []
    else:
        jst_result, jst_saved_count = fetch_douyin_orders(settings)
        if not jst_result.success:
            raise RuntimeError(f"Jushuitan douyin fetch failed: {jst_result.error_code} {jst_result.error_message}")
        jst_orders = jst_result.orders or []

    excel_path = Path(args.excel)
    log(f"loading Doudian Excel {excel_path}")
    excel_rows = parse_doudian_xlsx(excel_path)
    no_author_rows = [row for row in excel_rows if not is_influencer_row(row)]
    import_rows = excel_rows
    log(f"Excel rows={len(excel_rows)} influencer_rows={len(import_rows)} no_author_rows={len(no_author_rows)}")
    feishu_rows = doudian_influencer_rows(import_rows, excel_path, jst_orders)

    client = FeishuTableClient(settings, target_table)
    log(f"ensuring target table fields table={target_table}")
    created_fields = client.ensure_fields()
    delete_summary = {"platform": "\u6296\u97f3", "deleted_records": 0}
    if not args.no_delete_existing_douyin:
        log("deleting existing douyin rows from target table before reimport")
        delete_summary = client.delete_platform_records("\u6296\u97f3")
    log(f"upserting {len(feishu_rows)} douyin influencer rows")
    upsert_summary = client.upsert_rows(feishu_rows)
    log("reading back written douyin unique keys")
    readback = client.readback_unique_keys({row[F_UNIQUE_KEY] for row in feishu_rows})
    missing_keys = sorted(set(row[F_UNIQUE_KEY] for row in feishu_rows) - set(readback))
    mismatched = [
        row[F_UNIQUE_KEY]
        for row in feishu_rows
        if row[F_UNIQUE_KEY] in readback and str(readback[row[F_UNIQUE_KEY]].get(F_ORDER_NO) or "") != str(row[F_ORDER_NO])
    ]
    skipped_order_nos = [text_value(row, "\u8ba2\u5355id") for row in no_author_rows]
    summary = {
        "status": "success" if not missing_keys and not mismatched else "readback_mismatch",
        "target_table": target_table,
        "excel_file": str(excel_path),
        "excel_row_count": len(excel_rows),
        "skipped_no_author_count": 0,
        "skipped_no_author_order_nos_sample": skipped_order_nos[:30],
        "influencer_row_count": len(feishu_rows),
        "unique_key_prefix": "douyin_",
        "created_fields": created_fields,
        "delete_existing_douyin": delete_summary,
        "upsert": upsert_summary,
        "readback_count": len(readback),
        "missing_unique_keys": missing_keys[:30],
        "mismatched_unique_keys": mismatched[:30],
        "duplicate_check": client.duplicate_platform_order_keys("\u6296\u97f3"),
        "jushuitan": {
            "fetched": not args.skip_jushuitan,
            "success": bool(jst_result.success) if jst_result else None,
            "shop_id": os.getenv(DOUYIN_SHOP_ENV, "").strip(),
            "order_count": jst_result.order_count if jst_result else None,
            "saved_to_order_table_count": jst_saved_count,
            "total_amount": jst_result.total_amount if jst_result else None,
            "error_code": jst_result.error_code if jst_result else None,
            "error_message": jst_result.error_message if jst_result else None,
        },
        "author_top20": dict(Counter(row.get(F_AUTHOR_ACCOUNT) or "" for row in feishu_rows).most_common(20)),
        "jushuitan_match_status": dict(Counter(row.get(F_JST_MATCH_STATUS) or "" for row in feishu_rows)),
    }
    if args.evidence:
        evidence_path = Path(args.evidence)
        evidence_path.parent.mkdir(parents=True, exist_ok=True)
        evidence_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if summary["status"] == "success" else 4


if __name__ == "__main__":
    raise SystemExit(main())
