from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook

from shopops.services.product_breakdown import product_rules_from_records
from scripts.import_daily_files_to_feishu import (
    F_ACCESSORY_FLAG,
    F_AMOUNT_PER_DEAL,
    AD_FIELD_TYPES,
    F_CREATED_AT,
    F_DATE,
    F_DEAL_COUNT,
    F_DEAL_SPEND,
    F_NET_ACTUAL_ROI,
    F_NET_DEAL_COUNT,
    F_NET_TRADE_AMOUNT,
    INTERNAL_PRODUCT_BREAKDOWN_QUANTITY,
    F_ORDER_NO,
    F_PAID_AMOUNT,
    F_PLATFORM,
    F_QUANTITY,
    F_REFUND_AMOUNT,
    F_SETTLED_TRADE_AMOUNT,
    F_TOTAL_SPEND,
    F_TRADE_AMOUNT,
    F_TRADE_STATUS,
    F_UNIQUE_KEY,
    FeishuDailyClient,
    actual_sold_quantity,
    add_product_breakdown_to_orders,
    ad_unique_key,
    classify_file,
    collapse_order_rows,
    discover_daily_files,
    is_accessory_product,
    is_retryable_feishu_response,
    order_unique_key,
    parse_influencer_rows,
    parse_ad_rows,
    parse_order_rows,
    run_import,
)


def test_discovers_date_first_daily_folder_layout(tmp_path: Path):
    batch = tmp_path / "0608"
    tmall = batch / "天猫"
    douyin = batch / "抖音"
    tmall.mkdir(parents=True)
    douyin.mkdir(parents=True)
    order_file = tmall / "ExportOrderList.xlsx"
    temp_order_file = tmall / "~$ExportOrderList.xlsx"
    ad_file = tmall / "天猫投流.xlsx"
    douyin_order = douyin / "orders.csv"
    order_file.write_text("placeholder", encoding="utf-8")
    temp_order_file.write_text("placeholder", encoding="utf-8")
    ad_file.write_text("placeholder", encoding="utf-8")
    douyin_order.write_text("placeholder", encoding="utf-8")

    discovered = discover_daily_files(batch)

    assert discovered["天猫"]["orders"] == [order_file]
    assert discovered["天猫"]["ads"] == [ad_file]
    assert discovered["抖音"]["orders"] == [douyin_order]


def test_daily_import_ignores_taobao_folder(tmp_path: Path):
    batch = tmp_path / "0610"
    taobao = batch / "淘宝"
    taobao.mkdir(parents=True)
    order_file = taobao / "ExportOrderList.xlsx"
    order_file.write_text("placeholder", encoding="utf-8")

    discovered = discover_daily_files(batch)

    assert all(order_file not in paths for kinds in discovered.values() for paths in kinds.values())


def test_parse_tmall_order_uses_order_no_as_unique_key_and_refund_updates_amount(tmp_path: Path):
    path = tmp_path / "ExportOrderList.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["订单编号", "买家实付金额", "退款金额", "订单创建时间", "商品标题", "宝贝总数量", "订单状态"])
    sheet.append(["T1", "0.00", "169.00", "2026-06-08 10:00:00", "洗面奶", "1", "交易关闭"])
    workbook.save(path)

    rows = parse_order_rows("天猫", path)

    assert rows[0][F_UNIQUE_KEY] == "tmall_T1"
    assert rows[0][F_ORDER_NO] == "T1"
    assert rows[0][F_QUANTITY] == 0
    assert rows[0][F_PAID_AMOUNT] == 169.0
    assert rows[0][F_REFUND_AMOUNT] == 169.0


def test_parse_pdd_order_status_refund_without_explicit_refund(tmp_path: Path):
    path = tmp_path / "orders.csv"
    headers = ["商品", "订单号", "订单状态", "售后状态", "商品数量(件)", "订单成交时间", "支付时间", "商家实收金额(元)"]
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        writer.writerow(
            {
                "商品": "洗面奶",
                "订单号": "P1",
                "订单状态": "未发货，退款成功",
                "售后状态": "退款成功",
                "商品数量(件)": "1\t",
                "订单成交时间": "2026-06-07 09:30:00",
                "支付时间": "2026-06-08 10:00:00",
                "商家实收金额(元)": "168.99\t",
            }
        )

    rows = parse_order_rows("拼多多", path)

    assert rows[0][F_UNIQUE_KEY] == "pdd_P1"
    assert rows[0][F_CREATED_AT] == "2026-06-07 09:30:00"
    assert rows[0][F_QUANTITY] == 0
    assert rows[0][F_PAID_AMOUNT] == 168.99
    assert rows[0][F_REFUND_AMOUNT] == 168.99


def test_parse_pdd_order_falls_back_to_order_no_date_when_trade_time_is_blank(tmp_path: Path):
    path = tmp_path / "orders.csv"
    headers = ["商品", "订单号", "订单状态", "售后状态", "商品数量(件)", "订单成交时间", "支付时间", "商家实收金额(元)"]
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        writer.writerow(
            {
                "商品": "洗面奶",
                "订单号": "260607-400000294410900",
                "订单状态": "待付款",
                "售后状态": "无售后或售后取消",
                "商品数量(件)": "1",
                "订单成交时间": "",
                "支付时间": "",
                "商家实收金额(元)": "",
            }
        )

    rows = parse_order_rows("拼多多", path)

    assert rows[0][F_CREATED_AT] == "2026-06-07"
    assert rows[0][F_QUANTITY] == 0


def test_parse_pdd_price_difference_product_has_zero_actual_quantity(tmp_path: Path):
    path = tmp_path / "orders.csv"
    headers = ["商品", "订单号", "订单状态", "售后状态", "商品数量(件)", "订单成交时间", "支付时间", "商家实收金额(元)"]
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        writer.writerow(
            {
                "商品": "【购买前须联系客服确认】补收差价专用商品",
                "订单号": "260608-400000294410900",
                "订单状态": "已收货",
                "售后状态": "无售后或售后取消",
                "商品数量(件)": "500",
                "订单成交时间": "2026-06-08 09:00:00",
                "支付时间": "2026-06-08 09:00:00",
                "商家实收金额(元)": "0",
            }
        )

    rows = parse_order_rows("拼多多", path)

    assert rows[0][F_QUANTITY] == 0


def test_accessory_product_has_flag_and_zero_actual_quantity():
    product = "洁面乳打泡机配件"

    assert is_accessory_product(product)
    assert actual_sold_quantity(
        quantity=2,
        product=product,
        unit_price=39.9,
        refund_amount=0,
        trade_status="已收货",
        fulfill_status="无售后或售后取消",
    ) == 0


def test_regular_product_keeps_actual_quantity():
    product = "趣白全自动洗面奶打泡机"

    assert not is_accessory_product(product)
    assert actual_sold_quantity(
        quantity=2,
        product=product,
        unit_price=39.9,
        refund_amount=0,
        trade_status="已收货",
        fulfill_status="无售后或售后取消",
    ) == 2


def test_collapsed_order_does_not_let_accessory_line_zero_main_product():
    rows = collapse_order_rows(
        [
            {
                F_UNIQUE_KEY: "douyin_O1",
                F_ORDER_NO: "O1",
                F_QUANTITY: 1,
                F_PAID_AMOUNT: 39.9,
                F_REFUND_AMOUNT: 0,
                F_ACCESSORY_FLAG: "否",
            },
            {
                F_UNIQUE_KEY: "douyin_O1",
                F_ORDER_NO: "O1",
                F_QUANTITY: 0,
                F_PAID_AMOUNT: 0,
                F_REFUND_AMOUNT: 0,
                F_ACCESSORY_FLAG: "是",
            },
        ]
    )

    assert rows[0][F_ACCESSORY_FLAG] == "否"


def test_collapsed_order_marks_accessory_only_order_as_accessory():
    rows = collapse_order_rows(
        [
            {
                F_UNIQUE_KEY: "douyin_O1",
                F_ORDER_NO: "O1",
                F_QUANTITY: 0,
                F_PAID_AMOUNT: 0,
                F_REFUND_AMOUNT: 0,
                F_ACCESSORY_FLAG: "是",
            },
            {
                F_UNIQUE_KEY: "douyin_O1",
                F_ORDER_NO: "O1",
                F_QUANTITY: 0,
                F_PAID_AMOUNT: 0,
                F_REFUND_AMOUNT: 0,
                F_ACCESSORY_FLAG: "是",
            },
        ]
    )

    assert rows[0][F_ACCESSORY_FLAG] == "是"


def test_parse_ad_rows_aggregates_by_platform_date(tmp_path: Path):
    path = tmp_path / "天猫投流.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["日期", "花费", "展现量", "点击量", "总成交金额"])
    sheet.append(["2026-06-07 00:00:00", "100", "1000", "50", "300"])
    sheet.append(["2026-06-07 00:00:00", "20", "200", "10", "80"])
    workbook.save(path)

    rows = parse_ad_rows("天猫", path)

    assert len(rows) == 1
    assert rows[0][F_UNIQUE_KEY] == ad_unique_key("天猫", "2026-06-07")
    assert rows[0][F_PLATFORM] == "天猫"
    assert rows[0][F_DATE] == "2026-06-07"
    assert rows[0]["实际消耗"] == 120
    assert rows[0]["成交金额"] == 380


def test_parse_pdd_ad_rows_preserves_platform_specific_metrics(tmp_path: Path):
    path = tmp_path / "商品推广_账户_分天数据.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "日期",
            "成交花费(元)",
            "交易额(元)",
            "实际投产比",
            "总花费(元)",
            "净交易额(元)",
            "净实际投产比",
            "净成交笔数",
            "结算交易额(元)",
            "成交笔数",
            "每笔成交金额(元)",
            "曝光量",
            "点击量",
        ]
    )
    sheet.append(["2026-06-07", 1426.86, 5407.68, 3.79, 1426.86, 5069.7, 3.55, 30, 5069.7, 32, 168.99, 26765, 1028])
    workbook.save(path)

    rows = parse_ad_rows("拼多多", path)

    assert len(rows) == 1
    row = rows[0]
    assert row[F_UNIQUE_KEY] == ad_unique_key("拼多多", "2026-06-07")
    assert row[F_DEAL_SPEND] == 1426.86
    assert row[F_TOTAL_SPEND] == 1426.86
    assert row[F_TRADE_AMOUNT] == 5407.68
    assert row["成交金额"] == 5407.68
    assert row[F_NET_TRADE_AMOUNT] == 5069.7
    assert row[F_NET_ACTUAL_ROI] == 3.55
    assert row[F_NET_DEAL_COUNT] == 30
    assert row[F_SETTLED_TRADE_AMOUNT] == 5069.7
    assert row[F_DEAL_COUNT] == 32
    assert row[F_AMOUNT_PER_DEAL] == 168.99


def test_run_import_can_filter_pdd_ads_to_requested_dates(tmp_path: Path):
    batch = tmp_path / "0609"
    pdd_dir = batch / "拼多多"
    pdd_dir.mkdir(parents=True)
    path = pdd_dir / "商品推广_账户_分天数据.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["日期", "成交花费(元)", "交易额(元)", "总花费(元)", "成交笔数", "曝光量", "点击量"])
    sheet.append(["2026-06-06", 100, 300, 100, 2, 1000, 50])
    sheet.append(["2026-06-07", 1426.86, 5407.68, 1426.86, 32, 26765, 1028])
    sheet.append(["2026-06-08", 568.23, 2027.88, 568.23, 12, 5861, 253])
    workbook.save(path)

    evidence = tmp_path / "evidence.json"
    summary = run_import(
        batch_dir=batch,
        dry_run=True,
        evidence=evidence,
        platforms={"拼多多"},
        kinds={"ads"},
        dates={"2026-06-07", "2026-06-08"},
    )

    assert summary["ad_count"] == 2
    assert summary["ad_dates"] == ["2026-06-07", "2026-06-08"]
    assert summary["date_filter"] == ["2026-06-07", "2026-06-08"]
    assert summary["files"]["拼多多"]["ads"][0]["rows"] == 2
    assert [row[F_DATE] for row in summary["sample_ad_rows"]] == ["2026-06-07", "2026-06-08"]
    assert summary["sample_ad_rows"][0][F_DEAL_SPEND] == 1426.86


def test_upsert_rows_does_not_create_or_require_nonexistent_optional_fields():
    calls: list[tuple[str, str, dict | None, dict | None]] = []

    class FakeClient(FeishuDailyClient):
        def __init__(self) -> None:
            self.app_token = "app"

        def field_names(self, table_id: str) -> set[str]:
            return {"unique_key", "订单号", "平台"}

        def iter_records(self, table_id: str, field_names=None):
            yield {"record_id": "rec1", "fields": {"unique_key": "tmall_T1", "订单号": "T1"}}

        def request(self, method: str, path: str, payload=None, params=None):
            calls.append((method, path, payload, params))
            return {}

    result = FakeClient().upsert_rows(
        table_id="tbl",
        rows=[{"unique_key": "tmall_T1", "订单号": "T1", "平台": "天猫", "不存在字段": "drop"}],
        required_fields=["unique_key", "订单号"],
        fallback_match_fields=("订单号",),
    )

    assert result["updated"] == 1
    assert result["created"] == 0
    assert result["dropped_nonexistent_fields"] == {"不存在字段": 1}
    assert calls[0][0] == "POST"
    assert "/fields" not in calls[0][1]
    assert calls[0][2]["records"][0]["fields"] == {"平台": "天猫"}


def test_ensure_missing_fields_for_rows_creates_only_typed_nonempty_fields():
    calls: list[tuple[str, str, dict | None, dict | None]] = []

    class FakeClient(FeishuDailyClient):
        def __init__(self) -> None:
            self.app_token = "app"

        def field_names(self, table_id: str) -> set[str]:
            return {"unique_key", "平台", "投放日期"}

        def request(self, method: str, path: str, payload=None, params=None):
            calls.append((method, path, payload, params))
            return {}

    created = FakeClient().ensure_missing_fields_for_rows(
        "tbl",
        [
            {"unique_key": "ads_pdd_2026-06-07", "成交花费(元)": 1426.86, "净交易额(元)": 5069.7, "不支持字段": 1},
            {"unique_key": "ads_pdd_2026-06-08", "成交花费(元)": 568.23, "结算交易额(元)": None},
        ],
        AD_FIELD_TYPES,
    )

    assert created == ["净交易额(元)", "成交花费(元)"]
    assert [call[2] for call in calls] == [
        {"field_name": "净交易额(元)", "type": 2},
        {"field_name": "成交花费(元)", "type": 2},
    ]


def test_existing_order_update_can_be_limited_to_trade_status():
    calls: list[tuple[str, str, dict | None, dict | None]] = []

    class FakeClient(FeishuDailyClient):
        def __init__(self) -> None:
            self.app_token = "app"

        def field_names(self, table_id: str) -> set[str]:
            return {"unique_key", "订单号", "交易状态", "实收款"}

        def iter_records(self, table_id: str, field_names=None):
            yield {"record_id": "rec1", "fields": {"unique_key": "tmall_T1", "订单号": "T1"}}

        def request(self, method: str, path: str, payload=None, params=None):
            calls.append((method, path, payload, params))
            return {}

    result = FakeClient().upsert_rows(
        table_id="tbl",
        rows=[{"unique_key": "tmall_T1", "订单号": "T1", "交易状态": "卖家已发货", "实收款": 169}],
        required_fields=["unique_key", "订单号"],
        fallback_match_fields=("订单号",),
        update_existing_fields={"交易状态"},
    )

    assert result["updated"] == 1
    assert result["created"] == 0
    assert calls[0][2]["records"][0]["fields"] == {"交易状态": "卖家已发货"}


def test_existing_order_update_sends_only_changed_fields():
    calls: list[tuple[str, str, dict | None, dict | None]] = []

    class FakeClient(FeishuDailyClient):
        def __init__(self) -> None:
            self.app_token = "app"

        def field_names(self, table_id: str) -> set[str]:
            return {F_UNIQUE_KEY, F_ORDER_NO, F_TRADE_STATUS, F_PAID_AMOUNT}

        def iter_records(self, table_id: str, field_names=None):
            yield {
                "record_id": "rec1",
                "fields": {
                    F_UNIQUE_KEY: "tmall_T1",
                    F_ORDER_NO: "T1",
                    F_TRADE_STATUS: "交易成功",
                    F_PAID_AMOUNT: 169,
                },
            }

        def request(self, method: str, path: str, payload=None, params=None):
            calls.append((method, path, payload, params))
            return {}

    result = FakeClient().upsert_rows(
        table_id="tbl",
        rows=[{F_UNIQUE_KEY: "tmall_T1", F_ORDER_NO: "T1", F_TRADE_STATUS: "交易成功", F_PAID_AMOUNT: 0}],
        required_fields=[F_UNIQUE_KEY, F_ORDER_NO],
        fallback_match_fields=(F_ORDER_NO,),
        update_existing_fields={F_TRADE_STATUS, F_PAID_AMOUNT},
    )

    assert result["updated"] == 1
    assert result["created"] == 0
    assert calls[0][2]["records"][0]["fields"] == {F_PAID_AMOUNT: 0}


def test_deduplicate_records_deletes_only_repeated_platform_order_keys():
    calls: list[tuple[str, str, dict | None, dict | None]] = []

    class FakeClient(FeishuDailyClient):
        def __init__(self) -> None:
            self.app_token = "app"

        def field_names(self, table_id: str) -> set[str]:
            return {"平台", "订单号"}

        def iter_records(self, table_id: str, field_names=None):
            yield {"record_id": "keep1", "fields": {"平台": "视频号", "订单号": "O1"}}
            yield {"record_id": "drop1", "fields": {"平台": "视频号", "订单号": "O1"}}
            yield {"record_id": "keep2", "fields": {"平台": "抖音", "订单号": "O1"}}
            yield {"record_id": "keep3", "fields": {"平台": "视频号", "订单号": "O2"}}

        def request(self, method: str, path: str, payload=None, params=None):
            calls.append((method, path, payload, params))
            return {}

    result = FakeClient().deduplicate_records("tbl", ("平台", "订单号"))

    assert result["deleted_duplicate_records"] == 1
    assert result["duplicate_keys"] == 1
    assert calls == [
        (
            "POST",
            "/bitable/v1/apps/app/tables/tbl/records/batch_delete",
            {"records": ["drop1"]},
            None,
        )
    ]


def test_canonicalize_ad_unique_keys_updates_legacy_douyin_keys_and_deletes_duplicates():
    calls: list[tuple[str, str, dict | None, dict | None]] = []

    class FakeClient(FeishuDailyClient):
        def __init__(self) -> None:
            self.app_token = "app"

        def field_names(self, table_id: str) -> set[str]:
            return {"unique_key", "平台", "投放日期"}

        def iter_records(self, table_id: str, field_names=None):
            yield {"record_id": "rec_old", "fields": {"unique_key": "douyin_ads_3810515145814311090_2026-03-25", "平台": "抖音", "投放日期": "2026-03-25"}}
            yield {"record_id": "rec_dup", "fields": {"unique_key": "ads_douyin_2026-03-25", "平台": "抖音", "投放日期": "2026-03-25"}}
            yield {"record_id": "rec_ok", "fields": {"unique_key": "ads_pdd_2026-03-25", "平台": "拼多多", "投放日期": "2026-03-25"}}

        def request(self, method: str, path: str, payload=None, params=None):
            calls.append((method, path, payload, params))
            return {}

    result = FakeClient().canonicalize_ad_unique_keys("tbl")

    assert result["updated"] == 1
    assert result["deleted_duplicate_records"] == 1
    assert calls[0][1].endswith("/records/batch_delete")
    assert calls[0][2] == {"records": ["rec_dup"]}
    assert calls[1][1].endswith("/records/batch_update")
    assert calls[1][2] == {
        "records": [{"record_id": "rec_old", "fields": {"unique_key": "ads_douyin_2026-03-25", "平台": "抖音"}}]
    }


def test_classify_file_prefers_headers_over_filename(tmp_path: Path):
    order_path = tmp_path / "random-name.csv"
    with order_path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["订单号", "订单状态", "商品"])
        writer.writerow(["P1", "已发货", "商品A"])

    ad_path = tmp_path / "another-random.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["日期", "花费", "点击量", "展现量"])
    sheet.append(["2026-06-09", 100, 20, 1000])
    workbook.save(ad_path)

    assert classify_file("拼多多", order_path) == "orders"
    assert classify_file("天猫", ad_path) == "ads"


def test_douyin_order_export_never_generates_influencer_rows(tmp_path: Path):
    path = tmp_path / "orders.csv"
    headers = ["主订单编号", "订单状态", "选购商品", "达人昵称", "达人实际承担优惠金额"]
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        writer.writerow({"主订单编号": "D1", "订单状态": "订单付款", "选购商品": "商品A", "达人昵称": "达人A", "达人实际承担优惠金额": "3.5"})

    assert classify_file("抖音", path) == "orders"
    assert parse_influencer_rows("抖音", path) == []


def test_order_rows_can_be_enriched_with_product_breakdown_fields():
    rules = product_rules_from_records([{"fields": {"商品名称": "洗面奶", "搜索关键词": "洗面奶"}}])
    rows = [
        {
            "商品名称": "趣白全自动洗面奶打泡机",
            "数量": 2,
            "实收款": 338,
            "退款金额": 0,
        }
    ]

    add_product_breakdown_to_orders(rows, rules)

    assert rows[0]["洗面奶数量"] == 2
    assert rows[0]["洗面奶有效销售额"] == 338


def test_order_breakdown_uses_internal_source_quantity_before_accessory_zeroing():
    rules = product_rules_from_records([{"fields": {"商品名称": "配件", "搜索关键词": "配件"}}])
    rows = [
        {
            "商品名称": "洁面乳打泡机配件",
            "数量": 0,
            INTERNAL_PRODUCT_BREAKDOWN_QUANTITY: 8,
            "实收款": 14,
            "退款金额": 0,
        }
    ]

    add_product_breakdown_to_orders(rows, rules)

    assert rows[0]["配件数量"] == 8
    assert rows[0]["配件有效销售额"] == 14
    assert INTERNAL_PRODUCT_BREAKDOWN_QUANTITY not in rows[0]


def test_douyin_influencer_rows_require_commission_excel(tmp_path: Path):
    path = tmp_path / "48a95f74-af9d-088a-e36e-81eab8857874_3824661741573898566.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["订单id", "商品id", "商品名称", "作者账号", "抖音/火山号", "支付金额", "佣金率", "预估佣金支出", "结算金额", "实际佣金支出", "订单状态", "付款时间", "店铺名称", "商品数量", "订单来源", "流量来源"])
    sheet.append(["6953547102608758621\t", "3810515145814311090\t", "商品A", "作者A", "66799442664", "169.00", "32.00%", "54.08", "0.00", "0.00", "订单付款", "2026-06-09 21:25:25", "趣白", "1", "抖音", "视频"])
    workbook.save(path)

    rows = parse_influencer_rows("抖音", path)

    assert classify_file("抖音", path) == "influencer"
    assert rows[0]["unique_key"] == "douyin_6953547102608758621"
    assert rows[0]["订单号"] == "6953547102608758621"
    assert rows[0]["作者账号"] == "作者A"
    assert rows[0]["预估佣金支出"] == 54.08
    assert rows[0]["带货达人ID"] == "66799442664"
    assert rows[0]["带货达人昵称"] == "作者A"
    assert rows[0]["带货佣金率"] == "32.0"
    assert rows[0]["带货费用"] == 54.08
    assert rows[0]["带货费用口径"] == "预估佣金支出"
    assert rows[0]["流量来源"] == "视频"


def test_wechat_influencer_rows_fill_estimated_commission_from_cost(tmp_path: Path):
    path = tmp_path / "微信小店订单.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["订单号", "订单下单时间", "支付时间", "订单状态", "商品名称", "商品编码(平台)", "商品数量", "订单实际支付金额", "带货账号昵称", "带货费用", "带货佣金率", "带货费用类型"])
    sheet.append(["W1", "2026-06-08 12:00:00", "2026-06-08 12:01:00", "已付款", "商品B", "P1", 1, 169, "达人B", 8.8, "5%", "佣金"])
    workbook.save(path)

    rows = parse_influencer_rows("视频号", path)

    assert rows[0]["unique_key"] == "视频号W1"
    assert rows[0]["带货费用"] == 8.8
    assert rows[0]["预估佣金支出"] == 8.8
    assert rows[0]["实际佣金支出"] is None
    assert rows[0]["佣金率"] == 5


def test_unique_key_rules_are_stable():
    assert order_unique_key("抖音", "\t6926972861706959874") == "douyin_6926972861706959874"
    assert ad_unique_key("拼多多", "2026-06-07") == "ads_pdd_2026-06-07"


def test_feishu_data_not_ready_response_is_retryable():
    assert is_retryable_feishu_response(400, {"code": 1254607})
    assert is_retryable_feishu_response(429, {"code": 99991663})
    assert not is_retryable_feishu_response(400, {"code": 1254000})
